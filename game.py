from datetime import datetime, time
from flask import (
    Flask,
    Response,
    abort,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    send_from_directory,
    session,
    url_for,
    flash,
)
import hashlib
import json
import mimetypes
import os
import random
import secrets
import string
from pathlib import Path

from sqlalchemy import desc, func, inspect
from sqlalchemy.types import JSON
from werkzeug.utils import secure_filename
from flask_login import (
    UserMixin,
    login_user,
    LoginManager,
    current_user,
    logout_user,
    login_required,
)
from flask_wtf.csrf import CSRFProtect, CSRFError
from dotenv import load_dotenv

from extensions import db, socketio
from number_voice import number_to_audio
from flask_socketio import emit, join_room




load_dotenv()
app = Flask(__name__, template_folder="static/")
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get("DB_PATH")
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY") #убрать в переменные среды
if app.config['SECRET_KEY'] is None:
    raise ValueError("ОШИБКА: Переменная окружения SECRET_KEY не установлена!")
app.secret_key = app.config["SECRET_KEY"] #os.urandom(32).hex

# CSRF-защита обычных HTTP-запросов. Проверка запускается вручную ниже,
# чтобы служебный транспорт Socket.IO (/socket.io) не блокировался.
app.config["WTF_CSRF_CHECK_DEFAULT"] = False
app.config["WTF_CSRF_TIME_LIMIT"] = None

csrf = CSRFProtect(app)
BASE_DIR = Path(__file__).resolve().parent
BRANDING_FILE = BASE_DIR / "config" / "branding.json"


with open(
    BASE_DIR / "config" / "branding.json",
    encoding="utf-8"
) as f:
    BRAND = json.load(f)
    
@app.context_processor
def inject_brand():

    return BRAND

def load_branding() -> dict:
    """Загружает настройки бренда из JSON."""
    default_branding = {
        "game_name": "The People Versus",
        "game_short_name": "TPV",
        "game_subtitle": "LIVE GAME MASTER CONSOLE",
    }

    if not BRANDING_FILE.exists():
        return default_branding

    try:
        with BRANDING_FILE.open("r", encoding="utf-8") as file:
            loaded = json.load(file)

        # Значения из JSON дополняют настройки по умолчанию.
        return {**default_branding, **loaded}

    except (OSError, json.JSONDecodeError) as error:
        app.logger.error("Не удалось загрузить branding.json: %s", error)
        return default_branding




@app.before_request
def protect_http_requests_from_csrf():
    """Проверяет CSRF для изменяющих HTTP-запросов, кроме транспорта Socket.IO."""
    if request.path.startswith("/socket.io"):
        return None

    if request.method in {"POST", "PUT", "PATCH", "DELETE"}:
        csrf.protect()

    return None


@app.errorhandler(CSRFError)
def handle_csrf_error(error):
    """Возвращает JSON для fetch и понятную ошибку для обычных форм."""
    if request.is_json or request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return jsonify({
            "status": "error",
            "error": "csrf_failed",
            "message": "Сессия устарела или CSRF-токен недействителен. Обновите страницу."
        }), 400

    return render_template("csrf_error.html", reason=error.description), 400

socketio.init_app(app, cors_allowed_origins=os.environ.get("ALLOWED_ORIGINS")) # добавить конкретный домен
accepted_user = ""
db.init_app(app)
with app.app_context():
    print("DB URI:", app.config["SQLALCHEMY_DATABASE_URI"])
    print("ENGINE URL:", db.engine.url)
    print("DATABASE FILE:", db.engine.url.database)

    inspector = inspect(db.engine)
    print("TABLES:", inspector.get_table_names())
login_manager = LoginManager(app)
login_manager.session_protection = "strong"
login_manager.login_view = "login"
login_manager.login_message_category = "info"
app.config['TELEGRAM_BOT_TOKEN'] = ''
DEFAULT_ROOM_CODE = os.environ.get("DEFAULT_ROOM_CODE") #убрать в переменные среды
HOST_USERNAME = os.environ.get("HOST_USERNAME") #убрать в переменные среды

with app.app_context():
    inspector = inspect(db.engine)
    print(inspector.get_table_names())


@app.route("/error-test/<int:code>") 
def error_test_code(code):
    abort(code)

@socketio.on("connect")
def on_connect():
    print("Client connected")


@socketio.on("disconnect")
def on_disconnect():
    print("Client disconnected")


@socketio.on("ping:test")
def ping_test(data):
    print("Ping from client:", data)

    emit("pong:test", {
        "message": "Socket.IO работает"
    })
    
@socketio.on("room:join_slot")
def socket_join_room(data):
    room_code = str(data.get("room") or get_room_code() or "")
    role = data.get("role") or "unknown"
    username = data.get("username") or ""

    join_room(room_code)
    join_room(f"{room_code}:{role}")
    if username:
        join_room(f"{room_code}:user:{username}")

    print(f"Socket joined room={room_code}, role={role}, username={username}")
    update_list_users()
    emit("room:joined", {
        "room": room_code,
        "role": role,
        "username": username
    })   
    
    
@socketio.on("room:join_tpv")
def socket_join_room(data):
    room_code = str(data.get("room") or get_room_code() or "")
    role = data.get("role") or "unknown"
    username = data.get("username") or ""

    join_room(room_code)
    join_room(f"{room_code}:{role}")
    if username:
        join_room(f"{room_code}:user:{username}")

    print(f"Socket joined room={room_code}, role={role}, username={username}")
    update_users_tpv()
    emit("room:joined", {
        "room": room_code,
        "role": role,
        "username": username
    }) 

@socketio.on("count_answer_interactive")
def count_interactive(data):
    try:
        col = int(data["interactive"])
        socketio.emit("count_answer_interactive_for_spec",col,to=f"{DEFAULT_ROOM_CODE}:spectator")
    except:
        return

@login_manager.user_loader
def load_user(user_id):
    return db.session.get(Users, user_id)

class Room(db.Model):
    id = db.Column(db.Integer, primary_key=True, unique=True)
    game = db.Column(db.Text, unique=True)
    def __repr__(self):
        return '<Room %r>' %self.id


class Users(db.Model, UserMixin):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True, unique=True)
    username = db.Column(db.String(10), unique=True, nullable=False)
    answer = db.Column(db.Text)
    money = db.Column(db.Integer)
    time = db.Column(db.Text)
    status = db.Column(db.Text)
    main_money = db.Column(db.Integer)
    red_bomb = db.Column(db.Text)
    def __repr__(self):
        return '<Users %r>' %self.id
    
class Helps(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True, unique=True)
    h50_50 = db.Column(db.Text)
    alter = db.Column(db.Text)
    navi = db.Column(db.Text)
    x2 = db.Column(db.Text)
    auden = db.Column(db.Text)
    fact = db.Column(db.Text)
    def __repr__(self):
        return '<Helps %r>' %self.id

class Facts(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True, unique=True)
    slot = db.Column(db.Text, nullable=False)
    des_fact = db.Column(db.Text)
    def __repr__(self):
        return '<Facts %r>' %self.id

class Task(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True, unique=True)
    round = db.Column(db.Integer)
    fatal = db.Column(JSON)
    md5 = db.Column(db.Text)
    count_fatal = db.Column(db.Integer)
    b_bomb = db.Column(db.Integer)
    r_bomb = db.Column(db.Integer)
    def __repr__(self):
        return '<Task %r>' %self.id

class Answered(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True, unique=True)
    answer = db.Column(db.Text)
    round = db.Column(db.Integer)
    def __repr__(self):
        return '<Answered %r>' %self.id


def get_current_room():
    """Возвращает текущую игровую комнату или None."""
    return db.session.scalar(db.select(Room).limit(1))


def get_room_code():
    """Возвращает код текущей комнаты строкой."""
    room = get_current_room()
    return str(room.id) if room is not None else None


# ---------------------------------------------------------------------------
# TPV: централизованная отправка Socket.IO-событий
# ---------------------------------------------------------------------------

def emit_tpv_host(event, data=None):
    """Отправляет событие пульту ведущего TPV."""
    socketio.emit(event, data, to=f"{DEFAULT_ROOM_CODE}:host")


def emit_tpv_spectator(event, data=None):
    """Отправляет событие техническому экрану зрителя TPV."""
    socketio.emit(event, data, to=f"{DEFAULT_ROOM_CODE}:spectator")


def emit_tpv_players(event, data=None):
    """Отправляет событие всем игрокам текущей игровой комнаты TPV."""
    room_code = get_room_code()
    if room_code is not None:
        socketio.emit(event, data, to=room_code)


def emit_tpv_player(username, event, data=None):
    """Отправляет событие одному конкретному игроку TPV."""
    room_code = get_room_code()
    if room_code is not None and username:
        socketio.emit(event, data, to=f"{room_code}:user:{username}")


def init_game():
    db.session.execute(db.delete(Task))
    db.session.execute(db.delete(Answered))
    db.session.execute(db.delete(Helps))
    #if os.path.exists("answered.json"):
    #    os.remove("answered.json")
   # if os.path.exists("task.json"):
  #      os.remove("task.json")
    if os.path.exists("task_otbor.json"):
        os.remove("task_otbor.json")
   # if os.path.exists("helps.json"):
   #     os.remove("helps.json")
   # if os.path.exists("50_50.json"):
   #     os.remove("50_50.json")  
  #  if os.path.exists("alter.json"):
  #      os.remove("alter.json")
  #  if os.path.exists("navi.json"):
   #     os.remove("navi.json") 
  #  if os.path.exists("auden.json"):
    #    os.remove("auden.json") 
  #  if os.path.exists("fact.json"):
  #      os.remove("fact.json") 
  #  if os.path.exists("50_50_spec.json"):
   #     os.remove("50_50_spec.json")  
  #  if os.path.exists("alter_spec.json"):
  #      os.remove("alter_spec.json")
  #  if os.path.exists("navi_spec.json"):
   #     os.remove("navi_spec.json") 
 #   if os.path.exists("auden_spec.json"):
  #      os.remove("auden_spec.json") 
 #   if os.path.exists("fact_spec.json"):
  #      os.remove("fact_spec.json") 


@app.route('/')
def index(): 
    #print (url_for('join'))
    return render_template("index.html")


def check_id_room(room_id):
   # if not os.path.exists("room.json"):
    room = db.session.scalar(db.select(Room).where(Room.id == int(room_id)))
    if room == None:
            return False
    else:
        return True


def give_name_game(game_name):
    name = db.session.scalar(db.select(Room).where(Room.id==game_name))
    return name.game



@app.route('/join', methods=["POST", "GET"])
def join():
    if request.method == 'POST':
        if 'userLogged' in session:
            pass
        print(request.form)
        if request.form['user_name']== "":
           flash ('Требуется авторизация')
           return render_template("login.html")
        if  (request.form['user_name']!=HOST_USERNAME) and (request.form['room_id']==DEFAULT_ROOM_CODE):
           flash ('Неверный код комнаты')
           return render_template("login.html")
        if (request.form['user_name']==HOST_USERNAME)  and (request.form['room_id']==DEFAULT_ROOM_CODE):
           # _users[0] = request.form['user_name']
            init_game()
            return render_template("select.html")
        else:
            if check_id_room(request.form['room_id'])==False:
                flash ('Неверный код комнаты')
                return render_template("login.html")
            if give_name_game(request.form['room_id']) == 'slot':
                u = Users()
                u.username = request.form['user_name']
                u.answer = "0"
                u.money = 0
                u.time = datetime.now()
                u.status = "wait"
                u.main_money = 0
                u.red_bomb = "false"
                tmp = db.session.scalar(db.select(Users).where(Users.username==u.username))
                if tmp!=None:
                    if tmp.username == u.username:
                   # if tmp.username in session['username']:
                            print (url_for('join'))
                            return render_template("user_slot.html",value=u.username)

                            
                    #else:
                     #   return render_template("login.html")                       
                db.session.add(u)
                db.session.flush()
                db.session.commit()
                session['username'] = u.username
                ch = login_user(u)
                return render_template("user_slot.html",value=u.username)
            if give_name_game(request.form['room_id']) == 'tpv':
                find_user = db.session.scalar(db.select(UsersTpv).where(UsersTpv.username==request.form['user_name']))
                if find_user == None:
                    flash ('К сожалению, Ваша заявку на игру не найдена.')
                    return render_template("login.html")
                if find_user.flip=="false" or find_user.flip==None:
                    flash ('К сожалению, Ваша заявку на игру не одобрена! Отстуствует тема замены.')
                    return render_template("login.html")  
                if int(find_user.flip_col) < TPV_REQUIRED_FLIP_QUESTIONS:
                    flash ('К сожалению, Ваша заявку на игру не одобрена! Недостаточно вопросов замены.')
                    return render_template("login.html")  
                user_tpv = QueryTpv()
                user_tpv.username = request.form['user_name']
                tmp = db.session.scalar(db.select(QueryTpv).where(QueryTpv.username==user_tpv.username))
                if tmp!=None:
                    if tmp.username == user_tpv.username:
                   # if tmp.username in session['username']:
                            print (url_for('join'))
                            return render_template("tpv-user.html",value=user_tpv.username)
                            #return render_template("user_slot.html",value=u.username)
                user_tpv.money = find_user.money
                #user_tpv.money = 0
                user_tpv.flip = find_user.flip
                #user_tpv.flip = "test"
                #user_tpv.status = "wait"
                db.session.add(user_tpv)
                db.session.flush()
                db.session.commit()
                update_users_tpv()
                session['username'] = user_tpv.username 
                ch = login_user(user_tpv)
                return render_template("tpv-user.html",value=user_tpv.username)
                    
    return render_template("login.html")

@app.route('/select', methods=["POST", "GET"])
def select():
    if request.method == 'POST':
        if request.form.values == "Свободный слот":
         print (url_for('slot'))
         return render_template("slot.html")
    abort(403)
    print (url_for('join'))
    return render_template("login.html")

@app.route('/slot', methods=["POST", "GET"])
def slot():
    if request.method == 'POST':
        print (url_for('slot'))
        return render_template("slot.html")
    abort(403)
    print (url_for('join'))
    return render_template("login.html")

@app.route('/tpv', methods=["POST", "GET"])
def tpv():
    # Лаунчер TPV открывается только после POST из общего выбора игр.
    # Повторный GET разрешён только после успешного входа в TPV-лаунчер.
    if request.method == "POST":
        session["tpv_launcher_allowed"] = True
        session.pop("tpv_role", None)
        return render_template("tpv.html")

    if session.get("tpv_launcher_allowed") is True:
        return render_template("tpv.html")

    abort(403)


@app.route('/tpv_host', methods=["POST", "GET"])
def tpv_host():
    # Режим ведущего запускается только кнопкой из tpv.html.
    if request.method == "POST":
        if session.get("tpv_launcher_allowed") is not True:
            return redirect(url_for("tpv"))

        session["tpv_role"] = "host"
        return render_template("tpv-host.html")

    # Прямой ввод /tpv_host не открывает пульт.
    return redirect(url_for("tpv"))


@app.route('/tpv_spectator', methods=["POST", "GET"])
def tpv_spectator():
    # Экран зрителя запускается только кнопкой из tpv.html.
    if request.method == "POST":
        if session.get("tpv_launcher_allowed") is not True:
            return redirect(url_for("tpv"))

        session["tpv_role"] = "spectator"
        return render_template("tpv-spectator.html")

    # Прямой ввод /tpv_spectator не открывает экран зрителя.
    return redirect(url_for("tpv"))

@app.route('/tpv_user', methods=["GET"])
def tpv_user():
    if request.method == 'GET':
        abort(401)



@app.route('/user_slot', methods=["GET"])
def user_slot():
    if request.method == 'GET':
        abort(401)

@app.route('/spec_slot', methods=["POST", "GET"])
def spec_slot():
    if request.method == 'POST':
        print (url_for('slot'))
        return render_template("spec_slot.html")
    abort(400)
    print (url_for('join'))
    return render_template("login.html")




@app.route('/host_slot', methods=["POST", "GET"])
def host_slot():
    if request.method == 'POST':
        print (url_for('host_slot'))
        #for i in _users:
          #  flash (i)
        return render_template("host_slot.html")
    abort(403)
    print (url_for('join'))
    return render_template("login.html")

@app.route('/invite_user', methods=["POST", "GET"])
def invite_user():
    if request.method == 'POST':
        try:
            u = request.json['user_name']
            tmp = db.session.scalar(db.select(Users).where(Users.id==int(u)))
            u = str(tmp.username)
            if tmp.red_bomb == 'true':
                return json.dumps("red bomb")
        except:
            return json.dumps("fail")
        if tmp == None:
            return json.dumps("fail")
        tmp.status = 'main'
        tmp.answer = '0'
        tmp.time = 0
	#tmp.time = 0
        db.session.commit()
        socketio.emit("updated_status_user",tmp.status,to=f"{get_room_code()}:user:{tmp.username}");
        js = db.session.scalars(db.select(Users)).all()
        if len(js)!=1:
            for i in range(len(js)):
                if js[i].status !="main":
                    js[i].status = 'interactive'
                    js[i].answer = '0'
                    js[i].time = 0
                    socketio.emit("updated_status_user",js[i].status,to=f"{get_room_code()}:user:{js[i].username}");
            db.session.commit()
        update_list_users()    
        return json.dumps(u)
    print (url_for('host_slot'))
    return render_template("host_slot.html")




@app.route('/gen_task', methods=["POST", "GET"])
def gen_task():
    if request.method == 'POST':
        r = request.json["current_round"]
        tmp_bomb = request.json["bombs"]
        bomb = False
        if tmp_bomb == "true":
            bomb = True
        jsn = generate_string(int(r),bomb)
        if jsn == "null":
            return json.dumps("fail")
        get_task_user()
        return jsn
  #  print (url_for('host_slot'))
  #  return render_template("host_slot.html")


def get_md5_hash(stroka):
    characters = string.ascii_letters + string.punctuation
    random_string ="".join(secrets.choice(characters) for _ in range(12))
    result_str = str(stroka)+'_'+random_string
    return hashlib.md5(random_string.encode()).hexdigest()
    
    
def generate_string(round_id,is_bombed):
    secure_rnd = secrets.SystemRandom()
    current_round = int(round_id); #получить номер раунда, 0 - отборочный тур
    count_fatal = 0
    otbor_chislo = 0
    bomb = is_bombed
    if current_round == 0:
        otbor_chislo = secure_rnd.randint(10,999)
        a = secure_rnd.randint(10,otbor_chislo)
        b = secure_rnd.randint(otbor_chislo,999)
        md5_hash = get_md5_hash(otbor_chislo)
        result = [current_round,a,b,otbor_chislo,md5_hash]
        result_send = [current_round,a,b,None,md5_hash]
        socketio.emit("get_task_otbor",result_send,to=f"{DEFAULT_ROOM_CODE}:spectator")
        socketio.emit("get_task_otbor",result_send,to=f"{get_room_code()}:user")
        js = json.dumps(result)
        with open('task_otbor.json', 'w') as file:
            json.dump(result, file)
        return js
    else:
        match current_round:
            case 1: count_fatal=1
            case 2: count_fatal=2
            case 3: count_fatal=3
            case 4: count_fatal=5
            case 5: count_fatal=6
            case 6: count_fatal=8
            case 7: count_fatal=10
            case 8: count_fatal=12
            case 9: count_fatal=14
    
    if count_fatal == 1:
        fatal = secure_rnd.randint(1,15)
        #md5_hash = hashlib.md5(str(fatal).encode()).hexdigest()
        md5_hash = get_md5_hash(fatal)
        result = [current_round,fatal,md5_hash, count_fatal]
        result_send = [current_round,None,md5_hash, count_fatal]
        socketio.emit("get_task",result_send,to=f"{DEFAULT_ROOM_CODE}:spectator")
        socketio.emit("get_task",result_send,to=f"{get_room_code()}")
        db.session.execute(db.delete(Task))
        task = Task()
        task.round = current_round
        task.fatal = fatal
        task.md5 = md5_hash
        task.count_fatal = count_fatal
        db.session.add(task)
        db.session.flush()
        db.session.commit()
        js = json.dumps(result)
      #  with open('task.json','w') as file:
       #     json.dump(result,file)
        return js
    else:
        fatal = secure_rnd.sample([1,2,3,4,5,6,7,8,9,10,11,12,13,14,15], count_fatal)
        b_bomb = 'false'
        r_bomb = 'false'
        if bomb and (current_round >= 4):
            tmp_bombs = secure_rnd.sample(fatal,2)
            b_bomb = tmp_bombs[0]
            r_bomb = tmp_bombs[1]
        #md5_hash = hashlib.md5(str(fatal).encode()).hexdigest()
        md5_hash = get_md5_hash(fatal)
        result = [current_round,fatal,md5_hash,count_fatal,b_bomb,r_bomb]
        result_send = [current_round,None,md5_hash,count_fatal,None,None]
        socketio.emit("get_task",result_send,to=f"{DEFAULT_ROOM_CODE}:spectator")
        socketio.emit("get_task",result_send,to=f"{get_room_code()}")
        db.session.execute(db.delete(Task))
        task = Task()
        task.round = current_round
        task.fatal = json.dumps(fatal)
        task.md5 = md5_hash
        task.count_fatal = count_fatal
        if b_bomb != "false":
            task.b_bomb = b_bomb
        if r_bomb != "false":
            task.r_bomb = r_bomb
        db.session.add(task)
        db.session.flush()
        db.session.commit()
        js = json.dumps(result)
       # with open('task.json','w') as file:
           # json.dump(result,file)
        return js


            
@app.route('/get_fatal_host', methods=["POST", "GET"])
def get_fatal_host():
    if request.method == 'POST':
        r = request.json["answer"]
        r1 = request.json["round"]
        res = [r, r1]
       # db.session.execute(db.delete(Answered))
        answered = Answered()
        answered.answer = r
        answered.round = r1
        db.session.add(answered)
        db.session.flush()
        db.session.commit()
        #with open('answered.json', 'w') as file:
        #    json.dump(res, file)
        jsn = []
        task = db.session.scalar(db.select(Task).limit(1))
        jsn.append(task.round)
        if task.round == 1:
            jsn.append(task.fatal)
        else:
            jsn.append(json.loads(task.fatal))
        jsn.append(task.md5)
        jsn.append(task.count_fatal)
        if task.b_bomb == None:
            jsn.append("false")
        else:
            jsn.append(task.b_bomb)
        if task.r_bomb == None:
            jsn.append("false")
        else:
            jsn.append(task.r_bomb)
        #with open('task.json') as file:
        #   jsn = json.load(file)
    return jsn

@app.route('/h50_50', methods=["POST", "GET"])
def h50_50():
    if request.method == 'POST':
        secure_rnd = secrets.SystemRandom()
        r = request.json["round"]
        task = db.session.scalar(db.select(Task).limit(1))
        jsn =[]
            #with open('task.json') as file:
        jsn.append(task.round)
        if task.round == 1:
            jsn.append(task.fatal)
        else:
            jsn.append(json.loads(task.fatal))
        jsn.append(task.md5)
        jsn.append(task.count_fatal)
        if task.b_bomb == None:
            jsn.append("false")
        else:
            jsn.append(task.b_bomb)
        if task.r_bomb == None:
            jsn.append("false")
        else:
            jsn.append(task.r_bomb)
        f = jsn[1]
        cf = round(len(f)/2)
        res_f = secure_rnd.sample(f, cf)
        socketio.emit("response_50_50",res_f,to=f"{get_room_code()}")
        socketio.emit("response_50_50",res_f,to=f"{DEFAULT_ROOM_CODE}:spectator")
        #with open('50_50.json', 'w') as file:
        #    json.dump(res_f, file)
       # with open('50_50_spec.json', 'w') as file:
       #     json.dump(res_f, file)
        return res_f

@app.route('/alter', methods=["POST", "GET"])
def alter():
    if request.method == 'POST':
        secret_rnd = secrets.SystemRandom()
        r = request.json["round"]
        task = db.session.scalar(db.select(Task).limit(1))
        jsn =[]
            #with open('task.json') as file:
        jsn.append(task.round)
        if task.round == 1:
            jsn.append(task.fatal)
        else:
            jsn.append(json.loads(task.fatal))
        jsn.append(task.md5)
        jsn.append(task.count_fatal)
        if task.b_bomb == None:
            jsn.append("false")
        else:
            jsn.append(task.b_bomb)
        if task.r_bomb == None:
            jsn.append("false")
        else:
            jsn.append(task.r_bomb)
        f = jsn[1]
        cf = len(f)
        rf = secret_rnd.randint(1,cf-1)
        j = 0
        checked = False
        while (checked==False):
            checked = True
            j = secret_rnd.randint(1,15)
            for i in f:
                if i == j:
                    checked = False
                    break
        
        res = [f[rf],str(j)]
        socketio.emit("response_alter",res,to=f"{get_room_code()}")
        socketio.emit("response_alter",res,to=f"{DEFAULT_ROOM_CODE}:spectator")
       # with open('alter.json', 'w') as file:
       #     json.dump(res, file)
       # with open('alter_spec.json', 'w') as file:
       #     json.dump(res, file)
        return res
    
@app.route('/navi', methods=["POST", "GET"])
def navi():
    if request.method == 'POST':
        secret_rnd = secrets.SystemRandom()
        r = request.json["round"]
        task = db.session.scalar(db.select(Task).limit(1))
        jsn =[]
            #with open('task.json') as file:
        jsn.append(task.round)
        if task.round == 1:
            jsn.append(task.fatal)
        else:
            jsn.append(json.loads(task.fatal))
        jsn.append(task.md5)
        jsn.append(task.count_fatal)
        if task.b_bomb == None:
            jsn.append("false")
        else:
            jsn.append(task.b_bomb)
        if task.r_bomb == None:
            jsn.append("false")
        else:
            jsn.append(task.r_bomb)
        f = jsn[1]
        f.sort()
        row1 = [0,0,0,0,0]
        row2 = [0,0,0,0,0]
        row3 = [0,0,0,0,0]

        for i in f:
            if (i >=1) & (i<=5):
                row1[i-1] = i
            if (i>=6) & (i<=10):
                row2[i-6] = i
            if (i>=11) & (i<=15):
                row3[i-11] = i
        
        col1 = [row1[0],row2[0],row3[0]]
        col2 = [row1[1],row2[1],row3[1]]
        col3 = [row1[2],row2[2],row3[2]]
        col4 = [row1[3],row2[3],row3[3]]
        col5 = [row1[4],row2[4],row3[4]]
        
        col_s = 0
        for i in row1:
            if (i == 0):
                col_s+=1
                
        row1s = round(col_s/5*100)
        col_s = 0
        for i in row2:
            if (i == 0):
                col_s+=1
                
        row2s = round(col_s/5*100)
        col_s = 0
        for i in row3:
            if (i == 0):
                col_s+=1
                
        row3s = round(col_s/5*100)
        
        
        row_max = [row1s,row2s,row3s]
        
        col_s = 0
        for i in col1:
            if (i == 0):
                col_s+=1
                
        col1s = round(col_s/3*100)
        col_s = 0
        for i in col2:
            if (i == 0):
                col_s+=1
                
        col2s = round(col_s/3*100)
        col_s = 0
        for i in col3:
            if (i == 0):
                col_s+=1
                
        col3s = round(col_s/3*100)
        col_s = 0
        for i in col4:
            if (i == 0):
                col_s+=1
                
        col4s = round(col_s/3*100)
        col_s = 0
        for i in col5:
            if (i == 0):
                col_s+=1
                
        col5s = round(col_s/3*100)
        
        col_max=[col1s,col2s,col3s,col4s,col5s]
        
        max_c = max(col_max)
        max_r = max(row_max)
        if max_c > max_r:
            tmp = -1
            for i in range(0,5):
                if max_c == col_max[i]:
                    tmp = i
                    break
            if tmp==0:
                res = ["1","6","11"]
            if tmp==1:
                res = ["2","7","12"]
            if tmp==2:
                res = ["3","8","13"]
            if tmp==3:
                res = ["4","9","14"]
            if tmp==4:
                res = ["5","10","15"]
            socketio.emit("response_navi",res,to=f"{get_room_code()}")
            socketio.emit("response_navi",res,to=f"{DEFAULT_ROOM_CODE}:spectator")
            #with open('navi.json','w') as file:
            #    json.dump(res,file)
           # with open('navi_spec.json', 'w') as file:
            #    json.dump(res, file)
            return res
        if max_c < max_r:
            tmp = -1
            for i in range(0,3):
                if max_r == row_max[i]:
                    tmp = i
                    break
            if tmp==0:
                res = ["1","2","3","4","5"]
            if tmp==1:
                res = ["6","7","8","9","10"]
            if tmp==2:
                res = ["11","12","13","14","15"]
            socketio.emit("response_navi",res,to=f"{get_room_code()}")
            socketio.emit("response_navi",res,to=f"{DEFAULT_ROOM_CODE}:spectator")
            #with open('navi.json','w') as file:
            #    json.dump(res,file)
            #with open('navi_spec.json', 'w') as file:
            #    json.dump(res, file)
            return res
        if max_r == max_c:
            j = secret_rnd.randint(1,2)
            if j==1:
                tmp = -1
                for i in range(0,5):
                 if max_c == col_max[i]:
                     tmp = i
                     break
                if tmp==0:
                    res = ["1","6","11"]
                if tmp==1:
                    res = ["2","7","12"]
                if tmp==2:
                    res = ["3","8","13"]
                if tmp==3:
                    res = ["4","9","14"]
                if tmp==4:
                    res = ["5","10","15"]
                socketio.emit("response_navi",res,to=f"{get_room_code()}")
                socketio.emit("response_navi",res,to=f"{DEFAULT_ROOM_CODE}:spectator")
                #with open('navi.json','w') as file:
               #     json.dump(res,file)
               # with open('navi_spec.json', 'w') as file:
               #     json.dump(res, file)
                return res
            if j==2:
                tmp = -1
                for i in range(0,3):
                    if max_r == row_max[i]:
                        tmp = i
                        break
                if tmp==0:
                    res = ["1","2","3","4","5"]
                if tmp==1:
                    res = ["6","7","8","9","10"]
                if tmp==2:
                    res = ["11","12","13","14","15"]
                socketio.emit("response_navi",res,to=f"{get_room_code()}")
                socketio.emit("response_navi",res,to=f"{DEFAULT_ROOM_CODE}:spectator")
                #with open('navi.json','w') as file:
                #    json.dump(res,file)
               # with open('navi_spec.json', 'w') as file:
                #    json.dump(res, file)
                return res


    
@app.route('/check_answered', methods=["POST", "GET"])
def check_answered():
    if request.method == 'POST':
        r = request.json["check"]
    if db.session.scalar(db.select(Answered).limit(1))!=None:
        return "true"
    else:
        return "false"   
    
@app.route('/show_rights', methods=["POST", "GET"])
def show_rights():
    if request.method == 'POST':
        js = db.session.scalars(db.select(Users)).all()
        if len(js)==1:
            js[0].status = "check main"
        else:
            jjj = db.session.scalars(db.select(Users)).all()
            for i in range(0,len(jjj)):
                if jjj[i].status == "answered interactive":
                    jjj[i].status = "check interactive"
                if jjj[i].status == "answered main":
                    jjj[i].status = "check main"
                if jjj[i].status ==  "answered main x2":
                    jjj[i].status = "check main x2"
                if jjj[i].status == "interactive no answer":
                    jjj[i].status = "check interactive"
            db.session.commit()
        update_list_users()                                       
        r = request.json["round"]
        ans = db.session.scalar(db.select(Answered).limit(1))
        jsn = [ans.answer, ans.round]
        db.session.execute(db.delete(Answered))
        #if os.path.exists("answered.json"):
         #   with open('answered.json') as file:
         #       jsn = json.load(file)
          #  os.remove("answered.json")
       # if os.path.exists("50_50.json"):
       #     os.remove("50_50.json")
      #  if os.path.exists("alter.json"):
      #     os.remove("alter.json")
      #  if os.path.exists("navi.json"):
      #      os.remove("navi.json")
        check_answer()
        answered_check_spec()
        return jsn
           ## answered interactive
            
            
                

    
 
@app.route('/sounds/slot/<filename>')
def serve_audio(filename):
    CUSTOM_AUDIO_DIR = "sounds/slot/"
    sanitized_filename = secure_filename(filename)

    mime_type, _ = mimetypes.guess_type(sanitized_filename)
    if not mime_type or not mime_type.startswith('audio/'):
        abort(400, description="Unsupported audio format.")

    result = send_from_directory(
        CUSTOM_AUDIO_DIR,
        sanitized_filename,
        mimetype=mime_type,
        as_attachment=False
    )

    result.cache_control.public = True
    result.cache_control.max_age = 432000  # 5 дней
    result.headers["Cache-Control"] = "public, max-age=432000, immutable"

    return result 


    
#@app.route('/update_list_users', methods=["POST", "GET"])
@socketio.on("update_list_users")
def update_list_users():
    #if request.method == 'POST':
        js = db.session.scalars(db.select(Users)).all()
        if len(js)==1:
            id = js[0].id
            username = js[0].username
            answer = js[0].answer
            money = js[0].money
            time = js[0].time
            status = js[0].status
            main_money = js[0].main_money
            red_bomb = js[0].red_bomb
            jsn = [id,username,answer,money,time, status,main_money,red_bomb,"true"]
            result = jsn
            socketio.emit("updated_list_user", result, to=DEFAULT_ROOM_CODE)
            socketio.emit("updated_status_user",js[0].status,to=f"{get_room_code()}:user:{js[0].username}");
            update_for_spec()
        else:
            jsn = []
            for i in range(0,len(js)):
                id = js[i].id
                username = js[i].username
                answer = js[i].answer
                money = js[i].money
                time = js[i].time
                status = js[i].status
                main_money = js[i].main_money
                red_bomb = js[i].red_bomb
                tmp = [id,username,answer,money,time, status,main_money,red_bomb,"false"]
                socketio.emit("updated_status_user",js[i].status,to=f"{get_room_code()}:user:{js[i].username}");
                jsn.append(tmp)
            result = jsn
            socketio.emit("updated_list_user",result, to=DEFAULT_ROOM_CODE)
            update_for_spec()
            return result
            


@app.route('/open_room', methods=["POST", "GET"])
def open_room():
    if request.method == 'POST':
        db.session.execute(db.delete(Room))
        db.session.commit()
        game_type = request.json['game_type']
        try:
            room_code = Room()
            secret_rnd = secrets.SystemRandom()
            room_code.id = secret_rnd.randint(1000,9999)
            room_code.game = game_type
            db.session.add(room_code)
            db.session.flush()
            db.session.commit()
            join_url = f"{request.host_url.rstrip('/')}{url_for('join')}?room={room_code.id}"
            emit_tpv_spectator("room_code_show", {
                "room": room_code.id,
                "joinUrl": join_url,
            })
            return json.dumps(room_code.id)
        except:
            return json.dumps("fail")
    else:
        return json.dumps("fail")
  
 
@app.route('/game_over', methods=["POST", "GET"])
def game_over():
    if request.method == 'POST':
        try:
            lose = request.json['lose']
            main_money = request.json['money']
            u = request.json['user_name']
            rb = request.json['rb']
            user_u = db.session.scalar(db.select(Users).where(Users.id==u))
            match main_money:
                case '0': user_u.main_money = user_u.main_money + 0
                case '1 000': user_u.main_money = user_u.main_money + 1000
                case '3 000': user_u.main_money = user_u.main_money + 3000
                case '5 000': user_u.main_money = user_u.main_money + 5000
                case '10 000': user_u.main_money = user_u.main_money + 10000
                case '25 000': user_u.main_money = user_u.main_money + 25000
                case '50 000': user_u.main_money = user_u.main_money + 50000
                case '150 000': user_u.main_money = user_u.main_money + 150000
                case '500 000': user_u.main_money = user_u.main_money + 500000
                case '1 000 000': user_u.main_money = user_u.main_money + 1000000
            if rb == "true":
                user_u.red_bomb = "true"
            db.session.commit()
            jh = db.session.scalars(db.select(Users)).all()
            for i in range(len(jh)):
                if lose == "true":
                    jh[i].status = "game over lose"
                else:
                    jh[i].status = "game over"
            db.session.commit()
            update_list_users()
            return json.dumps("ok")
        except:
            return json.dumps("fail")
 
        
@app.route('/close_room', methods=["POST", "GET"])
def close_room():
    if request.method == 'POST':
        db.session.execute(db.delete(Room))
        db.session.commit()
        emit_tpv_spectator("room_code_hide", {})
        return json.dumps("ok")
    else:
        return json.dumps("fail")

@app.route('/get_user_status', methods=["POST", "GET"])
def get_user_status():
    if request.method == 'POST':
        try:
         tmp = request.json['user']
         user = Users()
         user = db.session.scalar(db.select(Users).where(Users.username==tmp))
         jsn = user.status
         return json.dumps(jsn)
        except:
            return json.dumps("fail")
        
@app.route('/reset_user_to_wait', methods=["POST", "GET"])
def reset_user_to_wait():
    if request.method == 'POST':
         #if os.path.exists('helps.json'):
         #   os.remove('helps.json')
         db.session.execute(db.delete(Helps))
        # if os.path.exists('answered.json'):
        #    os.remove('answered.json')
         db.session.execute(db.delete(Answered))
         db.session.execute(db.delete(Task))
       #  if os.path.exists("task.json"):
       #     os.remove("task.json")
         js = db.session.scalars(db.select(Users)).all()
         for i in range(0,len(js)):
            js[i].status = "wait"
            js[i].answer = "0"
            js[i].time = 0
            db.session.commit()
            socketio.emit("updated_status_user",js[i].status,to=f"{get_room_code()}:user:{js[i].username}");
         init_game()
         update_list_users()
         return json.dumps(" ")

#@app.route('/get_helps', methods=["POST", "GET"])
def get_helps():
    #if request.method == 'POST':
        #tmp_u = request.json['user']
        try:
            #if os.path.exists("helps.json"):
            user = db.session.scalars(db.select(Users)).all()
            result = []
            for i in range(len(user)):
                if (user[i].status == "main") or (user[i].status == "wait task main") or (user[i].status == "given task main"):
                    help = db.session.scalar(db.select(Helps).limit(1))
                    if help == None:
                       socketio.emit("get_helps", "fail", to=f"{get_room_code()}:user:{user[i].username}")
                       socketio.emit("get_helps", "fail",to=f"{DEFAULT_ROOM_CODE}:spectator")
                       return
                    else:
                        if help.h50_50 == "50:50":
                            result.append(help.h50_50)
                        if help.alter == "alter":
                            result.append(help.alter)
                        if help.navi == "navi":
                            result.append(help.navi)
                        if help.x2 == "x2":
                            result.append(help.x2)
                        if help.auden == "help_auden":
                            result.append(help.auden)
                        if help.fact == "fact":
                            result.append(help.fact)
                        socketio.emit("get_helps", result, to=f"{get_room_code()}:user:{user[i].username}")
                        socketio.emit("get_helps", result, to=f"{DEFAULT_ROOM_CODE}:spectator")
                     
            #    user = Users()
            #    user = db.session.scalar(db.select(Users).where(Users.username==tmp_u))
            #    if (user.status == "main") or (user.status == "wait task main") or (user.status == "given task main"):
            #        db.session.commit()
            #        with open('helps.json') as file:
            #            jsn = json.load(file)
            #        return jsn
            #    else:
            #        return json.dumps("no change")
            #else:
               #  return json.dumps("fail")   
        except:
            return json.dumps("fail")

@app.route('/send_helps', methods=["POST", "GET"])
def send_helps():
    if request.method == 'POST':
        tmp = request.json['helps']
        db.session.execute(db.delete(Helps))
        help = Helps()
        for i in tmp:
            if i == "50:50":
                help.h50_50 = i
            if i == "alter":
                help.alter = i
            if i == "navi":
                help.navi = i
            if i == "x2":
                help.x2 = i
            if i == "help_auden":
                help.auden = i
            if i == "fact":
                help.fact = i
        db.session.add(help)
        db.session.flush()
        db.session.commit()
        get_helps()
        #with open('helps.json','w') as file:
         #   json.dump(tmp,file)
        return json.dumps("OK") 
    else:
        return json.dumps("fail")
  
@app.route('/start_game', methods=["POST", "GET"])
def start_game():
    if request.method == 'POST':
        try:
            js = db.session.scalars(db.select(Users)).all()
            if len(js)!=1:
                for i in range(len(js)):
                    if js[i].status =="main":
                        js[i].status = "wait task main"
                    if js[i].status =="interactive":
                        js[i].status = "wait task interactive"
                    js[i].time = 0
                db.session.commit()
                update_list_users()
            else: 
                return json.dumps("fail")
            
            return json.dumps("ok")
        except:
            return json.dumps("fail")
  
#@app.route('/get_task_user', methods=["POST", "GET"])
def get_task_user():
    #if request.method == 'POST':
        #if not os.path.exists("task.json"):
        #    return json.dumps("fail")
       # with open('task.json') as file:
        #    jsn = json.load(file)
       # if request.json['user']=="spec":
        #    return json.dumps(jsn)
        try:
            task = db.first_or_404(db.select(Task))
            js = db.session.scalars(db.select(Users)).all()
            if len(js)!=1:
                for i in range(len(js)):
                    if js[i].status =="wait task main":
                        js[i].status = "given task main"
                    if js[i].status =="wait task interactive":
                        js[i].status = "given task interactive"
                db.session.commit()
            update_list_users()  
            #return json.dumps(jsn)
        except:
            return json.dumps("fail")
        

#@app.route('/check_answered_main', methods=["POST", "GET"])
def check_answered_main():
    #if request.method == 'POST':
        try:
            #u_tmp = request.json['user']
            #s_tmp = request.json['inter']
            #if not s_tmp:
            #    return json.dumps("fail")
            user2 = db.first_or_404(db.select(Users).where((Users.status == "answered main")|(Users.status == "answered main x2")))
            users = db.session.scalars(db.select(Users)).all()
            for i in range(len(users)):
                if (users[i].status == "given task interactive"):
                    users[i].status = "interactive no answer"
                    db.session.commit()
            update_list_users()
            socketio.emit("check_answered_main","ok",to=f"{get_room_code()}")
            
        except:
            return json.dumps("fail")
 
#@app.route('/answered_main_spec', methods=["POST", "GET"])
def answered_main_spec():
   # if request.method == 'POST':
        try:
            ans = db.first_or_404(db.select(Users).where((Users.status == "answered main")|(Users.status == "answered main x2")))
            if ans!=None:
                 #with open('answered.json') as file:
                 jsn = "o"+str(ans.answer)
                 socketio.emit("answered_main",jsn,to=f"{DEFAULT_ROOM_CODE}:spectator")
            #return json.dumps(jsn)
        except:
            return json.dumps("fail") 



#@app.route('/answered_check_spec', methods=["POST", "GET"])
def answered_check_spec():
    #if request.method == 'POST':
        try:
             task = db.session.scalar(db.select(Task).limit(1))
             jsn =[]
            #with open('task.json') as file:
             jsn.append(task.round)
             if task.round == 1:
                jsn.append(task.fatal)
             else:
                jsn.append(json.loads(task.fatal))
             jsn.append(task.md5)
             jsn.append(task.count_fatal)
             if task.b_bomb == None:
                jsn.append("false")
             else:
                jsn.append(task.b_bomb)
             if task.r_bomb == None:
                jsn.append("false")
             else:
                jsn.append(task.r_bomb)
             socketio.emit("answered_check_spec",jsn,to=f"{DEFAULT_ROOM_CODE}:spectator")
            #return json.dumps(jsn)
        except:
            return json.dumps("fail")  

    
@app.route('/send_answer', methods=["POST", "GET"])
def send_answer():
    if request.method == 'POST':
        try:
            u_tmp = request.json['user']
            a_tnp = request.json['answer_user']
            t_tmp = request.json['time_answer']
            user = db.session.scalar(db.select(Users).where(Users.username == u_tmp))
            if (user.status == "given task main") | (user.status == "check main x2") :
                user.status = "answered main"
                user.answer = a_tnp
                user.time = t_tmp
                db.session.commit()
                update_list_users()
                wait_answer_for_host()
                check_answered_main()
                answered_main_spec()
                return json.dumps("ok")
            if (user.status == "x2"):
                user.status = "answered main x2"
                user.answer = a_tnp
                user.time = t_tmp
                db.session.commit()
                update_list_users()
                wait_answer_for_host()
                check_answered_main()
                answered_main_spec()
                return json.dumps("ok")
            if (user.status == "given task interactive"):
                user.status = "answered interactive"
                user.answer = a_tnp
                user.time = t_tmp
                task = db.session.scalar(db.select(Task).limit(1))
                jsn =[]
                jsn.append(task.round)
                if task.round == 1:
                    jsn.append(task.fatal)
                else:
                    jsn.append(json.loads(task.fatal))
                jsn.append(task.md5)
                jsn.append(task.count_fatal)
                if task.b_bomb == None:
                    jsn.append("false")
                else:
                    jsn.append(task.b_bomb)
                if task.r_bomb == None:
                    jsn.append("false")
                else:
                    jsn.append(task.r_bomb)   
                #with open('task.json') as file:
                 #   jsn = json.load(file)
                fatals = jsn[1]
                c_fatals = jsn[3]
                r = jsn[0]
                wrong = False
                if r==1:
                    if int(user.answer)==fatals:
                        user.money = user.money - 50
                        wrong = True
                if r>1:
                    if (jsn[4]!="false") and (jsn[5]!="false"):
                        if int(user.answer)==jsn[4]:
                            user.money = 0
                            db.session.commit()
                            update_list_users()
                            return json.dumps("ok")
                        if int(user.answer)==jsn[5]:
                           user.money = user.money - 3000*c_fatals
                           db.session.commit()
                           update_list_users()
                           return json.dumps("ok")
                    for i in range(c_fatals):
                        if int(user.answer)==fatals[i]:
                            user.money = user.money - 50*c_fatals
                            wrong = True
                            break
                if not wrong:
                    user.money = user.money + 100*c_fatals
                db.session.commit()
            update_list_users()
            wait_answer_for_host()
            return json.dumps("ok")
        except:
            return json.dumps("fail")



#@app.route('/wait_answer_for_host', methods=["POST", "GET"])
def wait_answer_for_host():
    #if request.method == 'POST':
        try:
            user = db.session.scalar(db.select(Users).where((Users.status == "answered main")|(Users.status == "answered main x2")))
            if user == None:
                return json.dumps("fail")
            res = user.answer
            if res == '0':
                return json.dumps("fail")
            socketio.emit("user answered",res,to=f"{DEFAULT_ROOM_CODE}:host")
            #return json.dumps(res)
        except:
            return json.dumps("fail")

#@app.route('/check_answer', methods=["POST", "GET"])
def check_answer():
   # if request.method == 'POST':
        try:
            #tmp_u = request.json['user']
            users = db.session.scalars(db.select(Users)).all()
            for i in range(len(users)):
                if users[i].status == "check main":
                    users[i].status = "wait next round main"
                if users[i].status == "check interactive":
                    users[i].status = "wait next round interactive"
            #with open('task.json') as file:
            #        jsn = json.load(file)
            task = db.session.scalar(db.select(Task).limit(1))
            jsn =[]
            #with open('task.json') as file:
            jsn.append(task.round)
            if task.round == 1:
                jsn.append(task.fatal)
            else:
                jsn.append(json.loads(task.fatal))
            jsn.append(task.md5)
            jsn.append(task.count_fatal)
            if task.b_bomb == None:
                jsn.append("false")
            else:
                jsn.append(task.b_bomb)
            if task.r_bomb == None:
                jsn.append("false")
            else:
                jsn.append(task.r_bomb)
            db.session.commit()
            update_list_users()
            socketio.emit("checked answer", jsn, to=f"{get_room_code()}")
            #return json.dumps(jsn)
        except:
            return json.dumps("fail")

        
     




@app.route('/next_round', methods=["POST", "GET"])
def next_round():
    if request.method == 'POST':
        try:
            user = db.session.scalars(db.select(Users)).all()
            if len(user)==1:
                user[0].status = "wait task main"
                user[0].answer = "0"
                user[0].time = 0
            else:
                for i in range(len(user)):
                    if (user[i].status == "wait next round main") | (user[i].status == "check main x2"):
                        user[i].status = "wait task main"
                        user[i].answer = "0"
                    if (user[i].status == "wait next round interactive") | (user[i].status == "interactive no answer"):
                        user[i].status = "wait task interactive"
                        user[i].answer = "0"
                    user[i].time = 0
            #if os.path.exists('task.json'):
            #    os.remove('task.json')
            db.session.commit()
            db.session.execute(db.delete(Task))
            get_helps()
            update_list_users()
            return json.dumps("ok")
        except:
            return json.dumps("fail")


#@app.route('/get_50_50', methods=["POST", "GET"])

@socketio.on("get 50:50")
def get_50_50(data=None):
    #if request.method == 'POST':
        try:
         #   tmp_u = request.json['user']
          #  user = db.session.scalar(db.select(Users).where(Users.username==tmp_u))
          #  find = False
           # user.status = "50:50"
           # db.session.commit()
            #while not find:
           # if os.path.exists("50_50.json"):
            #    with open('50_50.json') as file:
            #        p = json.load(file)
               # os.remove("50_50.json")
            #return json.dumps(p)
            socketio.emit("request 50:50", "ok",to=f"{DEFAULT_ROOM_CODE}:host")
        except:
            return json.dumps("fail")
        
      

#@app.route('/get_alter', methods=["POST", "GET"])
@socketio.on("get alter")
def get_alter(data=None):
   # if request.method == 'POST':
        try:
           # tmp_u = request.json['user']
           # user = db.session.scalar(db.select(Users).where(Users.username==tmp_u))
            #find = False
           # user.status = "alter"
           # db.session.commit()
            #while not find:
           # if os.path.exists("alter.json"):
           #     with open('alter.json') as file:
           #         p = json.load(file)
                        #find = True
                       # user.status = "given task main"
            #db.session.commit()
               # os.remove("alter.json")
            socketio.emit("request alter", "ok",to=f"{DEFAULT_ROOM_CODE}:host")
            #return json.dumps(p)
        except:
            return json.dumps("fail")



#@app.route('/get_navi', methods=["POST", "GET"])
@socketio.on("get navi")
def get_navi(data=None):
   # if request.method == 'POST':
        try:
           # tmp_u = request.json['user']
          #  user = db.session.scalar(db.select(Users).where(Users.username==tmp_u))
           # find = False
            #user.status = "navi"
           # db.session.commit()
            #while not find:
           # if os.path.exists("navi.json"):
          #      with open('navi.json') as file:
          #          p = json.load(file)
                        #find = True
                        #user.status = "given task main"
            #db.session.commit()
                #os.remove("navi.json")
            socketio.emit("request navi", "ok",to=f"{DEFAULT_ROOM_CODE}:host")
           # return json.dumps(p)
        except:
            return json.dumps("fail")
        




#@app.route('/get_x2', methods=["POST", "GET"])
@socketio.on("get x2")
def get_x2(data=None):
   # if request.method == 'POST':
        try:
           # tmp_u = request.json['user']
            user = db.session.scalar(db.select(Users).where(Users.status=="given task main"))
            user.status = "x2"
            db.session.commit()
            update_list_users()
            socketio.emit("request x2", "ok", to=f"{DEFAULT_ROOM_CODE}:host")
            socketio.emit("response_x2", "ok", to=f"{DEFAULT_ROOM_CODE}:spectator")
            socketio.emit("response_x2", "ok", to=f"{get_room_code()}:user:{user.username}")
            return json.dumps("ok")
        except:
            return json.dumps("fail")


@app.route('/help_auden', methods=["POST", "GET"])
def help_auden():
    if request.method == 'POST':
        try:
            tmp_u = db.session.scalars(db.select(Users)).all()
            a1 = 0
            a2 = 0
            a3 = 0
            a4 = 0
            a5 = 0
            a6 = 0
            a7 = 0
            a8 = 0
            a9 = 0
            a10 = 0
            a11 = 0
            a12 = 0
            a13 = 0
            a14 = 0
            a15 = 0
            result = []
            for i in range (len(tmp_u)):
                if tmp_u[i].status == "answered interactive":
                    if tmp_u[i].answer == "1":
                        a1 += 1
                    if tmp_u[i].answer == "2":
                        a2 += 1
                    if tmp_u[i].answer == "3":
                        a3 += 1
                    if tmp_u[i].answer == "4":
                        a4 += 1
                    if tmp_u[i].answer == "5":
                        a5 += 1
                    if tmp_u[i].answer == "6":
                        a6 += 1
                    if tmp_u[i].answer == "7":
                        a7 += 1
                    if tmp_u[i].answer == "8":
                        a8 += 1
                    if tmp_u[i].answer == "9":
                        a9 += 1
                    if tmp_u[i].answer == "10":
                        a10 += 1
                    if tmp_u[i].answer == "11":
                        a11 += 1
                    if tmp_u[i].answer == "12":
                        a12 += 1
                    if tmp_u[i].answer == "13":
                        a13 += 1
                    if tmp_u[i].answer == "14":
                        a14 += 1
                    if tmp_u[i].answer == "15":
                        a15 += 1
            col_find_fatal = 0
            col_find_free = 0
            col_ans = (db.session.scalar(db.select(db.func.count(Users.id)).where(Users.status == "answered interactive")) or 0)
            if col_ans == 0:
                return jsonify("fail")
            result.append(round(a1/col_ans*100,2))
            result.append(round(a2/col_ans*100,2))
            result.append(round(a3/col_ans*100,2))
            result.append(round(a4/col_ans*100,2))
            result.append(round(a5/col_ans*100,2))
            result.append(round(a6/col_ans*100,2))
            result.append(round(a7/col_ans*100,2))
            result.append(round(a8/col_ans*100,2))
            result.append(round(a9/col_ans*100,2))
            result.append(round(a10/col_ans*100,2))
            result.append(round(a11/col_ans*100,2))
            result.append(round(a12/col_ans*100,2))
            result.append(round(a13/col_ans*100,2))
            result.append(round(a14/col_ans*100,2))
            result.append(round(a15/col_ans*100,2))
            jsn = []
            task = db.session.scalar(db.select(Task).limit(1))
            jsn.append(task.round)
            if task.round == 1:
                jsn.append(task.fatal)
            else:
                jsn.append(json.loads(task.fatal))
            jsn.append(task.md5)
            jsn.append(task.count_fatal)
            if task.b_bomb == None:
                jsn.append("false")
            else:
                jsn.append(task.b_bomb)
            if task.r_bomb == None:
                jsn.append("false")
            else:
                jsn.append(task.r_bomb)
            fatals = jsn[1]
            fatals.sort()
            for i in range(jsn[3]):
                for j in range(len(result)):
                    if ((j+1)==fatals[i]) & (result[j]>0):
                        col_find_fatal+=1
                        break
            
            result.append(round(col_find_fatal/col_ans*100))
            result.append(100-round(col_find_fatal/col_ans*100))   
            #with open('auden.json','w') as file:
           #     json.dump(result,file)
           # with open('auden_spec.json', 'w') as file:
            #    json.dump(result, file)
           # time.sleep(10)
            #if os.path.exists("auden.json"):
             #   os.remove("auden.json")        
            socketio.emit("response_auden",result,to=f"{DEFAULT_ROOM_CODE}:spectator")
            return json.dumps(result)
        except:
            return json.dumps("fail")

#@app.route('/get_auden', methods=["POST", "GET"])
@socketio.on("get auden")
def get_auden(data=None):
   # if request.method == 'POST':
        try:
         #   uu = request.json['user']
          #  user = db.session.scalar(db.select(Users).where(Users.username==uu))
            #user.status = "auden"
            #db.session.commit()
            #while (True):
           # if os.path.exists("auden.json"):
                    #break
            #    pass
            #user.status = "given task main"
            #db.session.commit()
           socketio.emit("request auden", "ok",to=f"{DEFAULT_ROOM_CODE}:host") 
           # return json.dumps("ok")
        except:
            return json.dumps("fail")



@app.route('/fact', methods=["POST", "GET"])
def fact(data=None):
    if request.method == 'POST':
        try:
            secret_rnd = secrets.SystemRandom()
            slot = secret_rnd.randint(1,15)
            jsn = []
            task = db.session.scalar(db.select(Task).limit(1))
            jsn.append(task.round)
            if task.round == 1:
                jsn.append(task.fatal)
            else:
                jsn.append(json.loads(task.fatal))
            jsn.append(task.md5)
            jsn.append(task.count_fatal)
            if task.b_bomb == None:
                jsn.append("false")
            else:
                jsn.append(task.b_bomb)
            if task.r_bomb == None:
                jsn.append("false")
            else:
                jsn.append(task.r_bomb)
            fat = jsn[1]
            find = False
            
            while not find:
                find = True
                for i in range(jsn[3]):
                    if fat[i]==slot:
                        slot = secret_rnd.randint(1,15)
                        find = False
                        break
            facts = db.session.scalar(db.select(Facts).where(Facts.slot==str(slot)))
            result = []
            result.append(facts.slot)
            result.append(facts.des_fact)
            socketio.emit("response_fact",result,to=f"{DEFAULT_ROOM_CODE}:spectator")
            #with open('fact.json','w') as file:
           #     json.dump(result,file)
           # with open('fact_spec.json', 'w') as file:
            #    json.dump(result, file)
            db.session.delete(facts)
            db.session.commit()
           # time.sleep(10)
            #if os.path.exists("fact.json"):
             #   os.remove("fact.json")   
            return json.dumps(result)
        except:
            return json.dumps("fail")

@socketio.on("get fact")
#@app.route('/get_fact', methods=["POST", "GET"])
def get_fact(data=None):
   # if request.method == 'POST':
        try:
          #  uu = request.json['user']
          #  user = db.session.scalar(db.select(Users).where(Users.username==uu))
           # user.status = "fact"
           # db.session.commit()
           # while (True):
            #if os.path.exists("fact.json"):
                #break
           #     pass
           # user.status = "given task main"
           # db.session.commit()
            #return json.dumps("ok")
            socketio.emit("request fact", "ok",to=f"{DEFAULT_ROOM_CODE}:host") 
        except:
            return json.dumps("fail")
        



##
@app.route('/clear_table', methods=["POST", "GET"])
def clear_table():
    if request.method == 'POST':
        user_all = db.session.execute(db.delete(Users))
        update_list_users()
        db.session.commit()
        return json.dumps("ok")
        
    return json.dumps("fail")



#@app.route('/update_for_spec', methods=["POST", "GET"])
def update_for_spec():
   # if request.method == 'POST':
        user_waits = (db.session.scalar(db.select(db.func.count(Users.id)).where(Users.status == "wait")) or 0)
        user_all = db.session.scalars(db.select(Users)).all()
        if len(user_all)==0:
            socketio.emit("updated_list_user_spec","fail",to=f"{DEFAULT_ROOM_CODE}:spectator")
            return json.dumps("fail")
        if (len(user_all)==user_waits):
            socketio.emit("updated_list_user_spec","wait",to=f"{DEFAULT_ROOM_CODE}:spectator")
            return json.dumps("wait")
        user_main = db.session.scalar(db.select(Users).where(Users.status == "main"))
        if (user_main != None):
            res = []
            res.append(user_main.status)
            res.append(user_main.username)
            socketio.emit("updated_list_user_spec",res,to=f"{DEFAULT_ROOM_CODE}:spectator")
            return json.dumps(res)
        user_main = db.session.scalar(db.select(Users).where(Users.status == "wait task main"))
        if (user_main != None):
            res = []
            res.append(user_main.status)
            res.append(user_main.username)
            socketio.emit("updated_list_user_spec",res,to=f"{DEFAULT_ROOM_CODE}:spectator")
            return json.dumps(res)
        user_main = db.session.scalar(db.select(Users).where(Users.status == "given task main"))
        if (user_main != None):
            res = []
            res.append(user_main.status)
            res.append(user_main.username)
            socketio.emit("updated_list_user_spec",res,to=f"{DEFAULT_ROOM_CODE}:spectator")
            return json.dumps(res)
        user_main = db.session.scalar(db.select(Users).where(Users.status == "answered main"))
        if (user_main != None):
            res = []
            res.append(user_main.status)
            res.append(user_main.username)
            socketio.emit("updated_list_user_spec",res,to=f"{DEFAULT_ROOM_CODE}:spectator")
            return json.dumps(res)
        user_main = db.session.scalar(db.select(Users).where(Users.status == "answered main x2"))
        if (user_main != None):
            res = []
            res.append(user_main.status)
            res.append(user_main.username)
            return json.dumps(res)
        
        user_main = db.session.scalar(db.select(Users).where(Users.status == "check main"))
        if (user_main != None):
            res = []
            res.append(user_main.status)
            res.append(user_main.username)
            socketio.emit("updated_list_user_spec",res,to=f"{DEFAULT_ROOM_CODE}:spectator")
            return json.dumps(res)
        user_main = db.session.scalar(db.select(Users).where(Users.status == "check main x2"))
        if (user_main != None):
            res = []
            res.append(user_main.status)
            res.append(user_main.username)
            socketio.emit("updated_list_user_spec",res,to=f"{DEFAULT_ROOM_CODE}:spectator")
            return json.dumps(res)
        user_main = db.session.scalar(db.select(Users).where(Users.status == "game over lose"))
        if (user_main != None):
            res = []
            res.append(user_main.status)
            res.append(user_main.username)
            socketio.emit("updated_list_user_spec",res,to=f"{DEFAULT_ROOM_CODE}:spectator")
            return json.dumps(res)
        user_main = db.session.scalar(db.select(Users).where(Users.status == "game over"))
        if (user_main != None):
            res = []
            res.append(user_main.status)
            res.append(user_main.username)
            socketio.emit("updated_list_user_spec",res,to=f"{DEFAULT_ROOM_CODE}:spectator")
            return json.dumps(res)
        user_main = db.session.scalar(db.select(Users).where(Users.status == "50:50"))
        if (user_main != None):
            res = []
            res.append(user_main.status)
            res.append(user_main.username)
            socketio.emit("updated_list_user_spec",res,to=f"{DEFAULT_ROOM_CODE}:spectator")
            return json.dumps(res)
        user_main = db.session.scalar(db.select(Users).where(Users.status == "alter"))
        if (user_main != None):
            res = []
            res.append(user_main.status)
            res.append(user_main.username)
            socketio.emit("updated_list_user_spec",res,to=f"{DEFAULT_ROOM_CODE}:spectator")
            return json.dumps(res)
        user_main = db.session.scalar(db.select(Users).where(Users.status == "x2"))
        if (user_main != None):
            res = []
            res.append(user_main.status)
            res.append(user_main.username)
            socketio.emit("updated_list_user_spec",res,to=f"{DEFAULT_ROOM_CODE}:spectator")
            return json.dumps(res)
        user_main = db.session.scalar(db.select(Users).where(Users.status == "navi"))
        if (user_main != None):
            res = []
            res.append(user_main.status)
            res.append(user_main.username)
            socketio.emit("updated_list_user_spec",res,to=f"{DEFAULT_ROOM_CODE}:spectator")
            return json.dumps(res)
        user_main = db.session.scalar(db.select(Users).where(Users.status == "auden"))
        if (user_main != None):
            res = []
            res.append(user_main.status)
            res.append(user_main.username)
            socketio.emit("updated_list_user_spec",res,to=f"{DEFAULT_ROOM_CODE}:spectator")
            return json.dumps(res)
        user_main = db.session.scalar(db.select(Users).where(Users.status == "fact"))
        if (user_main != None):
            res = []
            res.append(user_main.status)
            res.append(user_main.username)
            socketio.emit("updated_list_user_spec",res,to=f"{DEFAULT_ROOM_CODE}:spectator")
            return json.dumps(res)
        user_main = db.session.scalar(db.select(Users).where(Users.status == "otbor"))
        if (user_main != None):
            res = []
            res.append(user_main.status)
            res.append(user_main.username)
            socketio.emit("updated_list_user_spec",res,to=f"{DEFAULT_ROOM_CODE}:spectator")
            return json.dumps(res)
        user_main = db.session.scalar(db.select(Users).where(Users.status == "warning otbor"))
        if (user_main != None):
            res = []
            res.append(user_main.status)
            res.append(user_main.username)
            socketio.emit("updated_list_user_spec",res,to=f"{DEFAULT_ROOM_CODE}:spectator")
            return json.dumps(res)
        user_main = db.session.scalar(db.select(Users).where(Users.status == "start otbor"))
        if (user_main != None):
            res = []
            res.append(user_main.status)
            res.append(user_main.username)
            socketio.emit("updated_list_user_spec",res,to=f"{DEFAULT_ROOM_CODE}:spectator")
            return json.dumps(res)
        user_main = db.session.scalar(db.select(Users).where(Users.status == "winner otbor"))
        if (user_main != None):
            res = []
            res.append(user_main.status)
            res.append(user_main.username)
            socketio.emit("updated_list_user_spec",res,to=f"{DEFAULT_ROOM_CODE}:spectator")
            return json.dumps(res)
        user_main = db.session.scalar(db.select(Users).where(Users.status == "otbor end"))
        if (user_main != None):
            res = []
            res.append(user_main.status)
            res.append(user_main.username)
            socketio.emit("updated_list_user_spec",res,to=f"{DEFAULT_ROOM_CODE}:spectator")
            return json.dumps(res) 
        user_main = db.session.scalar(db.select(Users).where(Users.status == "show result"))
        if (user_main != None):
            res = get_result()
            socketio.emit("updated_list_user_spec",res,to=f"{DEFAULT_ROOM_CODE}:spectator")
            return json.dumps(res)
        user_main = db.session.scalar(db.select(Users).where(Users.status == "show total result"))
        if (user_main != None):
            res = get_total_result()
            socketio.emit("updated_list_user_spec",res,to=f"{DEFAULT_ROOM_CODE}:spectator")
            return json.dumps(res)


def get_result():
    result = []
    tmp = db.session.scalars(db.select(Users).order_by(desc(Users.money))).all()
    for i in range(len(tmp)):
        status = tmp[i].status
        username = tmp[i].username
        money = tmp[i].money
        t = [status, username, money]
        result.append(t)
    return result

def get_total_result():
    result = []
    tmp = db.session.scalars(db.select(Users).order_by(desc(Users.money+Users.main_money))).all()
    for i in range(len(tmp)):
        status = tmp[i].status
        username = tmp[i].username
        money = tmp[i].money + tmp[i].main_money
        t = [status, username, money]
        result.append(t)
    return result


#@app.route('/send_script', methods=["POST", "GET"])
@socketio.on("send_script")
def send_script(data):
    #if request.method == 'POST':
        try:
           # scr = request.json['script']
            scr = data.get("script")
            #with open('script.json', 'w') as file:
            #    json.dump(scr, file)
            socketio.emit("updated_script",scr,to=DEFAULT_ROOM_CODE) 
        except:
            return json.dumps("fail")
    
    
#@app.route('/send_fix', methods=["POST", "GET"])
@socketio.on("send_fix")
def send_fix(data):
    #if request.method == 'POST':
        try:
            #scr = request.json['fix']
            scr = data.get("fix")
            #with open('fix.json', 'w') as file:
            #    json.dump(scr, file)
           # return json.dumps("ok")
            socketio.emit("updated_fix",scr,to=DEFAULT_ROOM_CODE)  
        except:
            return json.dumps("fail")
        

#@app.route('/send_round', methods=["POST", "GET"])
@socketio.on("send_round")
def send_round(data):
    #if request.method == 'POST':
        try:
            scr = data.get("round")
            #scr = request.json['round']
           # with open('round.json', 'w') as file:
           #     json.dump(scr, file)
           # return json.dumps("ok")
            socketio.emit("updated_round",scr,to=DEFAULT_ROOM_CODE) 
        except:
            return json.dumps("fail")
        

@app.route('/get_tree', methods=["POST", "GET"])
def get_tree():
    if request.method == 'POST':
        try:
            res = []
            with open('script.json') as file:
                jsn = json.load(file)
            res.append(jsn)
            with open('fix.json') as file:
                jsn = json.load(file)
            res.append(jsn)
            with open('round.json') as file:
                jsn = json.load(file)
            res.append(jsn)
            return json.dumps(res)
        except:
            return json.dumps("fail")

@app.route('/otbor', methods=["POST", "GET"])
def otbor():
    if request.method == 'POST':
        try:
            u_all = db.session.scalars(db.select(Users)).all()
            for i in range(len(u_all)):
                if u_all[i].red_bomb == 'true':
                    continue
                u_all[i].status = "otbor"
                u_all[i].time = "0"
                socketio.emit("updated_status_user",u_all[i].status,to=f"{get_room_code()}:user:{u_all[i].username}");
            db.session.commit()
            update_list_users()
            return json.dumps("ok")
        except:
            return json.dumps("fail")
        
@app.route('/warning_otbor', methods=["POST", "GET"])
def warning_otbor():
    if request.method == 'POST':
        try:
            u_all = db.session.scalars(db.select(Users)).all()
            for i in range(len(u_all)):
                if u_all[i].red_bomb == 'true':
                    continue
                u_all[i].status = "warning otbor"
                db.session.commit()
                update_list_users()
                socketio.emit("updated_status_user",u_all[i].status,to=f"{get_room_code()}:user:{u_all[i].username}");
            return json.dumps("ok")
        except:
            return json.dumps("fail")

@app.route('/start_otbor', methods=["POST", "GET"])
def start_otbor():
    if request.method == 'POST':
        try:
            u_all = db.session.scalars(db.select(Users)).all()
            for i in range(len(u_all)):
                if u_all[i].red_bomb == 'true':
                    continue
                u_all[i].status = "start otbor"
                db.session.commit()
                update_list_users()
                socketio.emit("updated_status_user",u_all[i].status,to=f"{get_room_code()}:user:{u_all[i].username}");
            return json.dumps("ok")
        except:
            return json.dumps("fail")
        
@app.route('/show_answer_otbor', methods=["POST", "GET"])
def show_answer_otbor():
    if request.method == 'POST':
        try:
            u_all = db.session.scalars(db.select(Users)).all()
            with open('task_otbor.json') as file:
                p = json.load(file)
            for i in range(len(u_all)):
                if u_all[i].red_bomb == 'true':
                    continue
                u_all[i].status = "otbor end"
                db.session.commit()
                update_list_users()
                socketio.emit("updated_status_user",u_all[i].status,to=f"{get_room_code()}:user:{u_all[i].username}");
                socketio.emit("get_answer_otbor",p,to=f"{DEFAULT_ROOM_CODE}:spectator");
                socketio.emit("get_answer_otbor",p,to=f"{get_room_code()}:user:{u_all[i].username}");
            #os.remove('task_otbor.json')
            return json.dumps(p)
        except:
            return json.dumps("fail")


@app.route('/get_task_otbor', methods=["POST", "GET"])
def get_task_otbor():
    if request.method == 'POST':
        try:
            if os.path.exists("task_otbor.json"):
                with open('task_otbor.json') as file:
                    p = json.load(file)
                return json.dumps(p)
            else:
                return json.dumps("fail")
        except:
            return json.dumps("fail")

@app.route('/send_answer_otbor', methods=["POST", "GET"])
def send_answer_otbor():
    if request.method == 'POST':
        try:
            ans = request.json['ans_otbor']
            user = request.json['user']
            time = request.json['time_answer']
            user_db = db.session.scalar(db.select(Users).where(Users.username == user))
            user_db.answer = ans
            user_db.time = time
            db.session.commit()
            update_list_users()
            return json.dumps("ok")
        except:
            return json.dumps("fail")


@app.route('/show_result_otbor', methods=["POST", "GET"])
def show_result_otbor():
    if request.method == 'POST':
        try:
            user_all = db.session.scalars(db.select(Users)).all()
            with open('task_otbor.json') as file:
                    p = json.load(file)
            abs_arr = []
            for i in range(len(user_all)):
                if user_all[i].answer == '0':
                    abs_arr.append(1000000)
                else:
                    try:
                        abs_a = abs(int(p[3])-int(user_all[i].answer))
                        abs_arr.append(abs_a)
                    except:
                        abs_arr.append(1000000)
            min_abs_arr = min(abs_arr)
            users_ans_win = []
            for i in range(len(abs_arr)):
                if abs_arr[i] == min_abs_arr:
                    users_ans_win.append(i)
            if len(users_ans_win)==1:
                user_winner = db.session.scalar(db.select(Users).where(Users.id==users_ans_win[0]+1))
                user_winner.status = "winner otbor"
                db.session.commit()
                #return json.dumps(user_winner)
            else:
                arr_time = []
                for i in range(len(users_ans_win)):
                    tmp_time = user_all[users_ans_win[i]].time
                    arr_time.append(tmp_time)
                min_time = min(arr_time)
                for i in range(len(user_all)):
                    if ((user_all[i].time == min_time) and (min_abs_arr==abs(int(user_all[i].answer)-int(p[3])))):
                        user_winner = user_all[i]
                        user_winner.status = "winner otbor"
                        db.session.commit()
                       # return json.dumps(user_winner)   
            result = []
            result.append(user_winner.id)
            result.append(user_winner.username)
            result.append(user_winner.answer)
            result.append(user_winner.time)
            
            socketio.emit("show_winner_otbor",update_list_users(),to=f"{DEFAULT_ROOM_CODE}:spectator")
            return json.dumps(result)
        except:
            return json.dumps("fail")

@app.route('/show_result_interactive', methods=["POST", "GET"])
def show_result_interactive():
    if request.method == 'POST':
        try:
            action = request.json['action']
            user_all = db.session.scalars(db.select(Users)).all()
            for i in range(len(user_all)):
                if action == "show":
                    user_all[i].status = "show result"
                if action == "show total":
                    user_all[i].status = "show total result"
                if action == "hide":
                    user_all[i].status = "wait"
            db.session.commit()
            update_list_users()
            return json.dumps("ok")
        except:
            return json.dumps("fail")

@socketio.on("wait_4_min")
def wait_4_min():
    socketio.emit("wait_4min","ok",to=f"{DEFAULT_ROOM_CODE}:spectator")
    
@socketio.on("wait_1_min")
def wait_1_min():
    socketio.emit("wait_1min","ok",to=f"{DEFAULT_ROOM_CODE}:spectator")

@socketio.on("host_show_credits")
def host_show_credits():
    socketio.emit("show_credits", {
    "title": "Спасибо за игру!",
    "lines": [
        "Ведущий: Mokaque",
        "Оригинальная идея: Сергей Бойцов,  Игорь Черкасов",
        "Композитор: Дмитрий Яковлев",
        "Адаптация правил и игры: Mokaque",
        "Техническая реализация: Mokaque",
        "Графика: ChatGPT",
        "Никто из участников создания данной адаптации игры не претендует на авторские права на формат оригинальной игры 'Свободный слот'",
        "Данный проект выпущен исключительно в развлекательных целях и не преследует целей получение материальной выгоды",
        "Оригинальная игра 'Свободный слот' проводится в онлайн формате на https://www.twitch.tv/fighter_kit",
        "До встречи в следующей игре!"
    ]
}, to=f"{DEFAULT_ROOM_CODE}:spectator")
    

@socketio.on("host_hide_credits")
def host_hide_credits():
    socketio.emit("hide_credits", {}, to=f"{DEFAULT_ROOM_CODE}:spectator")


@socketio.on("show_intro")
def handle_show_intro():
    emit("show_intro", to=f"{DEFAULT_ROOM_CODE}:spectator")




#---------------------------------------------
#TPV


@app.route('/sounds/tpv/<filename>')
def serve_audio_tpv(filename):
    CUSTOM_AUDIO_DIR = "sounds/tpv/"
    sanitized_filename = secure_filename(filename)

    mime_type, _ = mimetypes.guess_type(sanitized_filename)
    #if not mime_type or not mime_type.startswith('audio/'):
     #   abort(400, description="Unsupported audio format.")

    result = send_from_directory(
        CUSTOM_AUDIO_DIR,
        sanitized_filename,
        mimetype=mime_type,
        as_attachment=False
    )

    result.cache_control.public = True
    result.cache_control.max_age = 432000  # 5 дней
    result.headers["Cache-Control"] = "public, max-age=432000, immutable"

    return result  

@app.route('/sounds/tpv/bong-game/<filename>')
def serve_audio_tpv_bong(filename):
    CUSTOM_AUDIO_DIR = "sounds/tpv/bong-game/"
    sanitized_filename = secure_filename(filename)

    mime_type, _ = mimetypes.guess_type(sanitized_filename)
    if not mime_type or not mime_type.startswith('audio/'):
        abort(400, description="Unsupported audio format.")

    result = send_from_directory(
        CUSTOM_AUDIO_DIR,
        sanitized_filename,
        mimetype=mime_type,
        as_attachment=False
    )

    result.cache_control.public = True
    result.cache_control.max_age = 432000  # 5 дней
    result.headers["Cache-Control"] = "public, max-age=432000, immutable"

    return result  


@app.post("/api/voice-number")
def api_voice_number():
    data = request.get_json(silent=True) or {}

    try:
        number = int(data.get("number"))
    except (TypeError, ValueError):
        return jsonify({
            "ok": False,
            "error": "Необходимо передать целое число",
        }), 400

    include_currency = bool(
        data.get("include_currency", False)
    )

    try:
        filenames = number_to_audio(
            number,
            include_currency=include_currency,
        )
    except ValueError as error:
        return jsonify({
            "ok": False,
            "error": str(error),
        }), 400

    urls = [
        url_for(
            "serve_audio_tpv_bong",
            filename=f"{filename}",
        )
        for filename in filenames
    ]

    return jsonify({
        "ok": True,
        "number": number,
        "files": filenames,
        "urls": urls,
    })


class UsersTpv(db.Model):
    id = db.Column(
        db.Integer,
        primary_key=True,
        autoincrement=True,
    )
    username = db.Column(
        db.String(10),
        unique=True,
        nullable=False,
    )
    flip = db.Column(db.Text)
    money = db.Column(db.Integer, default=0)
    approve = db.Column(db.Text)
    flip_col = db.Column(db.Integer, default=0)
    def __repr__(self):
            return '<UsersTpv %r>' %self.id


class Questions_tpv(db.Model):

    id = db.Column(
        db.Integer,
        primary_key=True,
        autoincrement=True,
    )
    task = db.Column(db.Text)
    answer = db.Column(db.Text)
    comment = db.Column(db.Text)
    author = db.Column(db.Text)
    flip = db.Column(db.Text)
    show = db.Column(db.Text)
    def __repr__(self):
        return '<Questions_tpv %r>' %self.id


class QueryTpv(db.Model, UserMixin):
    id = db.Column(
        db.Integer,
        primary_key=True,
        autoincrement=True,
    )
    username = db.Column(
        db.String(10),
        unique=True,
        nullable=False,
    )
    flip = db.Column(db.Text)
    money = db.Column(db.Integer, default=0)
    status = db.Column(db.Text, default="wait")
    def __repr__(self):
            return '<QueryTpv %r>' %self.id
# Вставить в game.py после объявления UsersTpv и Questions_tpv.
TPV_REQUIRED_FLIP_QUESTIONS = 5
TPV_GENERAL_QUESTION_VALUES = {"", "false", "общий"}



# ============================================================================
# TPV EDITOR — ЭТАП 9: ИСТОРИЯ ИЗМЕНЕНИЙ
# ============================================================================
import json as _tpv_history_json
from datetime import datetime as _tpv_history_datetime
from sqlalchemy import inspect as _tpv_history_inspect


class TpvEditorHistory(db.Model):
    __tablename__ = "tpv_editor_history"

    id = db.Column(db.Integer, primary_key=True)
    created_at = db.Column(
        db.DateTime,
        nullable=False,
        default=_tpv_history_datetime.utcnow,
        index=True,
    )
    entity_type = db.Column(db.String(24), nullable=False, index=True)
    entity_id = db.Column(db.String(64), nullable=True, index=True)
    action = db.Column(db.String(24), nullable=False, index=True)
    title = db.Column(db.String(255), nullable=False)
    details = db.Column(db.Text, nullable=True)
    before_json = db.Column(db.Text, nullable=True)
    after_json = db.Column(db.Text, nullable=True)
    can_revert = db.Column(db.Boolean, nullable=False, default=False)
    reverted_at = db.Column(db.DateTime, nullable=True)
    revert_history_id = db.Column(db.Integer, nullable=True)


def tpv_editor_history_table_exists():
    try:
        return "tpv_editor_history" in _tpv_history_inspect(db.engine).get_table_names()
    except Exception:
        return False


def tpv_editor_history_create_table():
    TpvEditorHistory.__table__.create(bind=db.engine, checkfirst=True)


def tpv_editor_history_json(value):
    if value is None:
        return None
    return _tpv_history_json.dumps(value, ensure_ascii=False, sort_keys=True)


def tpv_editor_history_parse(value):
    if not value:
        return None
    try:
        return _tpv_history_json.loads(value)
    except (_tpv_history_json.JSONDecodeError, TypeError):
        return None


def tpv_editor_user_snapshot(user):
    if user is None:
        return None
    return {
        "id": user.id,
        "username": user.username,
        "money": int(user.money or 0),
        "flip": user.flip,
        "flip_col": int(user.flip_col or 0),
        "approve": user.approve,
    }


def tpv_editor_question_snapshot(question):
    if question is None:
        return None
    return {
        "id": question.id,
        "task": question.task,
        "answer": question.answer,
        "comment": question.comment,
        "author": question.author,
        "flip": question.flip,
        "show": question.show,
    }


def tpv_editor_history_add(
    entity_type,
    entity_id,
    action,
    title,
    before=None,
    after=None,
    details="",
    can_revert=False,
):
    if not tpv_editor_history_table_exists():
        return None

    row = TpvEditorHistory(
        entity_type=entity_type,
        entity_id=str(entity_id) if entity_id is not None else None,
        action=action,
        title=title,
        details=details or "",
        before_json=tpv_editor_history_json(before),
        after_json=tpv_editor_history_json(after),
        can_revert=bool(can_revert),
    )
    db.session.add(row)
    return row


def tpv_editor_history_to_dict(row):
    action_labels = {
        "create": "Создание",
        "update": "Изменение",
        "delete": "Удаление",
        "reset": "Сброс",
        "rename": "Переименование",
        "import": "Импорт",
        "fix": "Исправление",
        "revert": "Откат",
    }
    entity_labels = {
        "user": "Пользователь",
        "question": "Вопрос",
        "theme": "Тема",
        "bulk": "Массовая операция",
    }
    return {
        "id": row.id,
        "created_at": row.created_at.isoformat(timespec="seconds"),
        "entity_type": row.entity_type,
        "entity_label": entity_labels.get(row.entity_type, row.entity_type),
        "entity_id": row.entity_id,
        "action": row.action,
        "action_label": action_labels.get(row.action, row.action),
        "title": row.title,
        "details": row.details or "",
        "before": tpv_editor_history_parse(row.before_json),
        "after": tpv_editor_history_parse(row.after_json),
        "can_revert": bool(row.can_revert),
        "reverted": row.reverted_at is not None,
        "reverted_at": (
            row.reverted_at.isoformat(timespec="seconds")
            if row.reverted_at else None
        ),
        "revert_history_id": row.revert_history_id,
    }


def tpv_editor_restore_user(snapshot):
    if not snapshot:
        return None

    user = db.session.get(UsersTpv, int(snapshot["id"]))

    if user is None:
        duplicate = db.session.scalar(
            db.select(UsersTpv).where(
                func.lower(UsersTpv.username)
                == str(snapshot["username"]).casefold()
            )
        )
        if duplicate is not None:
            raise ValueError(
                "Нельзя восстановить пользователя: имя уже занято."
            )
        user = UsersTpv()
        user.id = int(snapshot["id"])
        db.session.add(user)

    user.username = snapshot["username"]
    user.money = int(snapshot.get("money") or 0)
    user.flip = snapshot.get("flip") or "false"
    user.flip_col = int(snapshot.get("flip_col") or 0)
    user.approve = snapshot.get("approve") or "false"
    return user


def tpv_editor_restore_question(snapshot):
    if not snapshot:
        return None

    question = db.session.get(Questions_tpv, int(snapshot["id"]))

    if question is None:
        question = Questions_tpv()
        question.id = int(snapshot["id"])
        db.session.add(question)

    question.task = snapshot.get("task") or ""
    question.answer = snapshot.get("answer") or ""
    question.comment = snapshot.get("comment") or ""
    question.author = snapshot.get("author") or ""
    question.flip = snapshot.get("flip") or "false"
    question.show = snapshot.get("show") or "false"
    return question


def tpv_editor_revert_history_row(row):
    before = tpv_editor_history_parse(row.before_json)
    after = tpv_editor_history_parse(row.after_json)

    if row.entity_type == "user":
        if row.action == "create":
            user = db.session.get(UsersTpv, int(row.entity_id))
            if user is not None:
                db.session.delete(user)
        elif row.action in {"update", "reset", "delete"}:
            tpv_editor_restore_user(before)
        else:
            raise ValueError("Эта операция пользователя не поддерживает откат.")

    elif row.entity_type == "question":
        if row.action == "create":
            question = db.session.get(Questions_tpv, int(row.entity_id))
            if question is not None:
                theme = question.flip
                db.session.delete(question)
                db.session.flush()
                tpv_editor_recalculate_theme(theme)
        elif row.action in {"update", "delete"}:
            old_theme = None
            current = db.session.get(Questions_tpv, int(row.entity_id))
            if current is not None:
                old_theme = current.flip
            restored = tpv_editor_restore_question(before)
            db.session.flush()
            if old_theme:
                tpv_editor_recalculate_theme(old_theme)
            tpv_editor_recalculate_theme(restored.flip)
        else:
            raise ValueError("Эта операция вопроса не поддерживает откат.")

    elif row.entity_type == "theme":
        if not isinstance(before, dict):
            raise ValueError("В истории отсутствует снимок темы.")

        for snapshot in before.get("questions", []):
            tpv_editor_restore_question(snapshot)
        for snapshot in before.get("users", []):
            tpv_editor_restore_user(snapshot)

        db.session.flush()
        for user in db.session.scalars(db.select(UsersTpv)).all():
            tpv_editor_update_approval(user)

    else:
        raise ValueError("Для этой операции откат недоступен.")

    row.reverted_at = _tpv_history_datetime.utcnow()


@app.get("/tpv_editor/api/history/status")
def tpv_editor_history_status():
    if not tpv_editor_allowed():
        return tpv_editor_error("Нет доступа к редактору.", 403)
    return jsonify({
        "ok": True,
        "table_exists": tpv_editor_history_table_exists(),
    })


@app.post("/tpv_editor/api/history/create-table")
def tpv_editor_history_create_table_route():
    if not tpv_editor_allowed():
        return tpv_editor_error("Нет доступа к редактору.", 403)
    tpv_editor_history_create_table()
    return jsonify({
        "ok": True,
        "message": "Таблица истории создана.",
    })


@app.get("/tpv_editor/api/history")
def tpv_editor_history_list():
    if not tpv_editor_allowed():
        return tpv_editor_error("Нет доступа к редактору.", 403)

    if not tpv_editor_history_table_exists():
        return jsonify({
            "ok": True,
            "table_exists": False,
            "items": [],
            "stats": {
                "total": 0,
                "today": 0,
                "revertible": 0,
                "reverted": 0,
            },
        })

    rows = db.session.scalars(
        db.select(TpvEditorHistory)
        .order_by(TpvEditorHistory.id.desc())
        .limit(1000)
    ).all()

    today = _tpv_history_datetime.utcnow().date()

    return jsonify({
        "ok": True,
        "table_exists": True,
        "items": [tpv_editor_history_to_dict(row) for row in rows],
        "stats": {
            "total": len(rows),
            "today": sum(row.created_at.date() == today for row in rows),
            "revertible": sum(
                row.can_revert and row.reverted_at is None
                for row in rows
            ),
            "reverted": sum(row.reverted_at is not None for row in rows),
        },
    })


@app.post("/tpv_editor/api/history/<int:history_id>/revert")
def tpv_editor_history_revert(history_id):
    if not tpv_editor_allowed():
        return tpv_editor_error("Нет доступа к редактору.", 403)

    if not tpv_editor_history_table_exists():
        return tpv_editor_error("Таблица истории не создана.", 409)

    row = db.session.get(TpvEditorHistory, history_id)

    if row is None:
        return tpv_editor_error("Запись истории не найдена.", 404)
    if not row.can_revert:
        return tpv_editor_error("Для этой операции откат недоступен.", 409)
    if row.reverted_at is not None:
        return tpv_editor_error("Операция уже была отменена.", 409)

    try:
        tpv_editor_revert_history_row(row)

        revert_row = tpv_editor_history_add(
            entity_type=row.entity_type,
            entity_id=row.entity_id,
            action="revert",
            title=f"Отмена операции: {row.title}",
            before=tpv_editor_history_parse(row.after_json),
            after=tpv_editor_history_parse(row.before_json),
            details=f"Отменена запись истории #{row.id}.",
            can_revert=False,
        )

        db.session.flush()
        if revert_row is not None:
            row.revert_history_id = revert_row.id
        db.session.commit()

        return jsonify({
            "ok": True,
            "message": "Изменение успешно отменено.",
        })
    except ValueError as exc:
        db.session.rollback()
        return tpv_editor_error(str(exc), 409)
    except Exception:
        db.session.rollback()
        raise


# ============================================================================
# TPV EDITOR — ЭТАП 8: КОНСТРУКТОР ИГРЫ
# ============================================================================
import random as _tpv_builder_random


class TpvGameBuild(db.Model):
    __tablename__ = "tpv_game_builds"

    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False)
    config_json = db.Column(db.Text, nullable=False, default="{}")
    question_ids_json = db.Column(db.Text, nullable=False, default="[]")
    is_active = db.Column(db.Boolean, nullable=False, default=False, index=True)
    created_at = db.Column(
        db.DateTime,
        nullable=False,
        default=_tpv_history_datetime.utcnow,
    )
    updated_at = db.Column(
        db.DateTime,
        nullable=False,
        default=_tpv_history_datetime.utcnow,
        onupdate=_tpv_history_datetime.utcnow,
        index=True,
    )


def tpv_editor_builder_table_exists():
    try:
        return "tpv_game_builds" in _tpv_history_inspect(db.engine).get_table_names()
    except Exception:
        return False


def tpv_editor_builder_create_table():
    TpvGameBuild.__table__.create(bind=db.engine, checkfirst=True)


def tpv_editor_builder_parse_json(value, default):
    if not value:
        return default
    try:
        result = _tpv_history_json.loads(value)
        return result if isinstance(result, type(default)) else default
    except (_tpv_history_json.JSONDecodeError, TypeError):
        return default


def tpv_editor_builder_config(row):
    return tpv_editor_builder_parse_json(row.config_json, {})


def tpv_editor_builder_question_ids(row):
    values = tpv_editor_builder_parse_json(row.question_ids_json, [])
    return [int(value) for value in values if str(value).isdigit()]


def tpv_editor_builder_normalize_payload(data):
    name = tpv_editor_normalize_text(data.get("name"))
    if not name:
        raise ValueError("Название набора обязательно.")
    if len(name) > 120:
        raise ValueError("Название набора должно содержать не более 120 символов.")

    try:
        limit = int(data.get("limit") or 30)
    except (TypeError, ValueError):
        raise ValueError("Количество вопросов должно быть целым числом.")

    if limit < 1 or limit > 1000:
        raise ValueError("Количество вопросов должно быть от 1 до 1000.")

    general_mode = str(data.get("general_mode") or "include")
    if general_mode not in {"include", "only", "exclude"}:
        raise ValueError("Некорректный режим общих вопросов.")

    themes = []
    seen_themes = set()
    for value in data.get("themes") or []:
        theme = tpv_editor_normalize_text(value)
        key = theme.casefold()
        if theme and not tpv_editor_is_general_theme(theme) and key not in seen_themes:
            seen_themes.add(key)
            themes.append(theme)

    excluded_authors = []
    seen_authors = set()
    for value in data.get("excluded_authors") or []:
        author = tpv_editor_normalize_text(value)
        key = author.casefold()
        if author and key not in seen_authors:
            seen_authors.add(key)
            excluded_authors.append(author)

    return {
        "name": name,
        "limit": limit,
        "general_mode": general_mode,
        "unused_only": bool(data.get("unused_only")),
        "randomize": bool(data.get("randomize", True)),
        "themes": themes,
        "excluded_authors": excluded_authors,
    }


def tpv_editor_builder_select_questions(config):
    questions = db.session.scalars(
        db.select(Questions_tpv).order_by(Questions_tpv.id)
    ).all()

    theme_keys = {
        tpv_editor_normalize_text(value).casefold()
        for value in config["themes"]
    }
    excluded_author_keys = {
        tpv_editor_normalize_text(value).casefold()
        for value in config["excluded_authors"]
    }

    result = []

    for question in questions:
        is_general = tpv_editor_is_general_theme(question.flip)
        theme_key = tpv_editor_normalize_text(question.flip).casefold()
        author_key = tpv_editor_normalize_text(question.author).casefold()

        if config["unused_only"] and str(question.show or "").casefold() == "true":
            continue

        if author_key in excluded_author_keys:
            continue

        if config["general_mode"] == "only" and not is_general:
            continue

        if config["general_mode"] == "exclude" and is_general:
            continue

        if (
            not is_general
            and theme_keys
            and theme_key not in theme_keys
        ):
            continue

        result.append(question)

    available_count = len(result)

    if config["randomize"]:
        _tpv_builder_random.SystemRandom().shuffle(result)

    return result[: config["limit"]], available_count


def tpv_editor_builder_question_dict(question):
    data = tpv_editor_question_to_dict(question)
    return {
        "id": data["id"],
        "task": data["task"],
        "author": data["author"],
        "flip_display": data["flip_display"],
        "is_general": data["is_general"],
        "show": data["show"],
    }


def tpv_editor_builder_to_dict(row, include_questions=True):
    config = tpv_editor_builder_config(row)
    question_ids = tpv_editor_builder_question_ids(row)
    questions = []

    if include_questions and question_ids:
        found = db.session.scalars(
            db.select(Questions_tpv).where(Questions_tpv.id.in_(question_ids))
        ).all()
        by_id = {item.id: item for item in found}
        questions = [
            tpv_editor_builder_question_dict(by_id[question_id])
            for question_id in question_ids
            if question_id in by_id
        ]

    return {
        "id": row.id,
        "name": row.name,
        "limit": int(config.get("limit") or len(question_ids) or 30),
        "general_mode": config.get("general_mode") or "include",
        "unused_only": bool(config.get("unused_only")),
        "randomize": bool(config.get("randomize", True)),
        "themes": config.get("themes") or [],
        "excluded_authors": config.get("excluded_authors") or [],
        "question_ids": question_ids,
        "question_count": len(question_ids),
        "questions": questions,
        "is_active": bool(row.is_active),
        "created_at": row.created_at.isoformat(timespec="seconds"),
        "updated_at": row.updated_at.isoformat(timespec="seconds"),
        "updated_at_label": row.updated_at.strftime("%d.%m.%Y %H:%M"),
    }


@app.get("/tpv_editor/api/game-builder")
def tpv_editor_builder_list():
    if not tpv_editor_allowed():
        return tpv_editor_error("Нет доступа к редактору.", 403)

    themes = tpv_editor_theme_list()
    authors = [
        value
        for value in db.session.scalars(
            db.select(Questions_tpv.author)
            .where(Questions_tpv.author.is_not(None))
            .distinct()
        ).all()
        if tpv_editor_normalize_text(value)
    ]
    authors = sorted(
        {tpv_editor_normalize_text(value) for value in authors},
        key=str.casefold,
    )

    unused_questions = int(db.session.scalar(
        db.select(func.count(Questions_tpv.id)).where(
            func.lower(Questions_tpv.show) != "true"
        )
    ) or 0)

    if not tpv_editor_builder_table_exists():
        return jsonify({
            "ok": True,
            "table_exists": False,
            "items": [],
            "themes": themes,
            "authors": authors,
            "stats": {
                "total": 0,
                "active_name": None,
                "unused_questions": unused_questions,
            },
        })

    rows = db.session.scalars(
        db.select(TpvGameBuild)
        .order_by(TpvGameBuild.is_active.desc(), TpvGameBuild.updated_at.desc())
    ).all()

    active = next((row for row in rows if row.is_active), None)

    return jsonify({
        "ok": True,
        "table_exists": True,
        "items": [tpv_editor_builder_to_dict(row) for row in rows],
        "themes": themes,
        "authors": authors,
        "stats": {
            "total": len(rows),
            "active_name": active.name if active else None,
            "unused_questions": unused_questions,
        },
    })


@app.post("/tpv_editor/api/game-builder/create-table")
def tpv_editor_builder_create_table_route():
    if not tpv_editor_allowed():
        return tpv_editor_error("Нет доступа к редактору.", 403)

    tpv_editor_builder_create_table()
    return jsonify({
        "ok": True,
        "message": "Таблица конструктора игры создана.",
    })


@app.post("/tpv_editor/api/game-builder/preview")
def tpv_editor_builder_preview():
    if not tpv_editor_allowed():
        return tpv_editor_error("Нет доступа к редактору.", 403)

    data = request.get_json(silent=True) or {}

    try:
        config = tpv_editor_builder_normalize_payload(data)
    except ValueError as exc:
        return tpv_editor_error(str(exc))

    questions, available_count = tpv_editor_builder_select_questions(config)

    return jsonify({
        "ok": True,
        "available_count": available_count,
        "questions": [
            tpv_editor_builder_question_dict(question)
            for question in questions
        ],
    })


@app.post("/tpv_editor/api/game-builder")
def tpv_editor_builder_create():
    if not tpv_editor_allowed():
        return tpv_editor_error("Нет доступа к редактору.", 403)

    if not tpv_editor_builder_table_exists():
        return tpv_editor_error("Таблица конструктора не создана.", 409)

    data = request.get_json(silent=True) or {}

    try:
        config = tpv_editor_builder_normalize_payload(data)
    except ValueError as exc:
        return tpv_editor_error(str(exc))

    question_ids = [
        int(value)
        for value in data.get("question_ids") or []
        if str(value).isdigit()
    ]

    if not question_ids:
        return tpv_editor_error("Сначала сформируйте выборку вопросов.")

    row = TpvGameBuild(
        name=config["name"],
        config_json=tpv_editor_history_json(config),
        question_ids_json=tpv_editor_history_json(question_ids),
        is_active=False,
    )
    db.session.add(row)
    db.session.flush()

    tpv_editor_history_add(
        "bulk",
        row.id,
        "create",
        f"Создан игровой набор «{row.name}»",
        after=tpv_editor_builder_to_dict(row),
        details=f"Вопросов в наборе: {len(question_ids)}.",
        can_revert=False,
    )

    db.session.commit()

    return jsonify({
        "ok": True,
        "message": "Игровой набор сохранён.",
        "item": tpv_editor_builder_to_dict(row),
    }), 201


@app.put("/tpv_editor/api/game-builder/<int:build_id>")
def tpv_editor_builder_update(build_id):
    if not tpv_editor_allowed():
        return tpv_editor_error("Нет доступа к редактору.", 403)

    row = db.session.get(TpvGameBuild, build_id)
    if row is None:
        return tpv_editor_error("Игровой набор не найден.", 404)

    data = request.get_json(silent=True) or {}

    try:
        config = tpv_editor_builder_normalize_payload(data)
    except ValueError as exc:
        return tpv_editor_error(str(exc))

    question_ids = [
        int(value)
        for value in data.get("question_ids") or []
        if str(value).isdigit()
    ]

    if not question_ids:
        return tpv_editor_error("Сначала сформируйте выборку вопросов.")

    before = tpv_editor_builder_to_dict(row)

    row.name = config["name"]
    row.config_json = tpv_editor_history_json(config)
    row.question_ids_json = tpv_editor_history_json(question_ids)
    row.updated_at = _tpv_history_datetime.utcnow()

    db.session.flush()

    tpv_editor_history_add(
        "bulk",
        row.id,
        "update",
        f"Обновлён игровой набор «{row.name}»",
        before=before,
        after=tpv_editor_builder_to_dict(row),
        can_revert=False,
    )

    db.session.commit()

    return jsonify({
        "ok": True,
        "message": "Игровой набор обновлён.",
        "item": tpv_editor_builder_to_dict(row),
    })


@app.delete("/tpv_editor/api/game-builder/<int:build_id>")
def tpv_editor_builder_delete(build_id):
    if not tpv_editor_allowed():
        return tpv_editor_error("Нет доступа к редактору.", 403)

    row = db.session.get(TpvGameBuild, build_id)
    if row is None:
        return tpv_editor_error("Игровой набор не найден.", 404)

    snapshot = tpv_editor_builder_to_dict(row)
    name = row.name
    db.session.delete(row)

    tpv_editor_history_add(
        "bulk",
        build_id,
        "delete",
        f"Удалён игровой набор «{name}»",
        before=snapshot,
        can_revert=False,
    )

    db.session.commit()

    return jsonify({
        "ok": True,
        "message": "Игровой набор удалён. Вопросы из базы не изменены.",
    })


@app.post("/tpv_editor/api/game-builder/<int:build_id>/activate")
def tpv_editor_builder_activate(build_id):
    if not tpv_editor_allowed():
        return tpv_editor_error("Нет доступа к редактору.", 403)

    row = db.session.get(TpvGameBuild, build_id)
    if row is None:
        return tpv_editor_error("Игровой набор не найден.", 404)

    active_rows = db.session.scalars(
        db.select(TpvGameBuild).where(TpvGameBuild.is_active.is_(True))
    ).all()

    for active in active_rows:
        active.is_active = False

    row.is_active = True
    row.updated_at = _tpv_history_datetime.utcnow()

    tpv_editor_history_add(
        "bulk",
        row.id,
        "update",
        f"Активирован игровой набор «{row.name}»",
        details=f"Вопросов в наборе: {len(tpv_editor_builder_question_ids(row))}.",
        can_revert=False,
    )

    db.session.commit()

    return jsonify({
        "ok": True,
        "message": f"Набор «{row.name}» сделан активным.",
    })


@app.get("/tpv_editor/api/game-builder/active")
def tpv_editor_builder_active():
    if not tpv_editor_builder_table_exists():
        return jsonify({
            "ok": True,
            "active": None,
        })

    row = db.session.scalar(
        db.select(TpvGameBuild)
        .where(TpvGameBuild.is_active.is_(True))
        .order_by(TpvGameBuild.updated_at.desc())
        .limit(1)
    )

    return jsonify({
        "ok": True,
        "active": tpv_editor_builder_to_dict(row) if row else None,
    })



# ============================================================================
# TPV EDITOR — ЭТАП 9: ЗАЯВКИ НА ВОПРОСЫ
# ============================================================================


class TpvQuestionApplication(db.Model):
    __tablename__ = "tpv_question_applications"

    id = db.Column(db.Integer, primary_key=True)
    author = db.Column(db.String(100), nullable=False, index=True)
    task = db.Column(db.Text, nullable=False)
    answer = db.Column(db.Text, nullable=False)
    comment = db.Column(db.Text, nullable=True)
    flip = db.Column(db.String(200), nullable=False, default="false", index=True)
    status = db.Column(db.String(16), nullable=False, default="pending", index=True)
    reject_reason = db.Column(db.Text, nullable=True)
    created_at = db.Column(
        db.DateTime,
        nullable=False,
        default=_tpv_history_datetime.utcnow,
        index=True,
    )
    reviewed_at = db.Column(db.DateTime, nullable=True)
    reviewed_by = db.Column(db.String(100), nullable=True)
    question_id = db.Column(db.Integer, nullable=True, index=True)


def tpv_editor_applications_table_exists():
    try:
        return (
            "tpv_question_applications"
            in _tpv_history_inspect(db.engine).get_table_names()
        )
    except Exception:
        return False


def tpv_editor_applications_create_table():
    TpvQuestionApplication.__table__.create(
        bind=db.engine,
        checkfirst=True,
    )


def tpv_editor_application_validate(data):
    author = tpv_editor_normalize_text(data.get("author"))
    task = tpv_editor_normalize_text(data.get("task"))
    answer = tpv_editor_normalize_text(data.get("answer"))
    comment = tpv_editor_normalize_text(data.get("comment"))
    flip = tpv_editor_normalize_theme(data.get("flip"))

    if not author:
        raise ValueError("Укажите имя автора.")
    if len(author) > 100:
        raise ValueError("Имя автора слишком длинное.")
    if not task:
        raise ValueError("Введите текст вопроса.")
    if not answer:
        raise ValueError("Введите ответ.")
    if len(task) > 3000:
        raise ValueError("Текст вопроса слишком длинный.")
    if len(answer) > 2000:
        raise ValueError("Ответ слишком длинный.")
    if len(comment) > 3000:
        raise ValueError("Комментарий слишком длинный.")

    if not tpv_editor_is_general_theme(flip):
        available_theme_values = db.session.scalars(
            db.select(Questions_tpv.flip)
            .where(Questions_tpv.flip.is_not(None))
            .distinct()
        ).all()

        available_themes = {
            tpv_editor_normalize_text(value).casefold()
            for value in available_theme_values
            if (
                tpv_editor_normalize_text(value)
                and not tpv_editor_is_general_theme(value)
            )
        }

        if tpv_editor_normalize_text(flip).casefold() not in available_themes:
            raise ValueError("Выбранная тема замены отсутствует в базе.")

    return {
        "author": author,
        "task": task,
        "answer": answer,
        "comment": comment,
        "flip": flip,
    }


def tpv_editor_application_to_dict(row):
    is_general = tpv_editor_is_general_theme(row.flip)
    author_exists = db.session.scalar(
        db.select(UsersTpv.id).where(
            func.lower(UsersTpv.username) == row.author.casefold()
        )
    ) is not None

    status_labels = {
        "pending": "На рассмотрении",
        "approved": "Утверждена",
        "rejected": "Отклонена",
    }

    return {
        "id": row.id,
        "author": row.author,
        "task": row.task,
        "answer": row.answer,
        "comment": row.comment or "",
        "flip": row.flip,
        "flip_display": "Общий" if is_general else row.flip,
        "is_general": is_general,
        "status": row.status,
        "status_label": status_labels.get(row.status, row.status),
        "reject_reason": row.reject_reason or "",
        "created_at": row.created_at.isoformat(timespec="seconds"),
        "created_at_label": row.created_at.strftime("%d.%m.%Y %H:%M"),
        "reviewed_at": (
            row.reviewed_at.isoformat(timespec="seconds")
            if row.reviewed_at else None
        ),
        "reviewed_by": row.reviewed_by or "",
        "question_id": row.question_id,
        "author_exists": author_exists,
    }


@app.get("/tpv_questions")
def tpv_question_application_page():
    return render_template("tpv-question-application.html")


@app.get("/api/tpv-question-applications/status")
def tpv_question_application_status():
    theme_values = db.session.scalars(
        db.select(Questions_tpv.flip)
        .where(Questions_tpv.flip.is_not(None))
        .distinct()
    ).all()

    themes = sorted(
        {
            tpv_editor_normalize_text(value)
            for value in theme_values
            if (
                tpv_editor_normalize_text(value)
                and not tpv_editor_is_general_theme(value)
            )
        },
        key=str.casefold,
    )

    return jsonify({
        "ok": True,
        "table_exists": tpv_editor_applications_table_exists(),
        "themes": themes,
    })


@app.post("/api/tpv-question-applications")
def tpv_question_application_submit():
    if not tpv_editor_applications_table_exists():
        return tpv_editor_error(
            "Приём заявок временно недоступен.",
            503,
        )

    data = request.get_json(silent=True) or {}

    try:
        values = tpv_editor_application_validate(data)
    except ValueError as exc:
        return tpv_editor_error(str(exc))

    duplicate = db.session.scalar(
        db.select(TpvQuestionApplication).where(
            func.lower(TpvQuestionApplication.task)
            == values["task"].casefold(),
            TpvQuestionApplication.status == "pending",
        )
    )

    if duplicate is not None:
        return tpv_editor_error(
            f"Похожая заявка уже ожидает модерации под номером {duplicate.id}.",
            409,
        )

    row = TpvQuestionApplication(
        author=values["author"],
        task=values["task"],
        answer=values["answer"],
        comment=values["comment"],
        flip=values["flip"],
        status="pending",
    )
    db.session.add(row)
    db.session.commit()

    return jsonify({
        "ok": True,
        "message": "Заявка отправлена на модерацию.",
        "application_id": row.id,
    }), 201


@app.get("/tpv_editor/api/question-applications")
def tpv_editor_application_list():
    if not tpv_editor_allowed():
        return tpv_editor_error("Нет доступа к редактору.", 403)

    if not tpv_editor_applications_table_exists():
        return jsonify({
            "ok": True,
            "table_exists": False,
            "items": [],
            "stats": {
                "total": 0,
                "pending": 0,
                "approved": 0,
                "rejected": 0,
            },
        })

    rows = db.session.scalars(
        db.select(TpvQuestionApplication)
        .order_by(TpvQuestionApplication.id.desc())
        .limit(2000)
    ).all()

    return jsonify({
        "ok": True,
        "table_exists": True,
        "items": [tpv_editor_application_to_dict(row) for row in rows],
        "stats": {
            "total": len(rows),
            "pending": sum(row.status == "pending" for row in rows),
            "approved": sum(row.status == "approved" for row in rows),
            "rejected": sum(row.status == "rejected" for row in rows),
        },
    })


@app.post("/tpv_editor/api/question-applications/create-table")
def tpv_editor_application_create_table_route():
    if not tpv_editor_allowed():
        return tpv_editor_error("Нет доступа к редактору.", 403)

    tpv_editor_applications_create_table()

    return jsonify({
        "ok": True,
        "message": "Таблица заявок создана.",
    })


@app.post("/tpv_editor/api/question-applications/<int:application_id>/approve")
def tpv_editor_application_approve(application_id):
    if not tpv_editor_allowed():
        return tpv_editor_error("Нет доступа к редактору.", 403)

    row = db.session.get(TpvQuestionApplication, application_id)

    if row is None:
        return tpv_editor_error("Заявка не найдена.", 404)
    if row.status != "pending":
        return tpv_editor_error("Заявка уже обработана.", 409)

    data = request.get_json(silent=True) or {}

    try:
        values = tpv_editor_application_validate(data)
    except ValueError as exc:
        return tpv_editor_error(str(exc))

    existing_question = db.session.scalar(
        db.select(Questions_tpv).where(
            func.lower(Questions_tpv.task)
            == values["task"].casefold()
        )
    )

    if existing_question is not None:
        return tpv_editor_error(
            f"Вопрос с такой формулировкой уже существует: #{existing_question.id}.",
            409,
        )

    user = db.session.scalar(
        db.select(UsersTpv).where(
            func.lower(UsersTpv.username)
            == values["author"].casefold()
        )
    )

    if user is None and bool(data.get("create_user")):
        if len(values["author"]) > 10:
            return tpv_editor_error(
                "Нельзя создать пользователя: текущая модель допускает имя до 10 символов.",
                409,
            )

        user = UsersTpv()
        user.username = values["author"]
        user.money = 0
        user.flip = "false"
        user.flip_col = 0
        user.approve = "false"
        db.session.add(user)
        db.session.flush()

        tpv_editor_history_add(
            "user",
            user.id,
            "create",
            f"Создан пользователь «{user.username}» при утверждении заявки",
            after=tpv_editor_user_snapshot(user),
            can_revert=True,
        )

    question = Questions_tpv()
    question.task = values["task"]
    question.answer = values["answer"]
    question.comment = values["comment"]
    question.author = values["author"]
    question.flip = values["flip"]
    question.show = "false"

    db.session.add(question)
    db.session.flush()

    row.author = values["author"]
    row.task = values["task"]
    row.answer = values["answer"]
    row.comment = values["comment"]
    row.flip = values["flip"]
    row.status = "approved"
    row.reject_reason = ""
    row.reviewed_at = _tpv_history_datetime.utcnow()
    row.reviewed_by = "TPV Editor"
    row.question_id = question.id

    tpv_editor_recalculate_theme(question.flip)

    tpv_editor_history_add(
        "question",
        question.id,
        "create",
        f"Утверждена заявка #{row.id}; создан вопрос #{question.id}",
        after=tpv_editor_question_snapshot(question),
        details=f"Автор заявки: {row.author}.",
        can_revert=True,
    )

    db.session.commit()

    return jsonify({
        "ok": True,
        "message": f"Заявка утверждена. Создан вопрос #{question.id}.",
        "question_id": question.id,
    })


@app.post("/tpv_editor/api/question-applications/<int:application_id>/reject")
def tpv_editor_application_reject(application_id):
    if not tpv_editor_allowed():
        return tpv_editor_error("Нет доступа к редактору.", 403)

    row = db.session.get(TpvQuestionApplication, application_id)

    if row is None:
        return tpv_editor_error("Заявка не найдена.", 404)
    if row.status != "pending":
        return tpv_editor_error("Заявка уже обработана.", 409)

    data = request.get_json(silent=True) or {}
    reason = tpv_editor_normalize_text(data.get("reject_reason"))

    if not reason:
        return tpv_editor_error("Укажите причину отклонения.")

    row.status = "rejected"
    row.reject_reason = reason
    row.reviewed_at = _tpv_history_datetime.utcnow()
    row.reviewed_by = "TPV Editor"

    tpv_editor_history_add(
        "bulk",
        row.id,
        "update",
        f"Отклонена заявка #{row.id}",
        details=f"Причина: {reason}",
        can_revert=False,
    )

    db.session.commit()

    return jsonify({
        "ok": True,
        "message": "Заявка отклонена.",
    })


@app.delete("/tpv_editor/api/question-applications/<int:application_id>")
def tpv_editor_application_delete(application_id):
    if not tpv_editor_allowed():
        return tpv_editor_error("Нет доступа к редактору.", 403)

    row = db.session.get(TpvQuestionApplication, application_id)

    if row is None:
        return tpv_editor_error("Заявка не найдена.", 404)

    snapshot = tpv_editor_application_to_dict(row)
    db.session.delete(row)

    tpv_editor_history_add(
        "bulk",
        application_id,
        "delete",
        f"Удалена заявка #{application_id}",
        before=snapshot,
        can_revert=False,
    )

    db.session.commit()

    return jsonify({
        "ok": True,
        "message": "Заявка удалена.",
    })



def tpv_editor_normalize_text(value):
    return " ".join(str(value or "").strip().split())


def tpv_editor_is_general_theme(value):
    return tpv_editor_normalize_text(value).casefold() in TPV_GENERAL_QUESTION_VALUES


def tpv_editor_normalize_theme(value):
    theme = tpv_editor_normalize_text(value)
    return "false" if tpv_editor_is_general_theme(theme) else theme


def tpv_editor_theme_key(value):
    """Ключ темы с поддержкой кириллицы, регистра и лишних пробелов."""
    return tpv_editor_normalize_text(value).casefold()


def tpv_editor_matching_questions(theme):
    """Возвращает вопросы темы. Сравнение выполняется в Python, не SQLite lower()."""
    key = tpv_editor_theme_key(theme)
    if not key or tpv_editor_is_general_theme(theme):
        return []

    questions = db.session.scalars(
        db.select(Questions_tpv).where(Questions_tpv.flip.is_not(None))
    ).all()

    return [
        question
        for question in questions
        if tpv_editor_theme_key(question.flip) == key
    ]


def tpv_editor_matching_users(theme):
    """Возвращает пользователей темы с корректной обработкой кириллицы."""
    key = tpv_editor_theme_key(theme)
    if not key or tpv_editor_is_general_theme(theme):
        return []

    users = db.session.scalars(
        db.select(UsersTpv).where(UsersTpv.flip.is_not(None))
    ).all()

    return [
        user
        for user in users
        if tpv_editor_theme_key(user.flip) == key
    ]


def tpv_editor_count_questions(theme):
    theme = tpv_editor_normalize_theme(theme)
    if tpv_editor_is_general_theme(theme):
        return 0
    return len(tpv_editor_matching_questions(theme))


def tpv_editor_update_approval(user):
    user.flip = tpv_editor_normalize_theme(user.flip)
    if tpv_editor_is_general_theme(user.flip):
        user.flip_col = 0
        user.approve = "false"
        return
    user.flip_col = tpv_editor_count_questions(user.flip)
    user.approve = "true" if user.flip_col >= TPV_REQUIRED_FLIP_QUESTIONS else "false"


def tpv_editor_author_question_count(username):
    return int(db.session.scalar(
        db.select(func.count(Questions_tpv.id)).where(Questions_tpv.author == username)
    ) or 0)


def tpv_editor_user_to_dict(user):
    flip_display = "" if tpv_editor_is_general_theme(user.flip) else (user.flip or "")
    approved = str(user.approve).lower() == "true"
    if not flip_display:
        label = "Тема не выбрана"
    elif approved:
        label = "Допущен"
    else:
        label = f"Недостаточно вопросов: {int(user.flip_col or 0)}/{TPV_REQUIRED_FLIP_QUESTIONS}"
    return {
        "id": user.id,
        "username": user.username,
        "money": int(user.money or 0),
        "flip": user.flip or "false",
        "flip_display": flip_display,
        "flip_col": int(user.flip_col or 0),
        "approve": "true" if approved else "false",
        "approve_label": label,
        "authored_questions": tpv_editor_author_question_count(user.username),
    }


def tpv_editor_allowed():
    return session.get("tpv_launcher_allowed") is True and session.get("tpv_role") == "editor"


def tpv_editor_error(message, status=400):
    return jsonify({"ok": False, "message": message}), status


@app.route("/tpv_editor", methods=["POST", "GET"])
def tpv_editor():
    if request.method == "POST":
        if session.get("tpv_launcher_allowed") is not True:
            return redirect(url_for("tpv"))
        session["tpv_role"] = "editor"
        return render_template("tpv-editor.html")
    if tpv_editor_allowed():
        return render_template("tpv-editor.html")
    return redirect(url_for("tpv"))


@app.get("/tpv_editor/api/users")
def tpv_editor_get_users():
    if not tpv_editor_allowed():
        return tpv_editor_error("Нет доступа к редактору.", 403)
    users = db.session.scalars(db.select(UsersTpv).order_by(UsersTpv.id)).all()
    return jsonify({"ok": True, "users": [tpv_editor_user_to_dict(u) for u in users]})


def tpv_editor_theme_list():
    """Возвращает список уникальных тематических значений Questions_tpv.flip."""
    raw = db.session.scalars(
        db.select(Questions_tpv.flip)
        .where(Questions_tpv.flip.is_not(None))
        .distinct()
    ).all()

    unique = {}

    for value in raw:
        theme = tpv_editor_normalize_text(value)

        if theme and not tpv_editor_is_general_theme(theme):
            unique.setdefault(theme.casefold(), theme)

    return sorted(unique.values(), key=str.casefold)


@app.get("/tpv_editor/api/themes")
def tpv_editor_get_themes():
    if not tpv_editor_allowed():
        return tpv_editor_error("Нет доступа к редактору.", 403)

    return jsonify({
        "ok": True,
        "themes": tpv_editor_theme_list(),
    })


@app.post("/tpv_editor/api/users")
def tpv_editor_create_user():
    if not tpv_editor_allowed():
        return tpv_editor_error("Нет доступа к редактору.", 403)
    data = request.get_json(silent=True) or {}
    username = tpv_editor_normalize_text(data.get("username"))
    if not username:
        return tpv_editor_error("Имя пользователя обязательно.")
    if len(username) > 10:
        return tpv_editor_error("Имя должно содержать не более 10 символов.")
    duplicate = db.session.scalar(db.select(UsersTpv).where(func.lower(UsersTpv.username) == username.casefold()))
    if duplicate:
        return tpv_editor_error("Пользователь с таким именем уже существует.", 409)
    try:
        money = int(data.get("money", 0) or 0)
    except (TypeError, ValueError):
        return tpv_editor_error("Баланс должен быть целым числом.")
    user = UsersTpv(username=username, money=money, flip=tpv_editor_normalize_theme(data.get("flip")), flip_col=0, approve="false")
    tpv_editor_update_approval(user)
    db.session.add(user)
    db.session.flush()
    tpv_editor_history_add(
        "user",
        user.id,
        "create",
        f"Создан пользователь «{user.username}»",
        after=tpv_editor_user_snapshot(user),
        can_revert=True,
    )
    db.session.commit()
    return jsonify({"ok": True, "message": "Пользователь создан.", "user": tpv_editor_user_to_dict(user)}), 201


@app.put("/tpv_editor/api/users/<int:user_id>")
def tpv_editor_update_user(user_id):
    if not tpv_editor_allowed():
        return tpv_editor_error("Нет доступа к редактору.", 403)
    user = db.session.get(UsersTpv, user_id)
    if not user:
        return tpv_editor_error("Пользователь не найден.", 404)
    data = request.get_json(silent=True) or {}
    username = tpv_editor_normalize_text(data.get("username"))
    if not username or len(username) > 10:
        return tpv_editor_error("Имя обязательно и должно быть не длиннее 10 символов.")
    duplicate = db.session.scalar(db.select(UsersTpv).where(func.lower(UsersTpv.username) == username.casefold(), UsersTpv.id != user_id))
    if duplicate:
        return tpv_editor_error("Пользователь с таким именем уже существует.", 409)
    try:
        money = int(data.get("money", 0) or 0)
    except (TypeError, ValueError):
        return tpv_editor_error("Баланс должен быть целым числом.")
    before_snapshot = tpv_editor_user_snapshot(user)
    old_username = user.username
    user.username = username
    user.money = money
    user.flip = tpv_editor_normalize_theme(data.get("flip"))
    tpv_editor_update_approval(user)
    tpv_editor_history_add(
        "user",
        user.id,
        "update",
        f"Изменён пользователь «{user.username}»",
        before=before_snapshot,
        after=tpv_editor_user_snapshot(user),
        can_revert=True,
    )
    db.session.commit()
    message = "Изменения сохранены."
    if old_username != username and tpv_editor_author_question_count(old_username):
        message += " Автор в существующих вопросах не переименован."
    return jsonify({"ok": True, "message": message, "user": tpv_editor_user_to_dict(user)})


@app.delete("/tpv_editor/api/users/<int:user_id>")
def tpv_editor_delete_user(user_id):
    if not tpv_editor_allowed():
        return tpv_editor_error("Нет доступа к редактору.", 403)
    user = db.session.get(UsersTpv, user_id)
    if not user:
        return tpv_editor_error("Пользователь не найден.", 404)
    before_snapshot = tpv_editor_user_snapshot(user)
    username = user.username
    count = tpv_editor_author_question_count(username)
    db.session.delete(user)
    tpv_editor_history_add(
        "user",
        user_id,
        "delete",
        f"Удалён пользователь «{username}»",
        before=before_snapshot,
        details=f"Связанных вопросов автора: {count}.",
        can_revert=True,
    )
    db.session.commit()
    message = f"Пользователь «{username}» удалён."
    if count:
        message += f" Его вопросы ({count}) сохранены."
    return jsonify({"ok": True, "message": message})


@app.post("/tpv_editor/api/users/<int:user_id>/reset-money")
def tpv_editor_reset_money(user_id):
    if not tpv_editor_allowed():
        return tpv_editor_error("Нет доступа к редактору.", 403)
    user = db.session.get(UsersTpv, user_id)
    if not user:
        return tpv_editor_error("Пользователь не найден.", 404)
    before_snapshot = tpv_editor_user_snapshot(user)
    user.money = 0
    tpv_editor_history_add(
        "user",
        user.id,
        "reset",
        f"Обнулён баланс пользователя «{user.username}»",
        before=before_snapshot,
        after=tpv_editor_user_snapshot(user),
        can_revert=True,
    )
    db.session.commit()
    return jsonify({"ok": True, "message": f"Баланс пользователя «{user.username}» обнулён."})


@app.post("/tpv_editor/api/users/recalculate-all")
def tpv_editor_recalculate_all():
    if not tpv_editor_allowed():
        return tpv_editor_error("Нет доступа к редактору.", 403)
    users = db.session.scalars(db.select(UsersTpv)).all()
    for user in users:
        tpv_editor_update_approval(user)
    db.session.commit()
    return jsonify({"ok": True, "message": f"Пересчитано пользователей: {len(users)}."})


# ============================================================================
# TPV EDITOR — ЭТАП 2: ВОПРОСЫ
# Добавить после блока первого этапа.
# ============================================================================

def tpv_editor_normalize_question_theme(value):
    theme = tpv_editor_normalize_text(value)
    return "false" if tpv_editor_is_general_theme(theme) else theme


def tpv_editor_question_to_dict(question):
    is_general = tpv_editor_is_general_theme(question.flip)
    return {
        "id": question.id,
        "task": question.task or "",
        "answer": question.answer or "",
        "comment": question.comment or "",
        "author": question.author or "",
        "flip": question.flip or "false",
        "flip_display": "общий" if is_general else question.flip,
        "is_general": is_general,
        "show": "true" if str(question.show).lower() == "true" else "false",
    }


def tpv_editor_recalculate_theme(theme):
    if tpv_editor_is_general_theme(theme):
        return 0

    users = tpv_editor_matching_users(theme)

    for user in users:
        tpv_editor_update_approval(user)

    return len(users)


def tpv_editor_question_duplicate(task, exclude_id=None):
    normalized = tpv_editor_normalize_text(task).casefold()
    if not normalized:
        return None
    query = db.select(Questions_tpv).where(
        func.lower(func.trim(Questions_tpv.task)) == normalized
    )
    if exclude_id is not None:
        query = query.where(Questions_tpv.id != exclude_id)
    return db.session.scalar(query.limit(1))


def tpv_editor_question_payload(data):
    task = tpv_editor_normalize_text(data.get("task"))
    answer = tpv_editor_normalize_text(data.get("answer"))
    author = tpv_editor_normalize_text(data.get("author"))
    comment = str(data.get("comment") or "").strip()
    flip = tpv_editor_normalize_question_theme(data.get("flip"))
    show = "true" if str(data.get("show")).lower() == "true" else "false"
    if not task:
        raise ValueError("Текст вопроса обязателен.")
    if not answer:
        raise ValueError("Ответ обязателен.")
    if not author:
        raise ValueError("Автор обязателен.")
    return task, answer, comment, author, flip, show


@app.get("/tpv_editor/api/questions")
def tpv_editor_get_questions():
    if not tpv_editor_allowed():
        return tpv_editor_error("Нет доступа к редактору.", 403)
    questions = db.session.scalars(
        db.select(Questions_tpv).order_by(Questions_tpv.id)
    ).all()
    raw_authors = db.session.scalars(
        db.select(Questions_tpv.author).where(Questions_tpv.author.is_not(None)).distinct()
    ).all()
    raw_themes = db.session.scalars(
        db.select(Questions_tpv.flip).where(Questions_tpv.flip.is_not(None)).distinct()
    ).all()
    authors = sorted({tpv_editor_normalize_text(x) for x in raw_authors if tpv_editor_normalize_text(x)}, key=str.casefold)
    themes = sorted({tpv_editor_normalize_text(x) for x in raw_themes if not tpv_editor_is_general_theme(x)}, key=str.casefold)
    return jsonify({"ok": True, "questions": [tpv_editor_question_to_dict(q) for q in questions], "authors": authors, "themes": themes})


@app.post("/tpv_editor/api/questions")
def tpv_editor_create_question():
    if not tpv_editor_allowed():
        return tpv_editor_error("Нет доступа к редактору.", 403)
    data = request.get_json(silent=True) or {}
    try:
        task, answer, comment, author, flip, show = tpv_editor_question_payload(data)
    except ValueError as exc:
        return tpv_editor_error(str(exc))
    duplicate = tpv_editor_question_duplicate(task)
    if duplicate:
        return tpv_editor_error(f"Такой вопрос уже существует: ID {duplicate.id}.", 409)
    question = Questions_tpv(task=task, answer=answer, comment=comment, author=author, flip=flip, show=show)
    db.session.add(question)
    db.session.flush()
    tpv_editor_recalculate_theme(flip)
    tpv_editor_history_add(
        "question",
        question.id,
        "create",
        f"Создан вопрос #{question.id}",
        after=tpv_editor_question_snapshot(question),
        can_revert=True,
    )
    db.session.commit()
    return jsonify({"ok": True, "message": "Вопрос создан.", "question": tpv_editor_question_to_dict(question)}), 201


@app.put("/tpv_editor/api/questions/<int:question_id>")
def tpv_editor_update_question(question_id):
    if not tpv_editor_allowed():
        return tpv_editor_error("Нет доступа к редактору.", 403)
    question = db.session.get(Questions_tpv, question_id)
    if not question:
        return tpv_editor_error("Вопрос не найден.", 404)
    data = request.get_json(silent=True) or {}
    try:
        task, answer, comment, author, flip, show = tpv_editor_question_payload(data)
    except ValueError as exc:
        return tpv_editor_error(str(exc))
    duplicate = tpv_editor_question_duplicate(task, exclude_id=question_id)
    if duplicate:
        return tpv_editor_error(f"Такой вопрос уже существует: ID {duplicate.id}.", 409)
    before_snapshot = tpv_editor_question_snapshot(question)
    old_theme = question.flip
    question.task, question.answer, question.comment = task, answer, comment
    question.author, question.flip, question.show = author, flip, show
    db.session.flush()
    tpv_editor_recalculate_theme(old_theme)
    if tpv_editor_normalize_text(old_theme).casefold() != tpv_editor_normalize_text(flip).casefold():
        tpv_editor_recalculate_theme(flip)
    tpv_editor_history_add(
        "question",
        question.id,
        "update",
        f"Изменён вопрос #{question.id}",
        before=before_snapshot,
        after=tpv_editor_question_snapshot(question),
        can_revert=True,
    )
    db.session.commit()
    return jsonify({"ok": True, "message": "Вопрос сохранён.", "question": tpv_editor_question_to_dict(question)})


@app.delete("/tpv_editor/api/questions/<int:question_id>")
def tpv_editor_delete_question(question_id):
    if not tpv_editor_allowed():
        return tpv_editor_error("Нет доступа к редактору.", 403)
    question = db.session.get(Questions_tpv, question_id)
    if not question:
        return tpv_editor_error("Вопрос не найден.", 404)
    before_snapshot = tpv_editor_question_snapshot(question)
    theme = question.flip
    db.session.delete(question)
    db.session.flush()
    tpv_editor_recalculate_theme(theme)
    tpv_editor_history_add(
        "question",
        question_id,
        "delete",
        f"Удалён вопрос #{question_id}",
        before=before_snapshot,
        can_revert=True,
    )
    db.session.commit()
    return jsonify({"ok": True, "message": f"Вопрос #{question_id} удалён."})


@app.post("/tpv_editor/api/questions/<int:question_id>/duplicate")
def tpv_editor_duplicate_question(question_id):
    if not tpv_editor_allowed():
        return tpv_editor_error("Нет доступа к редактору.", 403)
    source = db.session.get(Questions_tpv, question_id)
    if not source:
        return tpv_editor_error("Вопрос не найден.", 404)
    copy = Questions_tpv(task=f"{source.task} (копия)", answer=source.answer, comment=source.comment, author=source.author, flip=source.flip, show="false")
    db.session.add(copy)
    db.session.flush()
    tpv_editor_recalculate_theme(copy.flip)
    tpv_editor_history_add(
        "question",
        copy.id,
        "create",
        f"Создана копия вопроса #{source.id}",
        after=tpv_editor_question_snapshot(copy),
        details=f"Исходный вопрос: #{source.id}.",
        can_revert=True,
    )
    db.session.commit()
    return jsonify({"ok": True, "message": f"Создана копия вопроса #{copy.id}.", "question": tpv_editor_question_to_dict(copy)}), 201


@app.post("/tpv_editor/api/questions/reset-shown")
def tpv_editor_reset_question_show():
    if not tpv_editor_allowed():
        return tpv_editor_error("Нет доступа к редактору.", 403)
    questions = db.session.scalars(
        db.select(Questions_tpv).where(func.lower(Questions_tpv.show) == "true")
    ).all()
    for question in questions:
        question.show = "false"
    tpv_editor_history_add(
        "bulk",
        None,
        "reset",
        "Сброшены использованные вопросы",
        details=f"Обновлено вопросов: {len(questions)}.",
        can_revert=False,
    )
    db.session.commit()
    return jsonify({"ok": True, "message": f"Сброшено использованных вопросов: {len(questions)}.", "updated": len(questions)})

# ============================================================================
# КОНЕЦ БЛОКА TPV EDITOR — ЭТАП 2 (ЭТАП 3 НИЖЕ)
# ============================================================================


# ============================================================================
# TPV EDITOR — ЭТАП 3: ТЕМЫ
# Темы являются значениями Questions_tpv.flip и UsersTpv.flip.
# Отдельная таблица тем не создаётся.
# ============================================================================

def tpv_editor_theme_rows():
    questions_all = db.session.scalars(
        db.select(Questions_tpv).where(Questions_tpv.flip.is_not(None))
    ).all()
    users_all = db.session.scalars(
        db.select(UsersTpv).where(UsersTpv.flip.is_not(None))
    ).all()

    groups = {}

    for question in questions_all:
        value = tpv_editor_normalize_text(question.flip)
        if not value or tpv_editor_is_general_theme(value):
            continue

        key = tpv_editor_theme_key(value)
        group = groups.setdefault(key, {
            "variants": set(),
            "questions": [],
            "users": [],
        })
        group["variants"].add(value)
        group["questions"].append(question)

    for user in users_all:
        value = tpv_editor_normalize_text(user.flip)
        if not value or tpv_editor_is_general_theme(value):
            continue

        key = tpv_editor_theme_key(value)
        group = groups.setdefault(key, {
            "variants": set(),
            "questions": [],
            "users": [],
        })
        group["variants"].add(value)
        group["users"].append(user)

    result = []

    for group in groups.values():
        variants = sorted(group["variants"], key=str.casefold)
        questions = group["questions"]
        users = group["users"]
        question_count = len(questions)
        shown_count = sum(str(question.show).lower() == "true" for question in questions)
        approved_count = sum(str(user.approve).lower() == "true" for user in users)

        result.append({
            "name": variants[0],
            "variants": variants,
            "question_count": question_count,
            "shown_count": shown_count,
            "user_count": len(users),
            "approved_count": approved_count,
            "required_questions": TPV_REQUIRED_FLIP_QUESTIONS,
            "ready": question_count >= TPV_REQUIRED_FLIP_QUESTIONS,
        })

    return sorted(result, key=lambda item: item["name"].casefold())


@app.get("/tpv_editor/api/themes-dashboard")
def tpv_editor_themes_dashboard():
    if not tpv_editor_allowed():
        return tpv_editor_error("Нет доступа к редактору.", 403)
    return jsonify({"ok": True, "themes": tpv_editor_theme_rows()})


@app.post("/tpv_editor/api/themes/rename")
def tpv_editor_rename_theme():
    if not tpv_editor_allowed():
        return tpv_editor_error("Нет доступа к редактору.", 403)

    data = request.get_json(silent=True) or {}
    old_name = tpv_editor_normalize_text(data.get("old_name"))
    new_name = tpv_editor_normalize_text(data.get("new_name"))

    if not old_name or tpv_editor_is_general_theme(old_name):
        return tpv_editor_error("Исходная тема не указана.")
    if not new_name or tpv_editor_is_general_theme(new_name):
        return tpv_editor_error("Название темы не может быть пустым или общим.")
    if old_name.casefold() == new_name.casefold() and old_name == new_name:
        return tpv_editor_error("Название темы не изменилось.")

    questions = tpv_editor_matching_questions(old_name)
    users = tpv_editor_matching_users(old_name)
    history_before = {
        "questions": [tpv_editor_question_snapshot(item) for item in questions],
        "users": [tpv_editor_user_snapshot(item) for item in users],
    }

    if not questions and not users:
        return tpv_editor_error("Тема не найдена.", 404)

    for question in questions:
        question.flip = new_name
    for user in users:
        user.flip = new_name

    db.session.flush()
    # Пересчитываются пользователи целевой темы, включая тех, кто уже был в ней.
    affected_users = tpv_editor_matching_users(new_name)
    for user in affected_users:
        tpv_editor_update_approval(user)

    history_after = {
        "questions": [tpv_editor_question_snapshot(item) for item in questions],
        "users": [tpv_editor_user_snapshot(item) for item in users],
    }
    tpv_editor_history_add(
        "theme",
        old_name,
        "rename",
        f"Тема «{old_name}» переименована в «{new_name}»",
        before=history_before,
        after=history_after,
        can_revert=True,
    )
    db.session.commit()
    return jsonify({
        "ok": True,
        "message": (
            f"Тема «{old_name}» перенесена в «{new_name}». "
            f"Вопросов: {len(questions)}, пользователей: {len(users)}."
        ),
    })


@app.post("/tpv_editor/api/themes/delete")
def tpv_editor_delete_theme():
    if not tpv_editor_allowed():
        return tpv_editor_error("Нет доступа к редактору.", 403)

    data = request.get_json(silent=True) or {}
    name = tpv_editor_normalize_text(data.get("name"))
    target_raw = data.get("target")
    target = "false" if tpv_editor_is_general_theme(target_raw) else tpv_editor_normalize_text(target_raw)

    if not name or tpv_editor_is_general_theme(name):
        return tpv_editor_error("Удаляемая тема не указана.")
    if not target:
        target = "false"
    if not tpv_editor_is_general_theme(target) and name.casefold() == target.casefold():
        return tpv_editor_error("Нельзя перенести тему саму в себя.")

    questions = tpv_editor_matching_questions(name)
    users = tpv_editor_matching_users(name)
    history_before = {
        "questions": [tpv_editor_question_snapshot(item) for item in questions],
        "users": [tpv_editor_user_snapshot(item) for item in users],
    }

    if not questions and not users:
        return tpv_editor_error("Тема не найдена.", 404)

    for question in questions:
        question.flip = target
    for user in users:
        user.flip = target

    db.session.flush()

    # Если перенос в общие — пользователи теряют тему. Иначе пересчитывается цель.
    if tpv_editor_is_general_theme(target):
        for user in users:
            tpv_editor_update_approval(user)
        target_label = "общие вопросы / без темы"
    else:
        affected_users = tpv_editor_matching_users(target)
        for user in affected_users:
            tpv_editor_update_approval(user)
        target_label = f"тему «{target}»"

    history_after = {
        "questions": [tpv_editor_question_snapshot(item) for item in questions],
        "users": [tpv_editor_user_snapshot(item) for item in users],
    }
    tpv_editor_history_add(
        "theme",
        name,
        "delete",
        f"Удалена тема «{name}»",
        before=history_before,
        after=history_after,
        details=f"Перенос выполнен в: {target_label}.",
        can_revert=True,
    )
    db.session.commit()
    return jsonify({
        "ok": True,
        "message": (
            f"Тема «{name}» удалена. Связанные записи перенесены в {target_label}. "
            f"Вопросов: {len(questions)}, пользователей: {len(users)}."
        ),
    })


# ============================================================================
# ЭТАП 4 — ПРОВЕРКА БАЗЫ
# ============================================================================
import re as _tpv_re

_TPV_HTML_RE = _tpv_re.compile(r"<[^>]+>")
_TPV_CONTROL_RE = _tpv_re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]")


def tpv_editor_compact_text(value):
    return " ".join(str(value or "").strip().split())


def tpv_editor_quality_issue(code, level, entity, record_id, title, details, recommendation, fixable=False):
    return {
        "key": f"{entity}:{record_id}:{code}",
        "code": code,
        "level": level,
        "entity": entity,
        "record_id": record_id,
        "title": title,
        "details": details,
        "recommendation": recommendation,
        "fixable": bool(fixable),
    }


def tpv_editor_build_quality_report():
    questions = db.session.scalars(db.select(Questions_tpv).order_by(Questions_tpv.id)).all()
    users = db.session.scalars(db.select(UsersTpv).order_by(UsersTpv.id)).all()
    issues = []

    user_names = {tpv_editor_normalize_text(user.username).casefold() for user in users if user.username}
    duplicate_groups = {}

    for question in questions:
        task = str(question.task or "")
        answer = str(question.answer or "")
        author = str(question.author or "")
        comment = str(question.comment or "")
        flip = str(question.flip or "")
        show = str(question.show or "").casefold()

        if not task.strip():
            issues.append(tpv_editor_quality_issue("empty_task", "critical", "question", question.id, "Пустой текст вопроса", f"Вопрос #{question.id} не содержит формулировку.", "Откройте вопрос и заполните поле «Вопрос»."))
        if not answer.strip():
            issues.append(tpv_editor_quality_issue("empty_answer", "critical", "question", question.id, "Пустой ответ", f"Вопрос #{question.id} не содержит ответа.", "Откройте вопрос и заполните поле «Ответ»."))
        if not author.strip():
            issues.append(tpv_editor_quality_issue("empty_author", "warning", "question", question.id, "Не указан автор", f"У вопроса #{question.id} отсутствует автор.", "Назначьте автора вручную."))
        elif tpv_editor_normalize_text(author).casefold() not in user_names:
            issues.append(tpv_editor_quality_issue("unknown_author", "warning", "question", question.id, "Автор отсутствует в UsersTpv", f"Автор «{author}» не найден среди пользователей.", "Создайте пользователя либо исправьте имя автора."))

        fields = {"вопрос": task, "ответ": answer, "комментарий": comment, "автор": author, "тема": flip}
        dirty_fields = [name for name, value in fields.items() if value != tpv_editor_compact_text(value)]
        if dirty_fields:
            issues.append(tpv_editor_quality_issue("whitespace", "warning", "question", question.id, "Лишние пробелы", "Поля: " + ", ".join(dirty_fields) + ".", "Можно безопасно удалить пробелы по краям и повторяющиеся пробелы.", True))

        html_fields = [name for name, value in fields.items() if _TPV_HTML_RE.search(value)]
        if html_fields:
            issues.append(tpv_editor_quality_issue("html", "warning", "question", question.id, "Обнаружены HTML-теги", "Поля: " + ", ".join(html_fields) + ".", "Проверьте разметку вручную: автоматическое удаление может изменить смысл."))

        control_fields = [name for name, value in fields.items() if _TPV_CONTROL_RE.search(value)]
        if control_fields:
            issues.append(tpv_editor_quality_issue("control_chars", "warning", "question", question.id, "Управляющие символы", "Поля: " + ", ".join(control_fields) + ".", "Можно безопасно удалить непечатные управляющие символы.", True))

        if show not in {"true", "false"}:
            issues.append(tpv_editor_quality_issue("invalid_show", "warning", "question", question.id, "Некорректный статус show", f"Сохранено значение «{question.show}».", "Статус будет приведён к false.", True))

        if len(task.strip()) < 10 and task.strip():
            issues.append(tpv_editor_quality_issue("short_task", "info", "question", question.id, "Очень короткий вопрос", f"Длина формулировки: {len(task.strip())} символов.", "Проверьте, достаточно ли информации для игрока."))
        if len(task) > 500:
            issues.append(tpv_editor_quality_issue("long_task", "warning", "question", question.id, "Очень длинный вопрос", f"Длина формулировки: {len(task)} символов.", "Сократите формулировку или проверьте лимиты интерфейса."))

        duplicate_key = (tpv_editor_compact_text(task).casefold(), tpv_editor_compact_text(answer).casefold())
        if duplicate_key[0] and duplicate_key[1]:
            duplicate_groups.setdefault(duplicate_key, []).append(question)

    for group in duplicate_groups.values():
        if len(group) > 1:
            ids = [question.id for question in group]
            for question in group:
                others = [str(item) for item in ids if item != question.id]
                issues.append(tpv_editor_quality_issue("duplicate", "warning", "question", question.id, "Точный дубликат вопроса", "Совпадает с вопросами: " + ", ".join(others) + ".", "Сравните записи и удалите лишние вручную."))

    for user in users:
        theme = tpv_editor_normalize_theme(user.flip)
        expected_count = tpv_editor_count_questions(theme)
        expected_approve = "true" if (not tpv_editor_is_general_theme(theme) and expected_count >= TPV_REQUIRED_FLIP_QUESTIONS) else "false"
        actual_count = int(user.flip_col or 0)
        actual_approve = str(user.approve or "false").casefold()

        if actual_count != expected_count or actual_approve != expected_approve:
            issues.append(tpv_editor_quality_issue("stale_approval", "warning", "user", user.id, "Устаревший допуск пользователя", f"{user.username}: сохранено вопросов {actual_count}, должно быть {expected_count}; approve={actual_approve}, должно быть {expected_approve}.", "Можно безопасно пересчитать flip_col и approve.", True))

        username = str(user.username or "")
        if username != tpv_editor_compact_text(username):
            issues.append(tpv_editor_quality_issue("user_whitespace", "warning", "user", user.id, "Лишние пробелы в имени", f"Имя сохранено как «{username}».", "Исправьте имя вручную, чтобы не нарушить связь с author."))

    return {
        "issues": issues,
        "stats": {
            "total": len(issues),
            "critical": sum(item["level"] == "critical" for item in issues),
            "warning": sum(item["level"] == "warning" for item in issues),
            "info": sum(item["level"] == "info" for item in issues),
            "fixable": sum(item["fixable"] for item in issues),
            "scanned": len(questions) + len(users),
        },
    }


@app.get("/tpv_editor/api/quality-report")
def tpv_editor_quality_report():
    if not tpv_editor_allowed():
        return tpv_editor_error("Нет доступа к редактору.", 403)
    report = tpv_editor_build_quality_report()
    return jsonify({"ok": True, **report})


@app.post("/tpv_editor/api/quality/fix")
def tpv_editor_quality_fix():
    if not tpv_editor_allowed():
        return tpv_editor_error("Нет доступа к редактору.", 403)

    data = request.get_json(silent=True) or {}
    code = str(data.get("code") or "")
    entity = str(data.get("entity") or "")
    record_id = data.get("record_id")

    if entity == "question":
        question = db.session.get(Questions_tpv, int(record_id))
        if question is None:
            return tpv_editor_error("Вопрос не найден.", 404)

        if code == "whitespace":
            question.task = tpv_editor_compact_text(question.task)
            question.answer = tpv_editor_compact_text(question.answer)
            question.comment = tpv_editor_compact_text(question.comment)
            question.author = tpv_editor_compact_text(question.author)
            question.flip = tpv_editor_normalize_theme(question.flip)
        elif code == "control_chars":
            for field in ("task", "answer", "comment", "author", "flip"):
                setattr(question, field, _TPV_CONTROL_RE.sub("", str(getattr(question, field) or "")))
        elif code == "invalid_show":
            question.show = "false"
        else:
            return tpv_editor_error("Для этой проблемы нет безопасного автоматического исправления.")

        db.session.commit()
        return jsonify({"ok": True, "message": f"Вопрос #{question.id} исправлен."})

    if entity == "user" and code == "stale_approval":
        user = db.session.get(UsersTpv, int(record_id))
        if user is None:
            return tpv_editor_error("Пользователь не найден.", 404)
        tpv_editor_update_approval(user)
        db.session.commit()
        return jsonify({"ok": True, "message": f"Допуск пользователя «{user.username}» пересчитан."})

    return tpv_editor_error("Автоматическое исправление недоступно.")


@app.post("/tpv_editor/api/quality/fix-all-safe")
def tpv_editor_quality_fix_all_safe():
    if not tpv_editor_allowed():
        return tpv_editor_error("Нет доступа к редактору.", 403)

    questions = db.session.scalars(db.select(Questions_tpv)).all()
    users = db.session.scalars(db.select(UsersTpv)).all()
    fixed = 0

    for question in questions:
        original = (question.task, question.answer, question.comment, question.author, question.flip, question.show)
        question.task = _TPV_CONTROL_RE.sub("", tpv_editor_compact_text(question.task))
        question.answer = _TPV_CONTROL_RE.sub("", tpv_editor_compact_text(question.answer))
        question.comment = _TPV_CONTROL_RE.sub("", tpv_editor_compact_text(question.comment))
        question.author = _TPV_CONTROL_RE.sub("", tpv_editor_compact_text(question.author))
        question.flip = tpv_editor_normalize_theme(_TPV_CONTROL_RE.sub("", str(question.flip or "")))
        if str(question.show or "").casefold() not in {"true", "false"}:
            question.show = "false"
        current = (question.task, question.answer, question.comment, question.author, question.flip, question.show)
        if current != original:
            fixed += 1

    for user in users:
        before = (int(user.flip_col or 0), str(user.approve or "false"))
        tpv_editor_update_approval(user)
        after = (int(user.flip_col or 0), str(user.approve or "false"))
        if before != after:
            fixed += 1

    db.session.commit()
    return jsonify({"ok": True, "message": f"Безопасно исправлено записей: {fixed}.", "fixed": fixed})



@app.get("/tpv_editor/api/statistics")
def tpv_editor_statistics():
    if not tpv_editor_allowed():
        return tpv_editor_error("Нет доступа к редактору.", 403)

    users = db.session.scalars(
        db.select(UsersTpv).order_by(UsersTpv.username)
    ).all()

    questions = db.session.scalars(
        db.select(Questions_tpv).order_by(Questions_tpv.id)
    ).all()

    theme_rows = tpv_editor_theme_rows()

    total_questions = len(questions)
    general_questions = sum(
        tpv_editor_is_general_theme(question.flip)
        for question in questions
    )
    themed_questions = total_questions - general_questions
    shown_questions = sum(
        str(question.show or "").casefold() == "true"
        for question in questions
    )
    unused_questions = total_questions - shown_questions
    usage_percent = round(
        shown_questions * 100 / total_questions,
        1,
    ) if total_questions else 0

    approved_users = sum(
        str(user.approve or "").casefold() == "true"
        for user in users
    )
    users_without_theme = sum(
        tpv_editor_is_general_theme(user.flip)
        for user in users
    )
    users_not_approved = len(users) - approved_users

    ready_themes = sum(row["ready"] for row in theme_rows)
    shortage_themes = sum(
        not row["ready"]
        for row in theme_rows
    )

    author_groups = {}

    for question in questions:
        author_name = tpv_editor_normalize_text(question.author)
        author_key = author_name.casefold()

        group = author_groups.setdefault(author_key, {
            "name": author_name or "Без автора",
            "total": 0,
            "general": 0,
            "themed": 0,
            "shown": 0,
        })

        group["total"] += 1

        if tpv_editor_is_general_theme(question.flip):
            group["general"] += 1
        else:
            group["themed"] += 1

        if str(question.show or "").casefold() == "true":
            group["shown"] += 1

    authors = sorted(
        author_groups.values(),
        key=lambda item: (-item["total"], item["name"].casefold()),
    )

    themes = sorted(
        [
            {
                "name": row["name"],
                "question_count": row["question_count"],
                "shown_count": row["shown_count"],
                "user_count": row["user_count"],
                "approved_count": row["approved_count"],
                "ready": row["ready"],
                "missing": max(
                    0,
                    TPV_REQUIRED_FLIP_QUESTIONS - row["question_count"],
                ),
            }
            for row in theme_rows
        ],
        key=lambda item: (-item["question_count"], item["name"].casefold()),
    )

    user_rows = []

    for user in users:
        theme = "" if tpv_editor_is_general_theme(user.flip) else (
            tpv_editor_normalize_text(user.flip)
        )

        user_rows.append({
            "id": user.id,
            "username": user.username,
            "money": int(user.money or 0),
            "theme": theme,
            "question_count": int(user.flip_col or 0),
            "approved": str(user.approve or "").casefold() == "true",
        })

    return jsonify({
        "ok": True,
        "statistics": {
            "summary": {
                "users": len(users),
                "questions": total_questions,
                "themes": len(theme_rows),
                "approved_users": approved_users,
                "total_money": sum(int(user.money or 0) for user in users),
            },
            "questions": {
                "total": total_questions,
                "general": general_questions,
                "themed": themed_questions,
                "shown": shown_questions,
                "unused": unused_questions,
                "usage_percent": usage_percent,
            },
            "readiness": {
                "ready_themes": ready_themes,
                "shortage_themes": shortage_themes,
                "users_without_theme": users_without_theme,
                "users_not_approved": users_not_approved,
            },
            "themes": themes[:15],
            "authors": authors[:20],
            "users": user_rows,
        },
    })



# ---------------------------------------------------------------------------
# ЭТАП 6 — ИМПОРТ / ЭКСПОРТ
# Для XLSX нужен пакет: pip install openpyxl
# ---------------------------------------------------------------------------

import csv as tpv_csv
import io as tpv_io
import json as tpv_json
import os as tpv_os
import sqlite3 as tpv_sqlite3
import tempfile as tpv_tempfile
from datetime import datetime as tpv_datetime
from flask import send_file as tpv_send_file


def tpv_editor_export_users_rows():
    users = db.session.scalars(db.select(UsersTpv).order_by(UsersTpv.id)).all()
    return [
        {
            "id": user.id,
            "username": user.username or "",
            "money": int(user.money or 0),
            "flip": "false" if tpv_editor_is_general_theme(user.flip) else tpv_editor_normalize_text(user.flip),
            "flip_col": int(user.flip_col or 0),
            "approve": "true" if str(user.approve or "").casefold() == "true" else "false",
        }
        for user in users
    ]


def tpv_editor_export_questions_rows():
    questions = db.session.scalars(db.select(Questions_tpv).order_by(Questions_tpv.id)).all()
    return [
        {
            "id": question.id,
            "task": question.task or "",
            "answer": question.answer or "",
            "comment": question.comment or "",
            "author": question.author or "",
            "flip": "false" if tpv_editor_is_general_theme(question.flip) else tpv_editor_normalize_text(question.flip),
            "show": "true" if str(question.show or "").casefold() == "true" else "false",
        }
        for question in questions
    ]


def tpv_editor_download_bytes(content, filename, mimetype):
    buffer = tpv_io.BytesIO(content)
    buffer.seek(0)
    return tpv_send_file(buffer, as_attachment=True, download_name=filename, mimetype=mimetype)


def tpv_editor_csv_bytes(rows, fields):
    stream = tpv_io.StringIO(newline="")
    writer = tpv_csv.DictWriter(stream, fieldnames=fields, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(rows)
    return ("\ufeff" + stream.getvalue()).encode("utf-8")


def tpv_editor_xlsx_bytes(users=None, questions=None):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError as exc:
        raise RuntimeError("Для Excel установите пакет openpyxl: pip install openpyxl") from exc

    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)
    header_fill = PatternFill("solid", fgColor="12354A")
    header_font = Font(color="FFFFFF", bold=True)

    def add_sheet(name, rows, fields):
        sheet = workbook.create_sheet(name)
        sheet.append(fields)
        for row in rows:
            sheet.append([row.get(field, "") for field in fields])
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        sheet.freeze_panes = "A2"
        for column in sheet.columns:
            width = min(55, max(12, max(len(str(cell.value or "")) for cell in column) + 2))
            sheet.column_dimensions[column[0].column_letter].width = width

    if users is not None:
        add_sheet("Users", users, ["id", "username", "money", "flip", "flip_col", "approve"])
    if questions is not None:
        add_sheet("Questions", questions, ["id", "task", "answer", "comment", "author", "flip", "show"])

    output = tpv_io.BytesIO()
    workbook.save(output)
    return output.getvalue()


@app.get("/tpv_editor/export/users.<fmt>")
def tpv_editor_export_users(fmt):
    if not tpv_editor_allowed():
        return tpv_editor_error("Нет доступа к редактору.", 403)
    rows = tpv_editor_export_users_rows()
    stamp = tpv_datetime.now().strftime("%Y%m%d_%H%M%S")
    if fmt == "csv":
        return tpv_editor_download_bytes(tpv_editor_csv_bytes(rows, ["id", "username", "money", "flip", "flip_col", "approve"]), f"tpv_users_{stamp}.csv", "text/csv; charset=utf-8")
    if fmt == "json":
        data = tpv_json.dumps({"users": rows}, ensure_ascii=False, indent=2).encode("utf-8")
        return tpv_editor_download_bytes(data, f"tpv_users_{stamp}.json", "application/json")
    if fmt == "xlsx":
        try:
            data = tpv_editor_xlsx_bytes(users=rows)
        except RuntimeError as exc:
            return tpv_editor_error(str(exc), 500)
        return tpv_editor_download_bytes(data, f"tpv_users_{stamp}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    return tpv_editor_error("Неизвестный формат.", 404)


@app.get("/tpv_editor/export/questions.<fmt>")
def tpv_editor_export_questions(fmt):
    if not tpv_editor_allowed():
        return tpv_editor_error("Нет доступа к редактору.", 403)
    rows = tpv_editor_export_questions_rows()
    stamp = tpv_datetime.now().strftime("%Y%m%d_%H%M%S")
    if fmt == "csv":
        return tpv_editor_download_bytes(tpv_editor_csv_bytes(rows, ["id", "task", "answer", "comment", "author", "flip", "show"]), f"tpv_questions_{stamp}.csv", "text/csv; charset=utf-8")
    if fmt == "json":
        data = tpv_json.dumps({"questions": rows}, ensure_ascii=False, indent=2).encode("utf-8")
        return tpv_editor_download_bytes(data, f"tpv_questions_{stamp}.json", "application/json")
    if fmt == "xlsx":
        try:
            data = tpv_editor_xlsx_bytes(questions=rows)
        except RuntimeError as exc:
            return tpv_editor_error(str(exc), 500)
        return tpv_editor_download_bytes(data, f"tpv_questions_{stamp}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    return tpv_editor_error("Неизвестный формат.", 404)


@app.get("/tpv_editor/export/full.<fmt>")
def tpv_editor_export_full(fmt):
    if not tpv_editor_allowed():
        return tpv_editor_error("Нет доступа к редактору.", 403)
    users = tpv_editor_export_users_rows()
    questions = tpv_editor_export_questions_rows()
    stamp = tpv_datetime.now().strftime("%Y%m%d_%H%M%S")
    if fmt == "json":
        payload = {
            "format": "tpv-editor-backup",
            "version": 1,
            "created_at": tpv_datetime.now().isoformat(timespec="seconds"),
            "users": users,
            "questions": questions,
        }
        return tpv_editor_download_bytes(tpv_json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8"), f"tpv_backup_{stamp}.json", "application/json")
    if fmt == "xlsx":
        try:
            data = tpv_editor_xlsx_bytes(users=users, questions=questions)
        except RuntimeError as exc:
            return tpv_editor_error(str(exc), 500)
        return tpv_editor_download_bytes(data, f"tpv_backup_{stamp}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    return tpv_editor_error("Неизвестный формат.", 404)


@app.get("/tpv_editor/export/database.sqlite")
def tpv_editor_export_database():
    if not tpv_editor_allowed():
        return tpv_editor_error("Нет доступа к редактору.", 403)
    database_path = db.engine.url.database
    if not database_path:
        return tpv_editor_error("Не удалось определить файл SQLite.", 500)
    source_path = tpv_os.path.abspath(database_path)
    if not tpv_os.path.exists(source_path):
        return tpv_editor_error("Файл SQLite не найден.", 404)
    temp = tpv_tempfile.NamedTemporaryFile(prefix="tpv_backup_", suffix=".sqlite", delete=False)
    temp.close()
    try:
        with tpv_sqlite3.connect(source_path) as source, tpv_sqlite3.connect(temp.name) as target:
            source.backup(target)
        stamp = tpv_datetime.now().strftime("%Y%m%d_%H%M%S")
        return tpv_send_file(temp.name, as_attachment=True, download_name=f"tpv_database_{stamp}.sqlite", mimetype="application/vnd.sqlite3")
    except Exception:
        if tpv_os.path.exists(temp.name):
            tpv_os.unlink(temp.name)
        raise


def tpv_editor_read_import_file(upload, requested_entity="auto"):
    filename = (upload.filename or "").lower()
    raw = upload.read()
    users, questions = [], []

    if filename.endswith(".json"):
        payload = tpv_json.loads(raw.decode("utf-8-sig"))
        if isinstance(payload, list):
            if requested_entity == "users": users = payload
            elif requested_entity == "questions": questions = payload
            else: raise ValueError("Для JSON-массива явно выберите раздел импорта.")
        elif isinstance(payload, dict):
            users = payload.get("users") or []
            questions = payload.get("questions") or []
        else:
            raise ValueError("Некорректная структура JSON.")

    elif filename.endswith(".csv"):
        text = raw.decode("utf-8-sig")
        rows = list(tpv_csv.DictReader(tpv_io.StringIO(text)))
        entity = requested_entity
        if entity == "auto":
            headers = {str(name or "").strip().casefold() for name in (rows[0].keys() if rows else [])}
            entity = "questions" if "task" in headers or "answer" in headers else "users"
        if entity == "users": users = rows
        elif entity == "questions": questions = rows
        else: raise ValueError("CSV может содержать только один раздел.")

    elif filename.endswith(".xlsx"):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("Для Excel установите пакет openpyxl: pip install openpyxl") from exc
        workbook = load_workbook(tpv_io.BytesIO(raw), read_only=True, data_only=True)

        def sheet_rows(sheet):
            values = list(sheet.iter_rows(values_only=True))
            if not values: return []
            headers = [tpv_editor_normalize_text(value).casefold() for value in values[0]]
            return [dict(zip(headers, row)) for row in values[1:] if any(value is not None and str(value).strip() for value in row)]

        names = {name.casefold(): name for name in workbook.sheetnames}
        if "users" in names: users = sheet_rows(workbook[names["users"]])
        if "questions" in names: questions = sheet_rows(workbook[names["questions"]])
        if not users and not questions and len(workbook.sheetnames) == 1:
            rows = sheet_rows(workbook[workbook.sheetnames[0]])
            entity = requested_entity
            if entity == "auto":
                keys = set(rows[0].keys()) if rows else set()
                entity = "questions" if "task" in keys or "answer" in keys else "users"
            if entity == "users": users = rows
            elif entity == "questions": questions = rows
    else:
        raise ValueError("Поддерживаются только CSV, JSON и XLSX.")

    return users, questions


def tpv_editor_validate_import(users, questions):
    errors, warnings = [], []
    clean_users, clean_questions = [], []

    for index, row in enumerate(users, start=2):
        normalized = {str(key or "").strip().casefold(): value for key, value in dict(row).items()}
        username = tpv_editor_normalize_text(normalized.get("username"))
        if not username:
            errors.append(f"Пользователь, строка {index}: отсутствует username.")
            continue
        if len(username) > 10:
            errors.append(f"Пользователь «{username}»: имя длиннее 10 символов.")
            continue
        try:
            money = int(normalized.get("money") or 0)
        except (TypeError, ValueError):
            errors.append(f"Пользователь «{username}»: некорректный баланс.")
            continue
        clean_users.append({"username": username, "money": money, "flip": tpv_editor_normalize_theme(normalized.get("flip"))})

    seen_users = set()
    for row in clean_users:
        key = row["username"].casefold()
        if key in seen_users: warnings.append(f"Повтор пользователя в файле: {row['username']}.")
        seen_users.add(key)

    for index, row in enumerate(questions, start=2):
        normalized = {str(key or "").strip().casefold(): value for key, value in dict(row).items()}
        task = tpv_editor_normalize_text(normalized.get("task"))
        answer = tpv_editor_normalize_text(normalized.get("answer"))
        if not task:
            errors.append(f"Вопрос, строка {index}: отсутствует task.")
            continue
        if not answer:
            errors.append(f"Вопрос, строка {index}: отсутствует answer.")
            continue
        show = "true" if str(normalized.get("show") or "false").casefold() == "true" else "false"
        clean_questions.append({
            "task": task,
            "answer": answer,
            "comment": tpv_editor_normalize_text(normalized.get("comment")),
            "author": tpv_editor_normalize_text(normalized.get("author")),
            "flip": tpv_editor_normalize_theme(normalized.get("flip")),
            "show": show,
        })

    return clean_users, clean_questions, errors, warnings


def tpv_editor_parse_import_request():
    upload = request.files.get("file")
    if upload is None or not upload.filename:
        raise ValueError("Файл не выбран.")
    entity = request.form.get("entity", "auto")
    mode = request.form.get("mode", "merge")
    if entity not in {"auto", "users", "questions", "all"}: raise ValueError("Некорректный раздел импорта.")
    if mode not in {"merge", "replace"}: raise ValueError("Некорректный режим импорта.")
    users, questions = tpv_editor_read_import_file(upload, entity)
    clean_users, clean_questions, errors, warnings = tpv_editor_validate_import(users, questions)
    if entity == "users": clean_questions = []
    elif entity == "questions": clean_users = []
    return entity, mode, clean_users, clean_questions, errors, warnings


@app.post("/tpv_editor/api/import-preview")
def tpv_editor_import_preview():
    if not tpv_editor_allowed(): return tpv_editor_error("Нет доступа к редактору.", 403)
    try:
        entity, mode, users, questions, errors, warnings = tpv_editor_parse_import_request()
    except (ValueError, RuntimeError, tpv_json.JSONDecodeError, UnicodeDecodeError) as exc:
        return tpv_editor_error(str(exc))
    if not users and not questions and not errors:
        errors.append("В файле не найдено данных для импорта.")
    return jsonify({"ok": True, "preview": {"valid": not errors, "entity": entity, "mode": mode, "users_count": len(users), "questions_count": len(questions), "errors": errors, "warnings": warnings}})


@app.post("/tpv_editor/api/import")
def tpv_editor_import_apply():
    if not tpv_editor_allowed(): return tpv_editor_error("Нет доступа к редактору.", 403)
    try:
        entity, mode, users_data, questions_data, errors, warnings = tpv_editor_parse_import_request()
        if errors: return tpv_editor_error("Импорт отменён: " + "; ".join(errors))

        if mode == "replace":
            if users_data and (entity in {"users", "all", "auto"}):
                db.session.execute(db.delete(UsersTpv))
            if questions_data and (entity in {"questions", "all", "auto"}):
                db.session.execute(db.delete(Questions_tpv))
            db.session.flush()

        existing_users = db.session.scalars(db.select(UsersTpv)).all()
        user_map = {tpv_editor_normalize_text(user.username).casefold(): user for user in existing_users}
        users_added = users_updated = 0
        for row in users_data:
            key = row["username"].casefold()
            user = user_map.get(key)
            if user is None:
                user = UsersTpv()
                user.username = row["username"]
                user.flip_col = 0
                user.approve = "false"
                db.session.add(user)
                user_map[key] = user
                users_added += 1
            else:
                users_updated += 1
            user.money = row["money"]
            user.flip = row["flip"]

        existing_questions = db.session.scalars(db.select(Questions_tpv)).all()
        question_keys = {
            (tpv_editor_normalize_text(q.task).casefold(), tpv_editor_normalize_text(q.answer).casefold(), tpv_editor_normalize_text(q.author).casefold(), tpv_editor_theme_key(q.flip))
            for q in existing_questions
        }
        questions_added = questions_skipped = 0
        for row in questions_data:
            key = (row["task"].casefold(), row["answer"].casefold(), row["author"].casefold(), tpv_editor_theme_key(row["flip"]))
            if mode == "merge" and key in question_keys:
                questions_skipped += 1
                continue
            question = Questions_tpv()
            question.task = row["task"]
            question.answer = row["answer"]
            question.comment = row["comment"]
            question.author = row["author"]
            question.flip = row["flip"]
            question.show = row["show"]
            db.session.add(question)
            question_keys.add(key)
            questions_added += 1

        db.session.flush()
        for user in db.session.scalars(db.select(UsersTpv)).all():
            tpv_editor_update_approval(user)
        tpv_editor_history_add(
            "bulk",
            None,
            "import",
            "Выполнен импорт данных",
            details=(
                f"Режим: {mode}; пользователей добавлено: {users_added}; "
                f"обновлено: {users_updated}; вопросов добавлено: "
                f"{questions_added}; дублей пропущено: {questions_skipped}."
            ),
            can_revert=False,
        )
        db.session.commit()

        return jsonify({"ok": True, "message": f"Импорт завершён. Пользователи: добавлено {users_added}, обновлено {users_updated}. Вопросы: добавлено {questions_added}, пропущено дублей {questions_skipped}."})
    except (ValueError, RuntimeError, tpv_json.JSONDecodeError, UnicodeDecodeError) as exc:
        db.session.rollback()
        return tpv_editor_error(str(exc))
    except Exception:
        db.session.rollback()
        raise


# ============================================================================
# КОНЕЦ БЛОКА TPV EDITOR — ЭТАП 9
# ============================================================================



@socketio.on("update_users_tpv")
def update_users_tpv():
        js = db.session.scalars(db.select(QueryTpv)).all()
        if len(js)==1:
            id = js[0].id
            username = js[0].username
            flip = js[0].flip
            money = js[0].money
            status = js[0].status
            jsn = [id,username,flip,money,status,"true"]
            result = jsn
            emit_tpv_host("updated_users_tpv", result)
        else:
            jsn = []
            for i in range(0,len(js)):
                if js[i].status != "ended":
                    id = js[i].id
                    username = js[i].username
                    flip = js[i].flip
                    money = js[i].money
                    status = js[i].status
                    tmp = [id,username,flip,money,status,"false"]
                #socketio.emit("updated_user_tpv",js[i].status,to=f"{get_room_code()}:user:{js[i].username}");
                    jsn.append(tmp)
            result = jsn
            emit_tpv_host("updated_users_tpv", result)
            return result

@socketio.on("clean_db_tpv")
def clean_db_tpv():
    db.session.execute(db.delete(QueryTpv))
    db.session.commit()
    update_users_tpv()
    emit_tpv_host("DB_clean", "ok")


@socketio.on("tpv_spectator_ready")
def tpv_spectator_ready():
    """Восстанавливает код комнаты после перезагрузки экрана зрителя."""
    room_code = get_room_code()
    if room_code is None:
        emit("room_code_hide", {})
        return

    join_url = f"{request.host_url.rstrip('/')}{url_for('join')}?room={room_code}"
    emit("room_code_show", {
        "room": room_code,
        "joinUrl": join_url,
    })


@socketio.on("tpv_selection_start")
def tpv_selection_start():
    players = db.session.scalars(
        db.select(QueryTpv).where(QueryTpv.status == "wait")
    ).all()
    emit_tpv_spectator("tpv_spectator_select_start", {
        "players": [player.username for player in players]
    })


@socketio.on("tpv_versus")
def tpv_versus():
    emit_tpv_spectator("tpv_versus_spec", {"show": True})


@socketio.on("choose_player_random")
def choose_player_random():
    try:
        js = db.session.scalars(db.select(QueryTpv).where(QueryTpv.status=="wait")).all()
        if len(js)==0:
            return
        if len(js)==1:
            js[0].status = "selected"
            db.session.commit()
            id = js[0].id
            username = js[0].username
            flip = js[0].flip
            money = js[0].money
            status = js[0].status
            jsn = [id,username,flip,money,status]
            result = jsn
            socketio.emit("player_selected", result, to=f"{get_room_code()}:user:{username}")
            emit_tpv_host("player_selected", result)
            emit_tpv_spectator("tpv_spectator_player_selected", {
                "player": username,
                "topic": flip,
                "currentMoney": money,
            })
            update_users_tpv()
        else:
            secure_rnd = secrets.SystemRandom()
            num = secure_rnd.randrange(len(js))
            js[num].status = "selected"
            db.session.commit()
            id = js[num].id
            username = js[num].username
            flip = js[num].flip
            money = js[num].money
            status = js[num].status
            jsn = [id,username,flip,money,status]
            result = jsn
            update_users_tpv()
            socketio.emit("player_selected", result, to=f"{get_room_code()}:user:{username}")
            emit_tpv_host("player_selected", result)
            emit_tpv_spectator("tpv_spectator_player_selected", {
                "player": username,
                "topic": flip,
                "currentMoney": money,
            })
            update_users_tpv()
    except:
        pass
    
@socketio.on("choose_player_id")
def choose_player_id(data):
    try:
        num = data["id"]
        js = db.session.scalar(db.select(QueryTpv).where(QueryTpv.id==num))
        if js == None:
            return
        js.status = "selected"
        db.session.commit()
        id = js.id
        username = js.username
        flip = js.flip
        money = js.money
        status = js.status
        jsn = [id,username,flip,money,status]
        result = jsn
        socketio.emit("player_selected", result, to=f"{get_room_code()}:user:{username}")
        emit_tpv_host("player_selected", result)
        emit_tpv_spectator("tpv_spectator_player_selected", {
                "player": username,
                "topic": flip,
                "currentMoney": money,
            })
        update_users_tpv()
    except:
        pass

@socketio.on("reset_to_wait_tpv")
def reset_to_wait_tpv():
    try:
        js = db.session.scalars(db.select(QueryTpv)).all()
        if len(js)==1:
            if js[0].status != "ended":
                js[0].status = "wait"
                db.session.commit()
                update_users_tpv()
        else:
            for i in range(0,len(js)):
                if js[i].status != "ended":
                    js[i].status = "wait"
                    db.session.commit()
                    update_users_tpv()
        socketio.emit("reset", "wait", to=f"{get_room_code()}:user")
        emit_tpv_spectator("reset", "wait")
    except:
        return json.dump("fail")

def _emit_tpv_bong_to_player(event_name, data):
    """Пересылает этап гонг-игры выбранному игроку и экрану зрителя."""
    payload = dict(data or {})
    player = str(payload.pop("player", "") or "").strip()

    if player:
        socketio.emit(
            f"{event_name}_user",
            payload,
            to=f"{get_room_code()}:user:{player}",
        )

    emit_tpv_spectator(f"{event_name}_spec", payload)


@socketio.on("tpv_bong_prepare")
def tpv_bong_prepare(data):
    _emit_tpv_bong_to_player("tpv_bong_prepare", data)


@socketio.on("tpv_bong_selected")
def tpv_bong_selected(data):
    _emit_tpv_bong_to_player("tpv_bong_selected", data)


@socketio.on("tpv_bong_value")
def tpv_bong_value(data):
    _emit_tpv_bong_to_player("tpv_bong_value", data)


@socketio.on("tpv_bong_stop_ack")
def tpv_bong_stop_ack(data):
    _emit_tpv_bong_to_player("tpv_bong_stop_ack", data)


@socketio.on("tpv_bong_result")
def tpv_bong_result(data):
    _emit_tpv_bong_to_player("tpv_bong_result", data)


@socketio.on("tpv_bong_hide")
def tpv_bong_hide(data):
    _emit_tpv_bong_to_player("tpv_bong_hide", data)


@socketio.on("tpv_bong_stop_request")
def tpv_bong_stop_request(data):
    """Игрок нажал STOP. Запрос передаётся в комнату ведущего."""
    player = str((data or {}).get("player") or "").strip()

    if not player:
        return {"ok": False, "error": "player_required"}

    payload = {"player": player}

    # TPV-host подключается к технической комнате DEFAULT_ROOM_CODE.
    # Ранее запрос отправлялся только в get_room_code():host, поэтому
    # ведущий его не получал, а кнопка визуально казалась нерабочей.
    host_rooms = {
        f"{DEFAULT_ROOM_CODE}:host",
        f"{get_room_code()}:host",
    }

    for host_room in host_rooms:
        socketio.emit(
            "tpv_bong_stop_requested",
            payload,
            to=host_room,
        )

    return {"ok": True}


@socketio.on("generate_safe_bong_game")
def generate_safe_bong_game():
    secure_rnd = secrets.SystemRandom()
    num = secure_rnd.randint(1,3)
    emit_tpv_host("bong_game_safe_var", num)


@socketio.on("generate_sum_for_bong_game")
def generate_sum_for_bong_game(data):
    secure_rnd = secrets.SystemRandom()
    col = secure_rnd.randint(6,15)
    secure_rnd = secrets.SystemRandom()
    result = secure_rnd.sample(range(1,data['sum']),col)
    result.sort()
    result.append(data['sum'])
    emit_tpv_host("sum_generated", result)


@socketio.on("take_question")
def take_question(data):
    if data["flips"]=="false":
        js = db.session.scalar(db.select(Questions_tpv).where(Questions_tpv.flip=="false", Questions_tpv.author!=data['player'], Questions_tpv.show=="false").order_by(func.random()).limit(1))
        if js == None:
            emit_tpv_host("question_selected", "fail")
            return
        question = js.task
        answer = js.answer
        comment = js.comment
        author = js.author
        result = [question,answer,comment,author]
        js.show = "true"
        db.session.commit()
        result_user_spec = {
            "question": question,
            "author": author,
            "replacement": False,
            "questionNumber": data.get("questionNumber"),
        }
        socketio.emit("question_selected_user", result_user_spec, to=f"{get_room_code()}:user:{data['player']}")
        emit_tpv_spectator("question_selected_spec", result_user_spec)
        emit_tpv_host("question_selected", result)
    if data["flips"]!="false":
        js = db.session.scalar(db.select(Questions_tpv).where(Questions_tpv.flip==data["flips"], Questions_tpv.author!=data['player'], Questions_tpv.show=="false").order_by(func.random()).limit(1))
        if js == None:
            emit_tpv_host("question_selected", "fail")
            return
        question = js.task
        answer = js.answer
        comment = js.comment
        author = js.author
        result = [question,answer,comment,author]
        result_user_spec = {
            "question": question,
            "author": author,
            "replacement": True,
            "replacementTopic": data.get("flips"),
            "questionNumber": data.get("questionNumber"),
        }
        js.show = "true"
        db.session.commit()
        socketio.emit("question_selected_user", result_user_spec, to=f"{get_room_code()}:user:{data['player']}")
        emit_tpv_spectator("question_selected_spec", result_user_spec)        
        emit_tpv_host("question_selected", result)

@socketio.on("add_result_author")
def add_result_author(data):
    js = db.session.scalar(db.select(UsersTpv).where(UsersTpv.username==data["name_author"]))
    if js == None:
        u = UsersTpv()
        u.username = data["name_author"]
        u.flip = "false"
        u.money = data["sum_author"]
        u.approve = "false"
        u.flip_col = 0
        db.session.add(u)
        db.session.flush()
        db.session.commit()
    else:
        js.money = js.money + data["sum_author"]
        db.session.commit()

    emit_tpv_spectator(
        "tpv_author_win_user",
        {
            "amount": int(data.get("sum_author", 0) or 0),
            "author": data.get("name_author", ""),
        },
    )

@socketio.on("add_result_player")
def add_result_player(data):
    js = db.session.scalar(db.select(UsersTpv).where(UsersTpv.username==data["name_player"]))
    js.money = js.money + data["sum_player"]
    js1 = db.session.scalar(db.select(QueryTpv).where(QueryTpv.username==data["name_player"]))
    if js1 == None:
        return;
    js1.money = js1.money + data["sum_player"]
    js1.status = "ended"
    db.session.commit()

    player_win_payload = {
        "amount": int(data.get("sum_player", 0) or 0),
        "player": data.get("name_player", ""),
    }
    socketio.emit(
        "tpv_player_win_user",
        player_win_payload,
        to=f"{get_room_code()}:user:{data['name_player']}"
    )
    emit_tpv_spectator("tpv_player_win_user", player_win_payload)
    update_users_tpv()


@socketio.on("tpv_update_data_user_spec")
def tpv_update_data_user_spec(data):
    # v7 передаёт именованный объект состояния. Старый массив exp пока поддерживается.
    result = data.get("state", data.get("exp", {}))
    socketio.emit("update_data_user", result, to=f"{get_room_code()}:user")
    emit_tpv_spectator("update_data_spec", result)

@socketio.on("show_tree")
def show_tree(data):
    socketio.emit("show_tree_user", "show", to=f"{get_room_code()}:user:{data["player"]}")
    emit_tpv_spectator("show_tree_spec", "show")
    

@socketio.on("hide_tree")
def hide_tree(data):
    socketio.emit("hide_tree_user", "hide", to=f"{get_room_code()}:user:{data["player"]}")
    emit_tpv_spectator("hide_tree_spec", "hide")

@socketio.on("show_stats")
def show_stats(data):
    socketio.emit("show_stats_user", "show", to=f"{get_room_code()}:user:{data["player"]}")
    emit_tpv_spectator("show_stats_spec", "show")
    

@socketio.on("hide_stats")
def hide_stats(data):
    socketio.emit("hide_stats_user", "hide", to=f"{get_room_code()}:user:{data["player"]}")
    emit_tpv_spectator("hide_stats_spec", "hide")

@socketio.on("tpv_correct")
def tpv_correct(data):
    payload = {
        "answer": data.get("answer", ""),
        "questionNumber": data.get("questionNumber"),
        "correctCount": data.get("correctCount"),
        "round": data.get("round"),
        "roundFinished": bool(data.get("roundFinished", False)),
    }
    socketio.emit(
        "tpv_correct_user",
        payload,
        to=f"{get_room_code()}:user:{data['player']}"
    )
    emit_tpv_spectator("tpv_correct_spec", payload)

@socketio.on("tpv_pass")
def tpv_pass(data):
    payload = {
        "answer": data.get("answer", ""),
        "questionNumber": data.get("questionNumber"),
        "passCount": data.get("passCount"),
        "state": data.get("state"),
    }
    emit_tpv_player(data["player"], "tpv_pass_user", payload)
    emit_tpv_spectator("tpv_pass_spec", payload)

@socketio.on("tpv_flip")
def tpv_flip(data):
    payload = {
        "answer": data.get("answer", ""),
        "questionNumber": data.get("questionNumber"),
        "replacement": True,
        "state": data.get("state"),
    }
    emit_tpv_player(data["player"], "tpv_flip_user", payload)
    emit_tpv_spectator("tpv_flip_spec", payload)

@socketio.on("tpv_wrong")
def tpv_wrong(data):
    payload = {
        "answer": data.get("answer", ""),
        "questionNumber": data.get("questionNumber"),
        "wrongIndex": data.get("wrongIndex"),
        "state": data.get("state"),
    }
    
    emit_tpv_player(data["player"], "tpv_wrong_user", payload)
    emit_tpv_spectator("tpv_wrong_spec", payload)

@socketio.on("start_intro")
def start_intro():
     emit_tpv_spectator("start_intro",{"":""}),

@socketio.on("host_show_credits_tpv")
def host_show_credits():
    socketio.emit("show_credits_tpv", {
    "title": "Спасибо за игру!",
    "lines": [
        "Ведущий: Mokaque",
        "Редактор вопросов: Mokaque",
        "Оригинальная идея: David Briggs, Steve Knight, Mike Whitehill",
        "Голос Гонг Игры: Кирилл (Yandex SpeechKit)",
        "Техническая реализация: Mokaque",
        "Композиторы: Keith Strachan, Mattew Strachan",
        "Адаптация правил: Mokaque",
        "Графика: ChatGPT",
        "Оригинальный формат: Sony Pictures Entertainment",
        "Никто из участников создания данной адаптации игры не претендует на авторские права на формат оригинальной игры 'The People Versus'",
        "Данный проект выпущен исключительно в развлекательных целях и не преследует целей получение материальной выгоды",
        "До встречи в следующей игре!"
    ]
}, to=f"{DEFAULT_ROOM_CODE}:spectator")
   
   
@socketio.on("show_results_tpv")
def show_results_tpv():
    result = []
    tmp = db.session.scalars(db.select(UsersTpv).where(UsersTpv.money!=0).order_by(desc(UsersTpv.money))).all()
    for i in range(len(tmp)):
        username = tmp[i].username
        money = tmp[i].money
        t = [username, money]
        result.append(t)
    emit_tpv_spectator("show_results_tpv", result)
    

if __name__ == "__main__":
    _users = [' ']
    
    socketio.run(app,debug=False, host='0.0.0.0')
    


