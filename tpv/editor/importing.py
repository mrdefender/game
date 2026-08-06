"""Импорт данных TPV Editor — этап 13.6.5.

Поддерживаются JSON, CSV и XLSX. Модуль сохраняет прежние URL,
режимы merge/replace, preview и JSON-контракты.
"""

from __future__ import annotations

import csv
import io
import json
from typing import Any

from flask import jsonify, request

from .registry import EditorContext


class ImportService:
    """Чтение, проверка и применение импорта TPV Editor."""

    VALID_ENTITIES = {
        "auto",
        "users",
        "questions",
        "all",
    }
    VALID_MODES = {
        "merge",
        "replace",
    }

    def __init__(self, context: EditorContext) -> None:
        self.context = context
        self.db = context.db

        self.UsersTpv = self._dependency("UsersTpv")
        self.QuestionsTpv = self._dependency("Questions_tpv")

        self.normalize_text = self._dependency(
            "tpv_editor_normalize_text"
        )
        self.normalize_theme = self._dependency(
            "tpv_editor_normalize_theme"
        )
        self.theme_key = self._dependency(
            "tpv_editor_theme_key"
        )
        self.update_approval = self._dependency(
            "tpv_editor_update_approval"
        )
        self.history_add = self._dependency(
            "tpv_editor_history_add"
        )

    def _dependency(self, name: str) -> Any:
        value = self.context.get(name)
        if value is None:
            raise RuntimeError(
                "Импорт TPV Editor: "
                f"отсутствует зависимость {name}"
            )
        return value

    def read_file(
        self,
        upload: Any,
        requested_entity: str = "auto",
    ) -> tuple[list[Any], list[Any]]:
        filename = str(upload.filename or "").lower()
        raw = upload.read()

        users: list[Any] = []
        questions: list[Any] = []

        if filename.endswith(".json"):
            users, questions = self._read_json(
                raw,
                requested_entity,
            )
        elif filename.endswith(".csv"):
            users, questions = self._read_csv(
                raw,
                requested_entity,
            )
        elif filename.endswith(".xlsx"):
            users, questions = self._read_xlsx(
                raw,
                requested_entity,
            )
        else:
            raise ValueError(
                "Поддерживаются только CSV, JSON и XLSX."
            )

        return users, questions

    @staticmethod
    def _read_json(
        raw: bytes,
        requested_entity: str,
    ) -> tuple[list[Any], list[Any]]:
        payload = json.loads(raw.decode("utf-8-sig"))

        if isinstance(payload, list):
            if requested_entity == "users":
                return payload, []
            if requested_entity == "questions":
                return [], payload
            raise ValueError(
                "Для JSON-массива явно выберите раздел импорта."
            )

        if isinstance(payload, dict):
            return (
                payload.get("users") or [],
                payload.get("questions") or [],
            )

        raise ValueError("Некорректная структура JSON.")

    @staticmethod
    def _read_csv(
        raw: bytes,
        requested_entity: str,
    ) -> tuple[list[Any], list[Any]]:
        text = raw.decode("utf-8-sig")
        rows = list(
            csv.DictReader(io.StringIO(text))
        )

        entity = requested_entity
        if entity == "auto":
            headers = {
                str(name or "").strip().casefold()
                for name in (
                    rows[0].keys()
                    if rows
                    else []
                )
            }
            entity = (
                "questions"
                if "task" in headers or "answer" in headers
                else "users"
            )

        if entity == "users":
            return rows, []
        if entity == "questions":
            return [], rows

        raise ValueError(
            "CSV может содержать только один раздел."
        )

    def _read_xlsx(
        self,
        raw: bytes,
        requested_entity: str,
    ) -> tuple[list[Any], list[Any]]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError(
                "Для Excel установите пакет openpyxl: "
                "pip install openpyxl"
            ) from exc

        workbook = load_workbook(
            io.BytesIO(raw),
            read_only=True,
            data_only=True,
        )

        try:
            def sheet_rows(sheet: Any) -> list[dict[str, Any]]:
                values = list(
                    sheet.iter_rows(values_only=True)
                )
                if not values:
                    return []

                headers = [
                    self.normalize_text(value).casefold()
                    for value in values[0]
                ]

                return [
                    dict(zip(headers, row))
                    for row in values[1:]
                    if any(
                        value is not None
                        and str(value).strip()
                        for value in row
                    )
                ]

            names = {
                name.casefold(): name
                for name in workbook.sheetnames
            }

            users = (
                sheet_rows(workbook[names["users"]])
                if "users" in names
                else []
            )
            questions = (
                sheet_rows(workbook[names["questions"]])
                if "questions" in names
                else []
            )

            if (
                not users
                and not questions
                and len(workbook.sheetnames) == 1
            ):
                rows = sheet_rows(
                    workbook[workbook.sheetnames[0]]
                )
                entity = requested_entity

                if entity == "auto":
                    keys = (
                        set(rows[0].keys())
                        if rows
                        else set()
                    )
                    entity = (
                        "questions"
                        if "task" in keys or "answer" in keys
                        else "users"
                    )

                if entity == "users":
                    users = rows
                elif entity == "questions":
                    questions = rows

            return users, questions
        finally:
            workbook.close()

    def validate(
        self,
        users: list[Any],
        questions: list[Any],
    ) -> tuple[
        list[dict[str, Any]],
        list[dict[str, Any]],
        list[str],
        list[str],
    ]:
        errors: list[str] = []
        warnings: list[str] = []
        clean_users: list[dict[str, Any]] = []
        clean_questions: list[dict[str, Any]] = []

        for index, row in enumerate(users, start=2):
            normalized = self._normalize_row(row)
            username = self.normalize_text(
                normalized.get("username")
            )

            if not username:
                errors.append(
                    "Пользователь, строка "
                    f"{index}: отсутствует username."
                )
                continue

            if len(username) > 64:
                errors.append(
                    f"Пользователь «{username}»: "
                    "имя длиннее 64 символов."
                )
                continue

            try:
                money = int(normalized.get("money") or 0)
            except (TypeError, ValueError):
                errors.append(
                    f"Пользователь «{username}»: "
                    "некорректный баланс."
                )
                continue

            clean_users.append({
                "username": username,
                "money": money,
                "flip": self.normalize_theme(
                    normalized.get("flip")
                ),
            })

        seen_users: set[str] = set()
        for row in clean_users:
            key = row["username"].casefold()
            if key in seen_users:
                warnings.append(
                    "Повтор пользователя в файле: "
                    f"{row['username']}."
                )
            seen_users.add(key)

        for index, row in enumerate(questions, start=2):
            normalized = self._normalize_row(row)
            task = self.normalize_text(
                normalized.get("task")
            )
            answer = self.normalize_text(
                normalized.get("answer")
            )

            if not task:
                errors.append(
                    "Вопрос, строка "
                    f"{index}: отсутствует task."
                )
                continue

            if not answer:
                errors.append(
                    "Вопрос, строка "
                    f"{index}: отсутствует answer."
                )
                continue

            show = (
                "true"
                if str(
                    normalized.get("show") or "false"
                ).casefold() == "true"
                else "false"
            )

            clean_questions.append({
                "task": task,
                "answer": answer,
                "comment": self.normalize_text(
                    normalized.get("comment")
                ),
                "author": self.normalize_text(
                    normalized.get("author")
                ),
                "flip": self.normalize_theme(
                    normalized.get("flip")
                ),
                "show": show,
            })

        return (
            clean_users,
            clean_questions,
            errors,
            warnings,
        )

    @staticmethod
    def _normalize_row(
        row: Any,
    ) -> dict[str, Any]:
        return {
            str(key or "").strip().casefold(): value
            for key, value in dict(row).items()
        }

    def parse_request(
        self,
    ) -> tuple[
        str,
        str,
        list[dict[str, Any]],
        list[dict[str, Any]],
        list[str],
        list[str],
    ]:
        upload = request.files.get("file")
        if upload is None or not upload.filename:
            raise ValueError("Файл не выбран.")

        entity = request.form.get("entity", "auto")
        mode = request.form.get("mode", "merge")

        if entity not in self.VALID_ENTITIES:
            raise ValueError(
                "Некорректный раздел импорта."
            )
        if mode not in self.VALID_MODES:
            raise ValueError(
                "Некорректный режим импорта."
            )

        users, questions = self.read_file(
            upload,
            entity,
        )
        (
            clean_users,
            clean_questions,
            errors,
            warnings,
        ) = self.validate(users, questions)

        if entity == "users":
            clean_questions = []
        elif entity == "questions":
            clean_users = []

        return (
            entity,
            mode,
            clean_users,
            clean_questions,
            errors,
            warnings,
        )

    def apply(
        self,
        *,
        entity: str,
        mode: str,
        users_data: list[dict[str, Any]],
        questions_data: list[dict[str, Any]],
    ) -> dict[str, int]:
        if mode == "replace":
            if (
                users_data
                and entity in {
                    "users",
                    "all",
                    "auto",
                }
            ):
                self.db.session.execute(
                    self.db.delete(self.UsersTpv)
                )

            if (
                questions_data
                and entity in {
                    "questions",
                    "all",
                    "auto",
                }
            ):
                self.db.session.execute(
                    self.db.delete(self.QuestionsTpv)
                )

            self.db.session.flush()

        existing_users = self.db.session.scalars(
            self.db.select(self.UsersTpv)
        ).all()
        user_map = {
            self.normalize_text(
                user.username
            ).casefold(): user
            for user in existing_users
        }

        users_added = 0
        users_updated = 0

        for row in users_data:
            key = row["username"].casefold()
            user = user_map.get(key)

            if user is None:
                user = self.UsersTpv()
                user.username = row["username"]
                user.flip_col = 0
                user.approve = "false"

                self.db.session.add(user)
                user_map[key] = user
                users_added += 1
            else:
                users_updated += 1

            user.money = row["money"]
            user.flip = row["flip"]

        existing_questions = self.db.session.scalars(
            self.db.select(self.QuestionsTpv)
        ).all()
        question_keys = {
            (
                self.normalize_text(
                    question.task
                ).casefold(),
                self.normalize_text(
                    question.answer
                ).casefold(),
                self.normalize_text(
                    question.author
                ).casefold(),
                self.theme_key(question.flip),
            )
            for question in existing_questions
        }

        questions_added = 0
        questions_skipped = 0

        for row in questions_data:
            key = (
                row["task"].casefold(),
                row["answer"].casefold(),
                row["author"].casefold(),
                self.theme_key(row["flip"]),
            )

            if mode == "merge" and key in question_keys:
                questions_skipped += 1
                continue

            question = self.QuestionsTpv()
            question.task = row["task"]
            question.answer = row["answer"]
            question.comment = row["comment"]
            question.author = row["author"]
            question.flip = row["flip"]
            question.show = row["show"]

            self.db.session.add(question)
            question_keys.add(key)
            questions_added += 1

        self.db.session.flush()

        for user in self.db.session.scalars(
            self.db.select(self.UsersTpv)
        ).all():
            self.update_approval(user)

        self.history_add(
            "bulk",
            None,
            "import",
            "Выполнен импорт данных",
            details=(
                f"Режим: {mode}; "
                f"пользователей добавлено: {users_added}; "
                f"обновлено: {users_updated}; "
                f"вопросов добавлено: {questions_added}; "
                f"дублей пропущено: {questions_skipped}."
            ),
            can_revert=False,
        )

        self.db.session.commit()

        return {
            "users_added": users_added,
            "users_updated": users_updated,
            "questions_added": questions_added,
            "questions_skipped": questions_skipped,
        }


def register_importing(
    context: EditorContext,
) -> dict[str, Any]:
    """Зарегистрировать прежние API импорта."""
    service = ImportService(context)

    def error(message: str, status: int = 400):
        return jsonify({
            "ok": False,
            "message": message,
        }), status

    def require_access():
        if context.permissions.is_allowed():
            return None
        return error(
            "Нет доступа к редактору.",
            403,
        )

    def tpv_editor_import_preview():
        denied = require_access()
        if denied is not None:
            return denied

        try:
            (
                entity,
                mode,
                users,
                questions,
                errors,
                warnings,
            ) = service.parse_request()
        except (
            ValueError,
            RuntimeError,
            json.JSONDecodeError,
            UnicodeDecodeError,
        ) as exc:
            return error(str(exc))

        if not users and not questions and not errors:
            errors.append(
                "В файле не найдено данных для импорта."
            )

        return jsonify({
            "ok": True,
            "preview": {
                "valid": not errors,
                "entity": entity,
                "mode": mode,
                "users_count": len(users),
                "questions_count": len(questions),
                "errors": errors,
                "warnings": warnings,
            },
        })

    def tpv_editor_import_apply():
        denied = require_access()
        if denied is not None:
            return denied

        try:
            (
                entity,
                mode,
                users_data,
                questions_data,
                errors,
                _warnings,
            ) = service.parse_request()

            if errors:
                return error(
                    "Импорт отменён: "
                    + "; ".join(errors)
                )

            result = service.apply(
                entity=entity,
                mode=mode,
                users_data=users_data,
                questions_data=questions_data,
            )
        except (
            ValueError,
            RuntimeError,
            json.JSONDecodeError,
            UnicodeDecodeError,
        ) as exc:
            service.db.session.rollback()
            return error(str(exc))
        except Exception:
            service.db.session.rollback()
            raise

        return jsonify({
            "ok": True,
            "message": (
                "Импорт завершён. Пользователи: "
                f"добавлено {result['users_added']}, "
                f"обновлено {result['users_updated']}. "
                "Вопросы: "
                f"добавлено {result['questions_added']}, "
                "пропущено дублей "
                f"{result['questions_skipped']}."
            ),
        })

    rules = (
        (
            "/tpv_editor/api/import-preview",
            "tpv_editor_import_preview",
            tpv_editor_import_preview,
        ),
        (
            "/tpv_editor/api/import",
            "tpv_editor_import_apply",
            tpv_editor_import_apply,
        ),
    )

    for rule, endpoint, view_func in rules:
        context.app.add_url_rule(
            rule,
            endpoint=endpoint,
            view_func=view_func,
            methods=["POST"],
        )

    return {
        "service": service,
        "tpv_editor_read_import_file": (
            service.read_file
        ),
        "tpv_editor_validate_import": (
            service.validate
        ),
        "tpv_editor_parse_import_request": (
            service.parse_request
        ),
        "tpv_editor_import_preview": (
            tpv_editor_import_preview
        ),
        "tpv_editor_import_apply": (
            tpv_editor_import_apply
        ),
    }


__all__ = [
    "ImportService",
    "register_importing",
]
