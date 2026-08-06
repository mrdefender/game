"""Экспорт данных TPV Editor — этап 13.6.4."""

from __future__ import annotations

import csv
from datetime import datetime
import io
import json
import os
import sqlite3
import tempfile
from typing import Any

from flask import jsonify, send_file

from .registry import EditorContext


USERS_FIELDS = (
    "id",
    "username",
    "money",
    "flip",
    "flip_col",
    "approve",
)

QUESTIONS_FIELDS = (
    "id",
    "task",
    "answer",
    "comment",
    "author",
    "flip",
    "show",
)


class ExportService:
    """Подготовка CSV, JSON, XLSX и резервной копии SQLite."""

    def __init__(self, context: EditorContext) -> None:
        self.context = context
        self.db = context.db

        self.UsersTpv = self._dependency("UsersTpv")
        self.QuestionsTpv = self._dependency("Questions_tpv")

        self.is_general_theme = self._dependency(
            "tpv_editor_is_general_theme"
        )
        self.normalize_text = self._dependency(
            "tpv_editor_normalize_text"
        )

    def _dependency(self, name: str) -> Any:
        value = self.context.get(name)
        if value is None:
            raise RuntimeError(
                "Экспорт TPV Editor: "
                f"отсутствует зависимость {name}"
            )
        return value

    def users_rows(self) -> list[dict[str, Any]]:
        users = self.db.session.scalars(
            self.db.select(self.UsersTpv)
            .order_by(self.UsersTpv.id)
        ).all()

        return [
            {
                "id": user.id,
                "username": user.username or "",
                "money": int(user.money or 0),
                "flip": (
                    "false"
                    if self.is_general_theme(user.flip)
                    else self.normalize_text(user.flip)
                ),
                "flip_col": int(user.flip_col or 0),
                "approve": (
                    "true"
                    if str(user.approve or "").casefold() == "true"
                    else "false"
                ),
            }
            for user in users
        ]

    def questions_rows(self) -> list[dict[str, Any]]:
        questions = self.db.session.scalars(
            self.db.select(self.QuestionsTpv)
            .order_by(self.QuestionsTpv.id)
        ).all()

        return [
            {
                "id": question.id,
                "task": question.task or "",
                "answer": question.answer or "",
                "comment": question.comment or "",
                "author": question.author or "",
                "flip": (
                    "false"
                    if self.is_general_theme(question.flip)
                    else self.normalize_text(question.flip)
                ),
                "show": (
                    "true"
                    if str(question.show or "").casefold() == "true"
                    else "false"
                ),
            }
            for question in questions
        ]

    @staticmethod
    def download_bytes(
        content: bytes,
        filename: str,
        mimetype: str,
    ):
        buffer = io.BytesIO(content)
        buffer.seek(0)
        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype=mimetype,
        )

    @staticmethod
    def csv_bytes(
        rows: list[dict[str, Any]],
        fields: tuple[str, ...],
    ) -> bytes:
        stream = io.StringIO(newline="")
        writer = csv.DictWriter(
            stream,
            fieldnames=fields,
            extrasaction="ignore",
        )
        writer.writeheader()
        writer.writerows(rows)

        # BOM сохраняет корректную кириллицу в Excel.
        return ("\ufeff" + stream.getvalue()).encode("utf-8")

    @staticmethod
    def xlsx_bytes(
        *,
        users: list[dict[str, Any]] | None = None,
        questions: list[dict[str, Any]] | None = None,
    ) -> bytes:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError as exc:
            raise RuntimeError(
                "Для Excel установите пакет openpyxl: "
                "pip install openpyxl"
            ) from exc

        workbook = Workbook()
        workbook.remove(workbook.active)

        header_fill = PatternFill(
            "solid",
            fgColor="12354A",
        )
        header_font = Font(
            color="FFFFFF",
            bold=True,
        )

        def add_sheet(
            name: str,
            rows: list[dict[str, Any]],
            fields: tuple[str, ...],
        ) -> None:
            sheet = workbook.create_sheet(name)
            sheet.append(list(fields))

            for row in rows:
                sheet.append([
                    row.get(field, "")
                    for field in fields
                ])

            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(
                    horizontal="center"
                )

            sheet.freeze_panes = "A2"

            for column in sheet.columns:
                width = min(
                    55,
                    max(
                        12,
                        max(
                            len(str(cell.value or ""))
                            for cell in column
                        )
                        + 2,
                    ),
                )
                sheet.column_dimensions[
                    column[0].column_letter
                ].width = width

        if users is not None:
            add_sheet(
                "Users",
                users,
                USERS_FIELDS,
            )

        if questions is not None:
            add_sheet(
                "Questions",
                questions,
                QUESTIONS_FIELDS,
            )

        output = io.BytesIO()
        workbook.save(output)
        return output.getvalue()

    def users_export(self, fmt: str):
        rows = self.users_rows()
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        if fmt == "csv":
            return self.download_bytes(
                self.csv_bytes(rows, USERS_FIELDS),
                f"tpv_users_{stamp}.csv",
                "text/csv; charset=utf-8",
            )

        if fmt == "json":
            data = json.dumps(
                {"users": rows},
                ensure_ascii=False,
                indent=2,
            ).encode("utf-8")
            return self.download_bytes(
                data,
                f"tpv_users_{stamp}.json",
                "application/json",
            )

        if fmt == "xlsx":
            data = self.xlsx_bytes(users=rows)
            return self.download_bytes(
                data,
                f"tpv_users_{stamp}.xlsx",
                (
                    "application/vnd.openxmlformats-officedocument."
                    "spreadsheetml.sheet"
                ),
            )

        raise LookupError("Неизвестный формат.")

    def questions_export(self, fmt: str):
        rows = self.questions_rows()
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        if fmt == "csv":
            return self.download_bytes(
                self.csv_bytes(rows, QUESTIONS_FIELDS),
                f"tpv_questions_{stamp}.csv",
                "text/csv; charset=utf-8",
            )

        if fmt == "json":
            data = json.dumps(
                {"questions": rows},
                ensure_ascii=False,
                indent=2,
            ).encode("utf-8")
            return self.download_bytes(
                data,
                f"tpv_questions_{stamp}.json",
                "application/json",
            )

        if fmt == "xlsx":
            data = self.xlsx_bytes(questions=rows)
            return self.download_bytes(
                data,
                f"tpv_questions_{stamp}.xlsx",
                (
                    "application/vnd.openxmlformats-officedocument."
                    "spreadsheetml.sheet"
                ),
            )

        raise LookupError("Неизвестный формат.")

    def full_export(self, fmt: str):
        users = self.users_rows()
        questions = self.questions_rows()
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        if fmt == "json":
            payload = {
                "format": "tpv-editor-backup",
                "version": 1,
                "created_at": datetime.now().isoformat(
                    timespec="seconds"
                ),
                "users": users,
                "questions": questions,
            }
            return self.download_bytes(
                json.dumps(
                    payload,
                    ensure_ascii=False,
                    indent=2,
                ).encode("utf-8"),
                f"tpv_backup_{stamp}.json",
                "application/json",
            )

        if fmt == "xlsx":
            data = self.xlsx_bytes(
                users=users,
                questions=questions,
            )
            return self.download_bytes(
                data,
                f"tpv_backup_{stamp}.xlsx",
                (
                    "application/vnd.openxmlformats-officedocument."
                    "spreadsheetml.sheet"
                ),
            )

        raise LookupError("Неизвестный формат.")

    def database_export(self):
        database_path = self.db.engine.url.database
        if not database_path:
            raise RuntimeError(
                "Не удалось определить файл SQLite."
            )

        source_path = os.path.abspath(database_path)
        if not os.path.exists(source_path):
            raise FileNotFoundError("Файл SQLite не найден.")

        temporary = tempfile.NamedTemporaryFile(
            prefix="tpv_backup_",
            suffix=".sqlite",
            delete=False,
        )
        temporary.close()

        try:
            with sqlite3.connect(source_path) as source:
                with sqlite3.connect(temporary.name) as target:
                    source.backup(target)

            stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            return send_file(
                temporary.name,
                as_attachment=True,
                download_name=(
                    f"tpv_database_{stamp}.sqlite"
                ),
                mimetype="application/vnd.sqlite3",
            )
        except Exception:
            if os.path.exists(temporary.name):
                os.unlink(temporary.name)
            raise


def register_exporting(
    context: EditorContext,
) -> dict[str, Any]:
    """Зарегистрировать прежние маршруты экспорта."""
    service = ExportService(context)

    def error(message: str, status: int = 400):
        return jsonify({
            "ok": False,
            "message": message,
        }), status

    def require_access():
        if context.permissions.is_allowed():
            return None
        return error("Нет доступа к редактору.", 403)

    def tpv_editor_export_users(fmt: str):
        denied = require_access()
        if denied is not None:
            return denied

        try:
            return service.users_export(fmt)
        except RuntimeError as exc:
            return error(str(exc), 500)
        except LookupError as exc:
            return error(str(exc), 404)

    def tpv_editor_export_questions(fmt: str):
        denied = require_access()
        if denied is not None:
            return denied

        try:
            return service.questions_export(fmt)
        except RuntimeError as exc:
            return error(str(exc), 500)
        except LookupError as exc:
            return error(str(exc), 404)

    def tpv_editor_export_full(fmt: str):
        denied = require_access()
        if denied is not None:
            return denied

        try:
            return service.full_export(fmt)
        except RuntimeError as exc:
            return error(str(exc), 500)
        except LookupError as exc:
            return error(str(exc), 404)

    def tpv_editor_export_database():
        denied = require_access()
        if denied is not None:
            return denied

        try:
            return service.database_export()
        except FileNotFoundError as exc:
            return error(str(exc), 404)
        except RuntimeError as exc:
            return error(str(exc), 500)

    rules = (
        (
            "/tpv_editor/export/users.<fmt>",
            "tpv_editor_export_users",
            tpv_editor_export_users,
        ),
        (
            "/tpv_editor/export/questions.<fmt>",
            "tpv_editor_export_questions",
            tpv_editor_export_questions,
        ),
        (
            "/tpv_editor/export/full.<fmt>",
            "tpv_editor_export_full",
            tpv_editor_export_full,
        ),
        (
            "/tpv_editor/export/database.sqlite",
            "tpv_editor_export_database",
            tpv_editor_export_database,
        ),
    )

    for rule, endpoint, view_func in rules:
        context.app.add_url_rule(
            rule,
            endpoint=endpoint,
            view_func=view_func,
            methods=["GET"],
        )

    return {
        "service": service,
        "tpv_editor_export_users_rows": service.users_rows,
        "tpv_editor_export_questions_rows": (
            service.questions_rows
        ),
        "tpv_editor_download_bytes": service.download_bytes,
        "tpv_editor_csv_bytes": service.csv_bytes,
        "tpv_editor_xlsx_bytes": service.xlsx_bytes,
        "tpv_editor_export_users": tpv_editor_export_users,
        "tpv_editor_export_questions": (
            tpv_editor_export_questions
        ),
        "tpv_editor_export_full": tpv_editor_export_full,
        "tpv_editor_export_database": (
            tpv_editor_export_database
        ),
    }


__all__ = [
    "ExportService",
    "register_exporting",
]
